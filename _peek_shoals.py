from openpyxl import load_workbook
from pathlib import Path


def peek(path: str, maxr: int = 12, maxc: int = 16) -> None:
    p = Path(path)
    wb = load_workbook(p, read_only=True, data_only=True)
    print("===", p.name, "sheets:", wb.sheetnames)
    for sn in wb.sheetnames[:6]:
        ws = wb[sn]
        print("---", sn, "max_row", ws.max_row, "max_col", ws.max_column)
        for r in range(1, min(maxr, ws.max_row + 1)):
            row = []
            for c in range(1, min(maxc, ws.max_column + 1)):
                v = ws.cell(r, c).value
                s = "" if v is None else str(v).replace("\n", " ")[:45]
                row.append(s)
            if any(x.strip() for x in row):
                print(r, row)
        print()
    wb.close()


if __name__ == "__main__":
    base = Path(r"c:\Users\sheld\Documents\Work Stuff\Payroll\WE 03.22.26")
    peek(str(base / "Shoals" / "Shoals Invoice Audit WE 03.22.26.xlsx"))
    peek(str(base / "Imports" / "Shoals" / "Shoals CORE Timecards (1).xlsx"), maxr=20, maxc=14)
    peek(str(base / "Imports" / "Shoals" / "Shoals WED Timecards.xlsx"), maxr=20, maxc=14)
