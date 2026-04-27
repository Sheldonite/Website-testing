from __future__ import annotations

import base64
import io
import json
import os
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from types import SimpleNamespace
from urllib.parse import quote

from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.middleware.httpsredirect import HTTPSRedirectMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import Select, delete, desc, func, select
from sqlalchemy.exc import IntegrityError
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.sessions import SessionMiddleware
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import ensure_schema, session_scope
from .importers import (
    markup_1663_expected_total,
    parse_attendance_csv,
    parse_invoice_workbook,
    parse_ppe_csv,
    parse_roster_csv,
    parse_shoals_timecard_workbook,
    pay_type_uses_1663_markup,
)
from .models import (
    AdminProfile,
    AppUser,
    Attendance,
    EmploymentAssignmentArchive,
    ImportBatch,
    InvoiceCrossrefRowNote,
    InvoiceCrossrefSnapshot,
    InvoiceLine,
    InvoicePayTypeMapping,
    InvoiceSetting,
    InvoiceSettingsPreset,
    PpeDeductionLine,
    Person,
    PersonOnboardingDocument,
    PlannerEntry,
    Project,
    ProjectDepartment,
    ProjectParticipant,
    ProjectTask,
    ProjectTaskNote,
    ProjectTaskSupporter,
)
from .security import (
    get_security_settings,
    hash_password,
    normalize_username,
    safe_next_path,
    validate_new_password,
    verify_password,
)


APP_DIR = Path(__file__).resolve().parent
ONBOARDING_UPLOAD_DIR = APP_DIR / "uploads" / "onboarding"
MAX_ONBOARDING_UPLOAD_BYTES = 10 * 1024 * 1024
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))

# Invoice import batch `note` values (ImportBatch.note)
INVOICE_NOTE_PAID_BILLED = "invoice_paid_billed"
INVOICE_NOTE_MIM = "invoice_mim"
INVOICE_NOTE_PPE = "invoice_ppe"
INVOICE_NOTE_SHOALS_CORE = "invoice_shoals_core"
INVOICE_NOTE_SHOALS_WEEKEND = "invoice_shoals_weekend"
_INVOICE_IMPORT_NOTES_ALL = (
    INVOICE_NOTE_PAID_BILLED,
    INVOICE_NOTE_MIM,
    INVOICE_NOTE_PPE,
    INVOICE_NOTE_SHOALS_CORE,
    INVOICE_NOTE_SHOALS_WEEKEND,
)

# Persisted UI choice for import/export layout (dropdown). Inference from imports is only a fallback.
INVOICE_ACTIVE_LAYOUT_KEY = "invoice_active_import_layout"
_INVOICE_LAYOUT_VALUES = frozenset({"default", "material_motion", "shoals"})


def _latest_import_batch_by_note(s, note: str) -> ImportBatch | None:
    return (
        s.execute(select(ImportBatch).where(ImportBatch.note == note).order_by(desc(ImportBatch.imported_at)).limit(1))
        .scalars()
        .first()
    )


def _comparison_mim_line_batch_ids(s) -> list[int]:
    """Payroll-side batches for cross-reference: latest MIM, or Shoals core + weekend if no MIM."""
    mb = _latest_import_batch_by_note(s, INVOICE_NOTE_MIM)
    if mb:
        return [mb.id]
    ids: list[int] = []
    for note in (INVOICE_NOTE_SHOALS_CORE, INVOICE_NOTE_SHOALS_WEEKEND):
        b = _latest_import_batch_by_note(s, note)
        if b:
            ids.append(b.id)
    return ids


def _material_motion_only_batch_ids(s) -> list[int]:
    """Latest MIM payroll batch only (Material and Motion export; never Shoals batches)."""
    mb = _latest_import_batch_by_note(s, INVOICE_NOTE_MIM)
    return [mb.id] if mb else []


def _shoals_only_batch_ids(s) -> list[int]:
    """Shoals core + weekend payroll batches (no MIM batch)."""
    ids: list[int] = []
    for note in (INVOICE_NOTE_SHOALS_CORE, INVOICE_NOTE_SHOALS_WEEKEND):
        b = _latest_import_batch_by_note(s, note)
        if b:
            ids.append(b.id)
    return ids


def _crossref_line_batches_and_layout(s) -> tuple[list[int], str]:
    """
    Payroll-side import batches and comparison key mode for the active layout.
    Shoals layout uses core+weekend only and week-aware keys; Material in Motion uses MIM only.
    Default layout keeps legacy batch selection (MIM if present, else Shoals) and infers key mode.
    """
    layout = _get_active_invoice_layout(s)
    if layout == "shoals":
        return _shoals_only_batch_ids(s), "shoals"
    if layout == "material_motion":
        return _material_motion_only_batch_ids(s), "material_motion"
    mim_ids = _comparison_mim_line_batch_ids(s)
    if not _latest_import_batch_by_note(s, INVOICE_NOTE_MIM) and mim_ids:
        return mim_ids, "shoals"
    return mim_ids, "material_motion"


def _infer_invoice_layout_from_imports(s) -> str:
    """Fallback when no persisted layout: infer from which import batches exist."""
    if _latest_import_batch_by_note(s, INVOICE_NOTE_MIM) or _latest_import_batch_by_note(s, INVOICE_NOTE_PPE):
        return "material_motion"
    if _latest_import_batch_by_note(s, INVOICE_NOTE_SHOALS_CORE) or _latest_import_batch_by_note(s, INVOICE_NOTE_SHOALS_WEEKEND):
        return "shoals"
    return "default"


def _get_active_invoice_layout(s) -> str:
    """Persisted layout from settings, or infer from imports if unset."""
    st = s.execute(select(InvoiceSetting).where(InvoiceSetting.key == INVOICE_ACTIVE_LAYOUT_KEY)).scalar_one_or_none()
    if st and (st.value or "").strip() in _INVOICE_LAYOUT_VALUES:
        return (st.value or "").strip()
    return _infer_invoice_layout_from_imports(s)


def _set_active_invoice_layout(s, layout: str) -> None:
    if layout not in _INVOICE_LAYOUT_VALUES:
        return
    st = s.execute(select(InvoiceSetting).where(InvoiceSetting.key == INVOICE_ACTIVE_LAYOUT_KEY)).scalar_one_or_none()
    if st is None:
        s.add(InvoiceSetting(key=INVOICE_ACTIVE_LAYOUT_KEY, value=layout))
    else:
        st.value = layout


def _invoice_layout_switch_delete_notes(target_layout: str) -> list[str]:
    """Batch notes to delete when switching to `target_layout` (paid/billed is always kept)."""
    if target_layout == "default":
        return [INVOICE_NOTE_MIM, INVOICE_NOTE_PPE, INVOICE_NOTE_SHOALS_CORE, INVOICE_NOTE_SHOALS_WEEKEND]
    if target_layout == "material_motion":
        return [INVOICE_NOTE_SHOALS_CORE, INVOICE_NOTE_SHOALS_WEEKEND]
    if target_layout == "shoals":
        return [INVOICE_NOTE_MIM, INVOICE_NOTE_PPE]
    raise HTTPException(status_code=400, detail="Invalid target layout.")


# Buckets for person profile "Total Occurrences for Attendance" (CSV statuses are stored lowercase).
_OCCURRENCE_ORDER: tuple[tuple[str, str], ...] = (
    ("ncns", "No Call No Shows"),
    ("call_out", "Call Outs"),
    ("early_out", "Early Outs"),
    ("late_arrival", "Late Arrivals"),
    ("missed_punch", "Missed Punches"),
    ("long_lunch", "Long Lunches"),
)


def _occurrence_kind(status: str | None) -> str | None:
    """Map an attendance status string to an occurrence bucket key, or None if not one of the tracked types."""
    raw = (status or "").strip().lower()
    if not raw:
        return None
    s = " ".join(raw.replace("_", " ").replace("-", " ").split())
    compact = "".join(ch for ch in s if ch.isalnum())
    if "nocallnoshow" in compact or s == "ncns":
        return "ncns"
    if "callout" in compact or "call out" in s:
        return "call_out"
    if "earlyout" in compact or "early out" in s:
        return "early_out"
    if "latearrival" in compact or "late arrival" in s:
        return "late_arrival"
    if "missedpunch" in compact or "missed punch" in s:
        return "missed_punch"
    if "longlunch" in compact or "long lunch" in s:
        return "long_lunch"
    return None


def _attendance_occurrence_totals(attendance_list: list) -> tuple[int, list[tuple[str, int]]]:
    counts = {key: 0 for key, _ in _OCCURRENCE_ORDER}
    for row in attendance_list:
        k = _occurrence_kind(getattr(row, "status", None))
        if k in counts:
            counts[k] += 1
    breakdown = [(label, counts[key]) for key, label in _OCCURRENCE_ORDER]
    total = sum(counts.values())
    return total, breakdown


def _employment_assignment_calendar_days(
    hire: date | None,
    end: date | None,
    *,
    today: date,
) -> int | None:
    """Inclusive calendar days on assignment: hire → end date, or hire → today if no end date. None if no hire date."""
    if hire is None:
        return None
    if end is not None:
        if end < hire:
            return None
        return (end - hire).days + 1
    if today < hire:
        return None
    return (today - hire).days + 1


def _prior_completed_week_monday_sunday(today: date) -> tuple[date, date]:
    """Monday through Sunday of the calendar week before the week containing `today`."""
    monday_this_week = today - timedelta(days=today.weekday())
    monday_last_week = monday_this_week - timedelta(days=7)
    sunday_last_week = monday_last_week + timedelta(days=6)
    return monday_last_week, sunday_last_week


_WEEKDAY_NAMES: tuple[str, ...] = (
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
)


def _last_week_punch_rows(attendance_list: list, today: date | None = None) -> tuple[str, list[dict[str, object]]]:
    """One row per day Mon–Sun for the prior week; status is the stored daily record (no clock times in DB)."""
    ref = today or date.today()
    mon, sun = _prior_completed_week_monday_sunday(ref)
    range_label = f"{mon.strftime('%b %d, %Y')} – {sun.strftime('%b %d, %Y')}"
    by_day = {getattr(r, "attended_on"): r for r in attendance_list}
    rows: list[dict[str, object]] = []
    for i in range(7):
        d = mon + timedelta(days=i)
        r = by_day.get(d)
        rows.append(
            {
                "weekday": _WEEKDAY_NAMES[d.weekday()],
                "date": d,
                "status": (getattr(r, "status", None) or None) if r else None,
                "source_filename": (getattr(r, "source_filename", None) or None) if r else None,
            }
        )
    return range_label, rows


_ALLOWED_TALENT_STATUSES = frozenset({"active", "work_comp", "fmla", "terminated"})
_TALENT_STATUS_OPTIONS: tuple[tuple[str, str], ...] = (
    ("active", "Active"),
    ("work_comp", "Work Comp"),
    ("fmla", "FMLA"),
    ("terminated", "Terminated"),
)

_TERMINATION_END_REASONS: tuple[tuple[str, str], ...] = (
    ("job_abandonment", "Job Abandonment"),
    ("attendance", "Attendance"),
    ("resignation_other_job", "Resignation because they found another job"),
    ("resignation_dislike_job", "Resignation because they didn't like the job"),
    ("resignation_transportation", "Resignation because of transportation issues"),
    ("misconduct_policy", "Misconduct Policy Violation"),
)
_ALLOWED_TERMINATION_END_REASONS = frozenset(k for k, _ in _TERMINATION_END_REASONS)

_PROJECT_STATUS_OPTIONS: tuple[tuple[str, str], ...] = (
    ("planning", "Planning"),
    ("active", "Active"),
    ("on_hold", "On Hold"),
    ("completed", "Completed"),
)
_ALLOWED_PROJECT_STATUSES = frozenset(value for value, _ in _PROJECT_STATUS_OPTIONS)

_PROJECT_TASK_STATUS_OPTIONS: tuple[tuple[str, str], ...] = (
    ("backlog", "Backlog"),
    ("todo", "To Do"),
    ("in_progress", "In Progress"),
    ("blocked", "Stuck"),
    ("done", "Done"),
)
_ALLOWED_PROJECT_TASK_STATUSES = frozenset(value for value, _ in _PROJECT_TASK_STATUS_OPTIONS)

_PROJECT_TASK_PRIORITY_OPTIONS: tuple[tuple[str, str], ...] = (
    ("critical", "Critical"),
    ("high", "High"),
    ("medium", "Medium"),
    ("low", "Low"),
)
_ALLOWED_PROJECT_TASK_PRIORITIES = frozenset(value for value, _ in _PROJECT_TASK_PRIORITY_OPTIONS)

_PROJECT_TASK_RISK_OPTIONS: tuple[tuple[str, str], ...] = (
    ("low", "Low"),
    ("medium", "Medium"),
    ("high", "High"),
)
_ALLOWED_PROJECT_TASK_RISKS = frozenset(value for value, _ in _PROJECT_TASK_RISK_OPTIONS)

_PROJECT_TASK_NOTE_TYPE_OPTIONS: tuple[tuple[str, str], ...] = (
    ("general", "General"),
    ("decision", "Decision"),
    ("follow_up", "Follow-up"),
    ("risk", "Risk"),
    ("meeting", "Meeting"),
)
_ALLOWED_PROJECT_TASK_NOTE_TYPES = frozenset(value for value, _ in _PROJECT_TASK_NOTE_TYPE_OPTIONS)

_PROJECT_DEPARTMENT_COLORS: tuple[str, ...] = (
    "slate",
    "blue",
    "emerald",
    "amber",
    "rose",
    "cyan",
    "violet",
    "orange",
)
_DEFAULT_PROJECT_DEPARTMENTS: tuple[tuple[str, str], ...] = (
    ("Operations", "blue"),
    ("HR", "emerald"),
    ("IT", "violet"),
    ("Finance", "amber"),
    ("Training", "cyan"),
)


def _termination_reason_label(key: str | None) -> str:
    if not key:
        return ""
    for k, label in _TERMINATION_END_REASONS:
        if k == key:
            return label
    return key


def _quantize_hundredth(d: Decimal) -> Decimal:
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _invoice_num_display(value: object) -> str:
    """Round to nearest hundredth for cross-reference table display."""
    if value is None:
        return ""
    try:
        q = _quantize_hundredth(_as_decimal(value))
    except Exception:
        return str(value)
    return format(q, "f")


def _invoice_money_display(value: object) -> str:
    """Round to nearest hundredth and format as USD for cross-reference table display."""
    if value is None:
        return ""
    try:
        q = _quantize_hundredth(_as_decimal(value))
    except Exception:
        return str(value)
    return f"${format(q, ',.2f')}"


templates.env.filters["invoice_num"] = _invoice_num_display
templates.env.filters["invoice_money"] = _invoice_money_display

SECURITY_SETTINGS = get_security_settings()
_AUTH_EXEMPT_PATHS = frozenset({"/login", "/setup", "/favicon.ico"})
_AUTH_EXEMPT_PREFIXES = ("/static/",)
_UNSAFE_HTTP_METHODS = frozenset({"POST", "PUT", "PATCH", "DELETE"})
_SAFE_HTTP_METHODS = frozenset({"GET", "HEAD", "OPTIONS"})
_LOCAL_AUTH_BYPASS_ENV = "ATTENDANCE_TRACKER_AUTH_BYPASS"
_LOCAL_ADMIN_RESET_ENV = "ATTENDANCE_TRACKER_ALLOW_ADMIN_RESET"

_PERMISSION_GROUPS: tuple[tuple[str, tuple[tuple[str, str], ...]], ...] = (
    (
        "Accounts",
        (
            ("users_manage", "Create and edit user accounts"),
            ("impersonate_users", "Impersonate another user"),
        ),
    ),
    (
        "Projects",
        (
            ("projects_view", "View assigned projects"),
            ("projects_manage", "Create and edit projects"),
            ("projects_view_all", "View all projects (project manager)"),
        ),
    ),
    (
        "People",
        (
            ("people_view", "View people and attendance"),
            ("people_manage", "Edit employment and roster data"),
            ("people_sensitive_view", "View personal, emergency, and onboarding data"),
            ("people_sensitive_manage", "Edit personal, emergency, and onboarding data"),
        ),
    ),
    (
        "Imports",
        (
            ("imports_view", "View import screens and history"),
            ("imports_manage", "Run attendance and roster imports"),
        ),
    ),
    (
        "Invoices",
        (
            ("invoices_view", "View invoice crossreferencing"),
            ("invoices_manage", "Change invoice settings and imports"),
        ),
    ),
)
_ALL_PERMISSIONS = frozenset(key for _, group in _PERMISSION_GROUPS for key, _ in group)

app = FastAPI(title="Project: ATS")
if SECURITY_SETTINGS.allowed_hosts != ("*",):
    app.add_middleware(TrustedHostMiddleware, allowed_hosts=list(SECURITY_SETTINGS.allowed_hosts))
if SECURITY_SETTINGS.force_https:
    app.add_middleware(HTTPSRedirectMiddleware)
app.mount("/static", StaticFiles(directory=str(APP_DIR / "static")), name="static")


@app.on_event("startup")
def _startup() -> None:
    ensure_schema()
    ONBOARDING_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with session_scope() as s:
        if s.get(AdminProfile, 1) is None:
            s.add(AdminProfile(id=1))
            s.commit()
    _migrate_sensitive_people_data()
    _migrate_app_user_access()


def _redirect(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=303)


def _admin_profile_view(s) -> AdminProfile:
    p = s.get(AdminProfile, 1)
    if not p:
        p = AdminProfile(id=1)
        s.add(p)
        s.commit()
        s.refresh(p)
    return p


def _setup_available(s) -> bool:
    return not bool(s.execute(select(func.count(AppUser.id))).scalar_one())


def _same_origin_request(request: Request) -> bool:
    origin = (request.headers.get("origin") or "").strip()
    referer = (request.headers.get("referer") or "").strip()
    expected = f"{request.url.scheme}://{request.url.netloc}"
    if origin:
        return origin.rstrip("/") == expected
    if referer:
        return referer.startswith(expected + "/") or referer == expected
    return False


def _migrate_sensitive_people_data() -> None:
    with session_scope() as s:
        people = s.execute(select(Person)).scalars().all()
        changed = False
        for person in people:
            if person.legacy_date_of_birth is not None:
                person.date_of_birth = person.legacy_date_of_birth
                changed = True
            if person.legacy_social_security_number:
                person.social_security_number = person.legacy_social_security_number
                changed = True
            if person.legacy_direct_deposit_json:
                person.direct_deposit_json = person.legacy_direct_deposit_json
                changed = True
            if person.date_of_birth_encrypted and person.legacy_date_of_birth is not None:
                person.legacy_date_of_birth = None
                changed = True
            if person.social_security_number_encrypted and person.legacy_social_security_number:
                person.legacy_social_security_number = None
                changed = True
            if person.direct_deposit_json_encrypted and person.legacy_direct_deposit_json:
                person.legacy_direct_deposit_json = None
                changed = True
        if changed:
            s.commit()


def _permission_label(permission_key: str) -> str:
    for _, group in _PERMISSION_GROUPS:
        for key, label in group:
            if key == permission_key:
                return label
    return permission_key


def _normalize_permissions(values: set[str]) -> set[str]:
    perms = {value for value in values if value in _ALL_PERMISSIONS}
    if "projects_manage" in perms:
        perms.add("projects_view")
    if "people_manage" in perms:
        perms.add("people_view")
    if "people_sensitive_view" in perms:
        perms.add("people_view")
    if "people_sensitive_manage" in perms:
        perms.update({"people_sensitive_view", "people_view"})
    if "imports_manage" in perms:
        perms.add("imports_view")
    if "invoices_manage" in perms:
        perms.add("invoices_view")
    return perms


def _parse_permissions_json(raw: str | None) -> set[str]:
    if not raw or not raw.strip():
        return set()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return set()
    if not isinstance(data, list):
        return set()
    return _normalize_permissions({str(item).strip() for item in data if str(item).strip()})


def _permissions_json_from_form(raw_values: list[str]) -> str:
    permissions = sorted(_normalize_permissions(set(raw_values)))
    return json.dumps(permissions)


def _permissions_for_user(user: AppUser | SimpleNamespace | None) -> set[str]:
    if user is None or not bool(getattr(user, "is_active", True)):
        return set()
    permissions_attr = getattr(user, "permissions", None)
    if permissions_attr is not None:
        if bool(getattr(user, "is_superuser", False)):
            return set(_ALL_PERMISSIONS)
        return _normalize_permissions(set(permissions_attr))
    if getattr(user, "is_superuser", False):
        return set(_ALL_PERMISSIONS)
    return _parse_permissions_json(getattr(user, "permissions_json", None))


def _user_has_permission(user: AppUser | SimpleNamespace | None, permission_key: str) -> bool:
    return permission_key in _permissions_for_user(user)


def _user_display_name(user: AppUser | None) -> str:
    if user is None:
        return ""
    return (getattr(user, "full_name", None) or "").strip() or user.username


def _user_role_display(user: AppUser | None) -> str:
    if user is None:
        return ""
    if getattr(user, "is_superuser", False):
        return "Super Admin"
    return (getattr(user, "role_name", None) or "").strip() or "Custom access"


def _user_namespace(user: AppUser | None) -> SimpleNamespace | None:
    if user is None:
        return None
    permissions = _permissions_for_user(user)
    is_superuser = bool(user.is_superuser)
    return SimpleNamespace(
        id=user.id,
        username=user.username,
        full_name=(user.full_name or "").strip(),
        display_name=_user_display_name(user),
        role_name=(user.role_name or "").strip(),
        role_display=_user_role_display(user),
        is_superuser=is_superuser,
        is_active=bool(user.is_active),
        permissions=permissions,
        can_manage_users="users_manage" in permissions or is_superuser,
        can_impersonate="impersonate_users" in permissions or is_superuser,
        can_view_projects="projects_view" in permissions or is_superuser,
        can_manage_projects="projects_manage" in permissions or is_superuser,
        can_view_all_projects="projects_view_all" in permissions or is_superuser,
        project_manager_person_id=user.project_manager_person_id,
        person_id=user.person_id,
        can_view_people="people_view" in permissions or is_superuser,
        can_manage_people="people_manage" in permissions or is_superuser,
        can_view_sensitive_people="people_sensitive_view" in permissions or is_superuser,
        can_manage_sensitive_people="people_sensitive_manage" in permissions or is_superuser,
        can_view_imports="imports_view" in permissions or is_superuser,
        can_manage_imports="imports_manage" in permissions or is_superuser,
        can_view_invoices="invoices_view" in permissions or is_superuser,
        can_manage_invoices="invoices_manage" in permissions or is_superuser,
        last_login_at=user.last_login_at,
    )


def _local_auth_bypass_enabled(request: Request) -> bool:
    if os.environ.get(_LOCAL_AUTH_BYPASS_ENV) != "1":
        return False
    client_host = request.client.host if request.client else ""
    host_header = request.headers.get("host", "")
    return client_host in {"127.0.0.1", "::1", "localhost"} and host_header.startswith(
        ("127.0.0.1", "localhost", "[::1]")
    )


def _local_admin_reset_enabled(request: Request) -> bool:
    if os.environ.get(_LOCAL_ADMIN_RESET_ENV) != "1":
        return False
    client_host = request.client.host if request.client else ""
    host_header = request.headers.get("host", "")
    return client_host in {"127.0.0.1", "::1", "localhost"} and host_header.startswith(
        ("127.0.0.1", "localhost", "[::1]")
    )


def _local_preview_user_namespace() -> SimpleNamespace:
    permissions = set(_ALL_PERMISSIONS)
    return SimpleNamespace(
        id=0,
        username="local-preview",
        full_name="Local Preview",
        display_name="Local Preview Admin",
        role_name="Local Preview",
        role_display="Local Preview Admin",
        is_superuser=True,
        is_active=True,
        permissions=permissions,
        can_manage_users=True,
        can_impersonate=True,
        can_view_projects=True,
        can_manage_projects=True,
        can_view_all_projects=True,
        project_manager_person_id=None,
        person_id=None,
        can_view_people=True,
        can_manage_people=True,
        can_view_sensitive_people=True,
        can_manage_sensitive_people=True,
        can_view_imports=True,
        can_manage_imports=True,
        can_view_invoices=True,
        can_manage_invoices=True,
        last_login_at=None,
    )


def _assign_default_access(user: AppUser, *, make_superuser: bool = False) -> None:
    if make_superuser:
        user.is_superuser = True
        user.role_name = (user.role_name or "").strip() or "Super Admin"
        user.permissions_json = json.dumps(sorted(_ALL_PERMISSIONS))
        return
    user.is_superuser = False
    permissions = _normalize_permissions(_parse_permissions_json(user.permissions_json))
    user.permissions_json = json.dumps(sorted(permissions))


def _migrate_app_user_access() -> None:
    with session_scope() as s:
        users = s.execute(select(AppUser).order_by(AppUser.id.asc())).scalars().all()
        if not users:
            return
        changed = False
        for index, user in enumerate(users):
            if index == 0 and not user.is_superuser and not (user.permissions_json or "").strip():
                user.is_superuser = True
                user.role_name = (user.role_name or "").strip() or "Super Admin"
                user.permissions_json = json.dumps(sorted(_ALL_PERMISSIONS))
                changed = True
                continue
            normalized = sorted(_normalize_permissions(_parse_permissions_json(user.permissions_json)))
            desired_json = json.dumps(normalized)
            if user.permissions_json != desired_json:
                user.permissions_json = desired_json
                changed = True
            if user.is_superuser and not (user.role_name or "").strip():
                user.role_name = "Super Admin"
                changed = True
        if changed:
            s.commit()


def _path_allowed_for_user(user: AppUser | SimpleNamespace | None, path: str, *, method: str = "GET", query_tab: str | None = None) -> bool:
    if user is None or not bool(getattr(user, "is_active", True)):
        return False
    method_up = method.upper()
    path = path or "/"
    if path.startswith("/profile/impersonation/stop"):
        return True
    if path == "/profile" or path == "/logout":
        return True
    if path == "/home":
        return any(
            bool(getattr(user, attr, False))
            for attr in ("can_view_projects", "can_view_people", "can_view_imports", "can_view_invoices")
        )
    if path == "/" or path.startswith("/imports"):
        permission = "imports_manage" if method_up not in _SAFE_HTTP_METHODS else "imports_view"
        return _user_has_permission(user, permission)
    if path.startswith("/projects"):
        permission = "projects_manage" if method_up not in _SAFE_HTTP_METHODS else "projects_view"
        return _user_has_permission(user, permission)
    if path.startswith("/invoice-crossreferencing"):
        permission = "invoices_manage" if method_up not in _SAFE_HTTP_METHODS else "invoices_view"
        return _user_has_permission(user, permission)
    if path.startswith("/people"):
        if method_up not in _SAFE_HTTP_METHODS:
            if path.endswith("/personal-info") or path.endswith("/emergency-contacts") or "/onboarding-documents" in path:
                return _user_has_permission(user, "people_sensitive_manage")
            return _user_has_permission(user, "people_manage")
        if "/onboarding-documents/" in path:
            return _user_has_permission(user, "people_sensitive_view")
        tab = (query_tab or "").strip().lower()
        if tab in {"personal", "emergency", "onboarding"}:
            return _user_has_permission(user, "people_sensitive_view")
        return _user_has_permission(user, "people_view")
    if path.startswith("/profile/users/"):
        if path.endswith("/impersonate"):
            return _user_has_permission(user, "impersonate_users")
        return _user_has_permission(user, "users_manage")
    if path == "/profile/users":
        return _user_has_permission(user, "users_manage")
    return True


def _default_path_for_user(user: AppUser | SimpleNamespace | None) -> str:
    if _path_allowed_for_user(user, "/home"):
        return "/home"
    if _path_allowed_for_user(user, "/"):
        return "/"
    if _path_allowed_for_user(user, "/projects"):
        return "/projects"
    if _path_allowed_for_user(user, "/people"):
        return "/people"
    if _path_allowed_for_user(user, "/imports"):
        return "/imports"
    if _path_allowed_for_user(user, "/invoice-crossreferencing"):
        return "/invoice-crossreferencing"
    return "/profile"


def _safe_next_path_for_user(user: AppUser | SimpleNamespace | None, raw_next: str | None) -> str:
    next_path = safe_next_path(raw_next)
    tab = ""
    if "?" in next_path:
        _, _, query = next_path.partition("?")
        for part in query.split("&"):
            k, _, v = part.partition("=")
            if k == "tab":
                tab = v
                break
    if _path_allowed_for_user(user, next_path.split("?", 1)[0], query_tab=tab):
        return next_path
    return _default_path_for_user(user)


def _must_manage_users(request: Request) -> None:
    current = getattr(request.state, "current_user", None)
    if not current or not current.can_manage_users:
        raise HTTPException(status_code=403, detail="You do not have permission to manage user accounts.")


def _must_impersonate_users(request: Request) -> None:
    current = getattr(request.state, "current_user", None)
    if not current or not current.can_impersonate:
        raise HTTPException(status_code=403, detail="You do not have permission to impersonate users.")


class _RequestStateMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        with session_scope() as s:
            p = _admin_profile_view(s)
            request.state.admin_profile = SimpleNamespace(
                full_name=(p.full_name or "").strip(),
                email=(p.email or "").strip(),
                phone=(p.phone or "").strip(),
            )
            user_count = s.execute(select(func.count(AppUser.id))).scalar_one()
            has_users = bool(user_count)
            user_id = request.session.get("user_id")
            acting_user_id = request.session.get("acting_user_id")
            effective_user = s.get(AppUser, user_id) if user_id else None
            acting_user = s.get(AppUser, acting_user_id) if acting_user_id else None
            if effective_user is not None and not effective_user.is_active:
                effective_user = None
                request.session.pop("user_id", None)
                request.session.pop("acting_user_id", None)
            if acting_user is not None and not acting_user.is_active:
                acting_user = None
                request.session.pop("acting_user_id", None)
            if effective_user is None:
                acting_user = None
            request.state.current_user = _user_namespace(effective_user)
            request.state.acting_user = _user_namespace(acting_user)
            request.state.is_impersonating = bool(acting_user is not None and effective_user is not None and acting_user.id != effective_user.id)
            request.state.has_users = has_users

        if request.method.upper() in _UNSAFE_HTTP_METHODS and not _same_origin_request(request):
            return HTMLResponse("Forbidden", status_code=403)

        path = request.url.path or "/"
        if _local_auth_bypass_enabled(request):
            request.state.current_user = _local_preview_user_namespace()
            request.state.acting_user = None
            request.state.is_impersonating = False
            if path in {"/login", "/setup"}:
                return _redirect("/home")
            return await call_next(request)

        if not has_users:
            if path != "/setup" and not any(path.startswith(prefix) for prefix in _AUTH_EXEMPT_PREFIXES):
                return _redirect("/setup")
        else:
            is_exempt = path in _AUTH_EXEMPT_PATHS or any(path.startswith(prefix) for prefix in _AUTH_EXEMPT_PREFIXES)
            if path == "/reset-admin-password" and _local_admin_reset_enabled(request):
                is_exempt = True
            if effective_user is None and not is_exempt:
                next_path = safe_next_path(str(request.url.path) + (f"?{request.url.query}" if request.url.query else ""))
                return _redirect(f"/login?next={quote(next_path, safe='')}")
            if effective_user is not None and path in {"/login", "/setup"}:
                return _redirect(_default_path_for_user(request.state.current_user))
            if (
                effective_user is not None
                and path != "/profile/impersonation/stop"
                and not is_exempt
                and not _path_allowed_for_user(
                    request.state.current_user,
                    path,
                    method=request.method,
                    query_tab=request.query_params.get("tab"),
                )
            ):
                return HTMLResponse("Forbidden", status_code=403)

        return await call_next(request)


app.add_middleware(_RequestStateMiddleware)
app.add_middleware(
    SessionMiddleware,
    secret_key=SECURITY_SETTINGS.session_secret,
    session_cookie=SECURITY_SETTINGS.session_cookie_name,
    max_age=SECURITY_SETTINGS.session_max_age_seconds,
    same_site="lax",
    https_only=SECURITY_SETTINGS.secure_cookies,
)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, next: str | None = None):
    with session_scope() as s:
        if _setup_available(s):
            return _redirect("/setup")
    return templates.TemplateResponse(
        request,
        "login.html",
        {
            "title": "Sign in",
            "next_path": safe_next_path(next),
            "reset_admin_available": _local_admin_reset_enabled(request),
        },
    )


@app.post("/login", response_class=HTMLResponse)
def login_submit(
    request: Request,
    username: str = Form(""),
    password: str = Form(""),
    next: str = Form("/"),
):
    username_clean = (username or "").strip()
    next_path = safe_next_path(next)
    with session_scope() as s:
        if _setup_available(s):
            return _redirect("/setup")
        user = None
        try:
            user = s.execute(select(AppUser).where(AppUser.username == normalize_username(username_clean))).scalar_one_or_none()
        except ValueError:
            user = None
        if user is None or not user.is_active or not verify_password(password, user.password_hash):
            return templates.TemplateResponse(
                request,
                "login.html",
                {
                    "title": "Sign in",
                    "error": "Incorrect username or password.",
                    "next_path": next_path,
                    "submitted_username": username_clean,
                    "reset_admin_available": _local_admin_reset_enabled(request),
                },
                status_code=400,
            )
        request.session.clear()
        request.session["user_id"] = user.id
        redirect_path = _safe_next_path_for_user(_user_namespace(user), next_path)
        user.last_login_at = datetime.now()
        s.commit()
    return _redirect(redirect_path)


@app.get("/setup", response_class=HTMLResponse)
def setup_page(request: Request):
    with session_scope() as s:
        if not _setup_available(s):
            return _redirect("/login")
    return templates.TemplateResponse(
        request,
        "setup.html",
        {
            "title": "Create first admin login",
        },
    )


@app.post("/setup", response_class=HTMLResponse)
def setup_submit(
    request: Request,
    username: str = Form(""),
    password: str = Form(""),
    confirm_password: str = Form(""),
):
    username_clean = (username or "").strip()
    try:
        normalized_username = normalize_username(username_clean)
        password_clean = validate_new_password(password, confirm_password)
    except ValueError as e:
        return templates.TemplateResponse(
            request,
            "setup.html",
            {
                "title": "Create first admin login",
                "error": str(e),
                "submitted_username": username_clean,
            },
            status_code=400,
        )

    with session_scope() as s:
        if not _setup_available(s):
            return _redirect("/login")
        user = AppUser(username=normalized_username, password_hash=hash_password(password_clean), is_active=True)
        _assign_default_access(user, make_superuser=True)
        s.add(user)
        s.flush()
        request.session.clear()
        request.session["user_id"] = user.id
        redirect_path = _default_path_for_user(_user_namespace(user))
        s.commit()
    return _redirect(redirect_path)


@app.get("/reset-admin-password", response_class=HTMLResponse)
def reset_admin_password_page(request: Request):
    if not _local_admin_reset_enabled(request):
        raise HTTPException(status_code=404, detail="Not found")
    with session_scope() as s:
        admin_users = (
            s.execute(select(AppUser).where(AppUser.is_superuser == True).order_by(AppUser.username.asc()))
            .scalars()
            .all()
        )
        if not admin_users:
            admin_users = s.execute(select(AppUser).order_by(AppUser.username.asc())).scalars().all()
    return templates.TemplateResponse(
        request,
        "reset_admin_password.html",
        {
            "title": "Reset admin password",
            "admin_users": admin_users,
        },
    )


@app.post("/reset-admin-password", response_class=HTMLResponse)
def reset_admin_password_submit(
    request: Request,
    username: str = Form(""),
    password: str = Form(""),
    confirm_password: str = Form(""),
):
    if not _local_admin_reset_enabled(request):
        raise HTTPException(status_code=404, detail="Not found")
    username_clean = (username or "").strip()
    try:
        normalized_username = normalize_username(username_clean)
        password_clean = validate_new_password(password, confirm_password)
    except ValueError as e:
        with session_scope() as s:
            admin_users = (
                s.execute(select(AppUser).where(AppUser.is_superuser == True).order_by(AppUser.username.asc()))
                .scalars()
                .all()
            )
            if not admin_users:
                admin_users = s.execute(select(AppUser).order_by(AppUser.username.asc())).scalars().all()
        return templates.TemplateResponse(
            request,
            "reset_admin_password.html",
            {
                "title": "Reset admin password",
                "error": str(e),
                "admin_users": admin_users,
                "submitted_username": username_clean,
            },
            status_code=400,
        )

    with session_scope() as s:
        user = s.execute(select(AppUser).where(AppUser.username == normalized_username)).scalar_one_or_none()
        admin_users = (
            s.execute(select(AppUser).where(AppUser.is_superuser == True).order_by(AppUser.username.asc()))
            .scalars()
            .all()
        )
        if not admin_users:
            admin_users = s.execute(select(AppUser).order_by(AppUser.username.asc())).scalars().all()
        if user is None or (admin_users and user.id not in {admin.id for admin in admin_users}):
            return templates.TemplateResponse(
                request,
                "reset_admin_password.html",
                {
                    "title": "Reset admin password",
                    "error": "Choose a valid admin account.",
                    "admin_users": admin_users,
                    "submitted_username": username_clean,
                },
                status_code=400,
            )
        user.password_hash = hash_password(password_clean)
        user.is_active = True
        user.last_login_at = datetime.now()
        request.session.clear()
        request.session["user_id"] = user.id
        redirect_path = _default_path_for_user(_user_namespace(user))
        s.commit()
    return _redirect(redirect_path)


@app.post("/logout")
def logout(request: Request):
    request.session.clear()
    return _redirect("/login")


@app.get("/profile", response_class=HTMLResponse)
def admin_profile_page(
    request: Request,
    saved: str | None = None,
    edit: str | None = None,
    access_saved: str | None = None,
    access_error: str | None = None,
    user_edit: int | None = None,
):
    with session_scope() as s:
        profile = _admin_profile_view(s)
        has_contact = _admin_profile_has_contact(profile)
        users = s.execute(select(AppUser).order_by(AppUser.username.asc())).scalars().all()
        persons = s.execute(select(Person).order_by(Person.full_name.asc())).scalars().all()
        edited_user = s.get(AppUser, user_edit) if user_edit and getattr(request.state.current_user, "can_manage_users", False) else None
        account_rows = [
            {
                "id": user.id,
                "username": user.username,
                "display_name": _user_display_name(user),
                "full_name": (user.full_name or "").strip(),
                "role_display": _user_role_display(user),
                "role_name": (user.role_name or "").strip(),
                "is_superuser": bool(user.is_superuser),
                "is_active": bool(user.is_active),
                "permissions": sorted(_permissions_for_user(user)),
                "permissions_count": len(_permissions_for_user(user)),
                "last_login_at": user.last_login_at,
                "created_at": user.created_at,
                "is_current_user": bool(getattr(request.state.current_user, "id", None) == user.id),
                "is_acting_user": bool(getattr(request.state.acting_user, "id", None) == user.id),
                "person_id": user.person_id,
                "project_manager_person_id": user.project_manager_person_id,
            }
            for user in users
        ]
    editing = (edit or "").strip() == "1" and bool(getattr(request.state.current_user, "can_manage_users", False))
    return templates.TemplateResponse(
        request,
        "profile.html",
        {
            "title": "Your profile",
            "admin_profile": profile,
            "saved": saved == "1",
            "editing": editing,
            "admin_profile_has_contact": has_contact,
            "access_saved": access_saved == "1",
            "access_error": (access_error or "").strip(),
            "can_manage_users": bool(getattr(request.state.current_user, "can_manage_users", False)),
            "can_impersonate_users": bool(getattr(request.state.current_user, "can_impersonate", False)),
            "permission_groups": _PERMISSION_GROUPS,
            "account_rows": account_rows,
            "person_rows": [{"id": p.id, "full_name": p.full_name} for p in persons],
            "edited_user": edited_user,
            "edited_user_permissions": _permissions_for_user(edited_user),
            "is_impersonating": bool(getattr(request.state, "is_impersonating", False)),
            "acting_user": getattr(request.state, "acting_user", None),
            "current_user": getattr(request.state, "current_user", None),
        },
    )


def _admin_profile_has_contact(p: AdminProfile) -> bool:
    return bool(
        (p.full_name or "").strip()
        or (p.email or "").strip()
        or (p.phone or "").strip()
        or (p.notes or "").strip()
    )


@app.post("/profile")
def admin_profile_save(
    request: Request,
    full_name: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    notes: str = Form(""),
):
    _must_manage_users(request)
    with session_scope() as s:
        p = _admin_profile_view(s)
        fn = (full_name or "").strip()
        em = (email or "").strip()
        ph = (phone or "").strip()
        nt = (notes or "").strip()
        p.full_name = fn or None
        p.email = em or None
        p.phone = ph or None
        p.notes = nt or None
        p.updated_at = datetime.now()
        s.commit()
    return _redirect("/profile?saved=1")


@app.post("/profile/users")
def profile_users_create(
    request: Request,
    username: str = Form(""),
    full_name: str = Form(""),
    role_name: str = Form(""),
    password: str = Form(""),
    confirm_password: str = Form(""),
    permissions: list[str] = Form([]),
    is_superuser: str = Form(""),
    person_id: str = Form(""),
    project_manager_person_id: str = Form(""),
):
    _must_manage_users(request)
    username_clean = (username or "").strip()
    full_name_clean = (full_name or "").strip()
    role_name_clean = (role_name or "").strip()
    try:
        normalized_username = normalize_username(username_clean)
        password_clean = validate_new_password(password, confirm_password)
    except ValueError as e:
        return _redirect(f"/profile?access_error={quote(str(e))}")

    permissions_set = _normalize_permissions(set(permissions or []))
    make_superuser = (is_superuser or "").strip().lower() in {"1", "true", "yes", "on"}
    if not make_superuser and not permissions_set:
        return _redirect("/profile?access_error=Choose%20at%20least%20one%20permission%20or%20make%20the%20account%20a%20super%20admin.")

    person_id_val = None
    if person_id.strip():
        try:
            person_id_val = int(person_id.strip())
        except ValueError:
            pass

    project_manager_person_id_val = None
    if project_manager_person_id.strip():
        try:
            project_manager_person_id_val = int(project_manager_person_id.strip())
        except ValueError:
            pass

    with session_scope() as s:
        user = AppUser(
            username=normalized_username,
            full_name=full_name_clean or None,
            role_name=role_name_clean or None,
            password_hash=hash_password(password_clean),
            is_active=True,
            person_id=person_id_val,
            project_manager_person_id=project_manager_person_id_val,
        )
        if make_superuser:
            _assign_default_access(user, make_superuser=True)
        else:
            user.is_superuser = False
            user.permissions_json = _permissions_json_from_form(list(permissions_set))
        s.add(user)
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            return _redirect("/profile?access_error=That%20username%20already%20exists.")
    return _redirect("/profile?access_saved=1")


@app.post("/profile/users/{user_id}")
def profile_users_update(
    request: Request,
    user_id: int,
    full_name: str = Form(""),
    role_name: str = Form(""),
    password: str = Form(""),
    confirm_password: str = Form(""),
    permissions: list[str] = Form([]),
    is_superuser: str = Form(""),
    is_active: str = Form(""),
    person_id: str = Form(""),
    project_manager_person_id: str = Form(""),
):
    _must_manage_users(request)
    full_name_clean = (full_name or "").strip()
    role_name_clean = (role_name or "").strip()
    make_superuser = (is_superuser or "").strip().lower() in {"1", "true", "yes", "on"}
    keep_active = (is_active or "").strip().lower() in {"1", "true", "yes", "on"}
    permissions_set = _normalize_permissions(set(permissions or []))

    if password or confirm_password:
        try:
            password_clean = validate_new_password(password, confirm_password)
        except ValueError as e:
            return _redirect(f"/profile?user_edit={user_id}&access_error={quote(str(e))}")
    else:
        password_clean = None

    if not make_superuser and not permissions_set:
        return _redirect(f"/profile?user_edit={user_id}&access_error=Choose%20at%20least%20one%20permission%20or%20make%20the%20account%20a%20super%20admin.")

    owner_id = getattr(request.state.acting_user, "id", None) or getattr(request.state.current_user, "id", None)
    if not keep_active and owner_id == user_id:
        return _redirect(f"/profile?user_edit={user_id}&access_error=You%20cannot%20deactivate%20the%20account%20currently%20controlling%20this%20session.")

    person_id_val = None
    if person_id.strip():
        try:
            person_id_val = int(person_id.strip())
        except ValueError:
            pass

    project_manager_person_id_val = None
    if project_manager_person_id.strip():
        try:
            project_manager_person_id_val = int(project_manager_person_id.strip())
        except ValueError:
            pass

    with session_scope() as s:
        user = s.get(AppUser, user_id)
        if not user:
            raise HTTPException(status_code=404, detail="User not found.")
        user.full_name = full_name_clean or None
        user.role_name = role_name_clean or None
        user.is_active = keep_active
        user.person_id = person_id_val
        user.project_manager_person_id = project_manager_person_id_val
        if password_clean:
            user.password_hash = hash_password(password_clean)
        if make_superuser:
            _assign_default_access(user, make_superuser=True)
        else:
            user.is_superuser = False
            user.permissions_json = _permissions_json_from_form(list(permissions_set))
        s.commit()
    return _redirect("/profile?access_saved=1")


@app.post("/profile/users/{user_id}/impersonate")
def profile_users_impersonate(request: Request, user_id: int):
    _must_impersonate_users(request)
    current_id = getattr(request.state.current_user, "id", None)
    if current_id == user_id:
        return _redirect("/profile")
    with session_scope() as s:
        target = s.get(AppUser, user_id)
        if not target or not target.is_active:
            raise HTTPException(status_code=404, detail="Target user not found.")
        original_id = getattr(request.state.acting_user, "id", None) or current_id
        if original_id:
            request.session["acting_user_id"] = original_id
        request.session["user_id"] = target.id
        redirect_path = _default_path_for_user(_user_namespace(target))
    return _redirect(redirect_path)


@app.post("/profile/impersonation/stop")
def profile_impersonation_stop(request: Request):
    acting_user_id = request.session.get("acting_user_id")
    if not acting_user_id:
        return _redirect("/profile")
    with session_scope() as s:
        acting_user = s.get(AppUser, acting_user_id)
    request.session.pop("acting_user_id", None)
    if acting_user and acting_user.is_active:
        request.session["user_id"] = acting_user.id
        return _redirect(_default_path_for_user(_user_namespace(acting_user)))
    request.session.clear()
    return _redirect("/login")


def _norm_pay_type_key(value: str) -> str:
    return "".join(ch.lower() for ch in value.strip() if ch.isalnum())


def _norm_timesheet_id(value: str) -> str:
    v = value.strip()
    if v.endswith(".0"):
        v = v[:-2]
    return v.lower()


def _normalize_pay_type(value: str, alias_map: dict[str, str]) -> str:
    key = _norm_pay_type_key(value)
    return alias_map.get(key, value.strip())


def _invoice_key(timesheet_id: str, pay_type: str, alias_map: dict[str, str]) -> tuple[str, str]:
    normalized_pay_type = _normalize_pay_type(pay_type, alias_map)
    return (_norm_timesheet_id(timesheet_id), normalized_pay_type.strip().lower())


def _invoice_badge_cluster_key(line: object) -> str:
    """
    VMS ID for grouping; when blank, a surrogate from company / name / invoice / week so unrelated
    rows without a badge are not merged or deduped together.
    """
    raw = getattr(line, "external_timesheet_id", None)
    ts = _norm_timesheet_id(str(raw or ""))
    if ts:
        return ts
    en = (getattr(line, "employee_name", None) or "").strip().lower()
    co = (getattr(line, "company_name", None) or "").strip().lower()
    inv = getattr(line, "invoice_number", None)
    inv_s = str(inv).strip() if inv is not None else ""
    wk = ""
    wo = getattr(line, "week_ended_on", None)
    if wo is not None:
        wk = wo.isoformat() if hasattr(wo, "isoformat") else str(wo)[:10]
    return f"\x00no_badge:{co}\x1f{en}\x1f{inv_s}\x1f{wk}"


def _invoice_line_rate_qty_key(line: InvoiceLine) -> tuple[str, str]:
    """Pay rate and quantity from the imported line, bucketed for cross-reference identity."""
    if line.pay_rate is None:
        pr_s = ""
    else:
        pr_s = format(_as_decimal(line.pay_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f")
    if line.quantity is None:
        qt_s = ""
    else:
        qt_s = format(_as_decimal(line.quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f")
    return pr_s, qt_s


def _invoice_week_key(line: object) -> str:
    """ISO date string for week ending, or empty when unknown (legacy imports without a week column)."""
    w = getattr(line, "week_ended_on", None)
    if w is None:
        return ""
    # datetime is a subclass of date — handle explicitly so we always store a calendar date
    if isinstance(w, datetime):
        return w.date().isoformat()
    if isinstance(w, date):
        return w.isoformat()
    if isinstance(w, str) and w.strip():
        try:
            return date.fromisoformat(w.strip()[:10]).isoformat()
        except ValueError:
            return w.strip()
    return ""


def _invoice_week_display_line(line: object | None) -> str:
    """US short date for cross-reference display; empty if the import row has no week."""
    if not line:
        return ""
    wk = _invoice_week_key(line)
    if not wk:
        return ""
    try:
        return date.fromisoformat(wk[:10]).strftime("%m/%d/%Y")
    except ValueError:
        return wk


def _invoice_crossref_key(line: InvoiceLine, alias_map: dict[str, str]) -> tuple[str, str]:
    """
    Badge (VMS, or surrogate if VMS blank) + mapped pay type. Same merge key → merged for cross-ref.
    Shoals adds work week in `_merge_key_for_crossref` for per-week comparison.
    """
    badge = _invoice_badge_cluster_key(line)
    pt = _normalize_pay_type(line.pay_type, alias_map).strip().lower()
    return (badge, pt)


def _invoice_crossref_key_from_comparison(c: dict, alias_map: dict[str, str]) -> tuple[str, str]:
    """Same badge + pay type as `_invoice_crossref_key` for a snapshot / UI row dict."""
    wk: date | None = None
    wk_s = (str(c.get("week_ended_iso") or "")).strip()
    if wk_s:
        try:
            wk = date.fromisoformat(wk_s[:10])
        except ValueError:
            pass
    stub = SimpleNamespace(
        external_timesheet_id=c.get("external_timesheet_id"),
        pay_type=c.get("pay_type"),
        employee_name=c.get("employee_name"),
        company_name=c.get("company_name"),
        invoice_number=c.get("invoice_number"),
        week_ended_on=wk,
    )
    return _invoice_crossref_key(stub, alias_map)


def _merge_invoice_lines_for_crossref(lines: list[InvoiceLine]) -> InvoiceLine | SimpleNamespace:
    """Combine multiple imported lines that share badge + pay type (cross-ref identity)."""
    if len(lines) == 1:
        return lines[0]

    def wavg(attr: str) -> float | None:
        num = Decimal("0")
        den = Decimal("0")
        for x in lines:
            q = _as_decimal(x.quantity)
            v = getattr(x, attr, None)
            if v is None:
                continue
            num += _as_decimal(v) * q
            den += q
        if den > 0:
            return float((num / den).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))
        vals = [_as_decimal(getattr(x, attr)) for x in lines if getattr(x, attr, None) is not None]
        if not vals:
            return None
        return float(sum(vals) / len(vals))

    total_q = sum(_as_decimal(x.quantity) for x in lines)
    total_amt = sum(_as_decimal(x.total_amount) for x in lines)
    wk: date | None = None
    for x in lines:
        if x.week_ended_on is not None:
            wk = x.week_ended_on
            break
    inv = next((x.invoice_number for x in lines if x.invoice_number and str(x.invoice_number).strip()), None)
    en = next((x.employee_name for x in lines if x.employee_name and str(x.employee_name).strip()), None)
    co = next((x.company_name for x in lines if x.company_name and str(x.company_name).strip()), None)
    first = lines[0]
    return SimpleNamespace(
        external_timesheet_id=first.external_timesheet_id,
        pay_type=first.pay_type,
        employee_name=en or first.employee_name,
        company_name=co or first.company_name,
        invoice_number=inv or first.invoice_number,
        pay_rate=wavg("pay_rate"),
        bill_rate=wavg("bill_rate"),
        quantity=float(total_q) if any(x.quantity is not None for x in lines) else None,
        total_amount=float(total_amt) if any(x.total_amount is not None for x in lines) else None,
        week_ended_on=wk,
    )


def _as_decimal(v: Decimal | float | None) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    return Decimal(str(v))


def _invoice_line_total_billed_is_zero(line: InvoiceLine) -> bool:
    """True when the paid/billed row has $0 in the total-billed amount (omit from cross-reference)."""
    return _as_decimal(line.total_amount) == 0


def _invoice_loose_line_key(line: InvoiceLine, alias_map: dict[str, str]) -> tuple[str, str, str]:
    """Badge + mapped pay type + pay rate — ignores qty (detects duplicate phantoms)."""
    badge = _invoice_badge_cluster_key(line)
    pt = _normalize_pay_type(line.pay_type, alias_map).strip().lower()
    pr_s, _ = _invoice_line_rate_qty_key(line)
    return (badge, pt, pr_s)


def _is_invoice_zero_placeholder_line(line: InvoiceLine) -> bool:
    """True when both billed hours and extended amount are zero (duplicate line under another invoice #)."""
    return _as_decimal(line.quantity) == 0 and _as_decimal(line.total_amount) == 0


def _dedupe_invoice_zero_placeholders(lines: list[InvoiceLine], alias_map: dict[str, str]) -> list[InvoiceLine]:
    """
    If the file repeats the same person + pay code + pay rate with one row at 0 hrs / $0 and
    another with real hours (e.g. invoice 200878 vs 200880), keep only the substantive rows.
    """
    groups: defaultdict[tuple[str, str, str], list[InvoiceLine]] = defaultdict(list)
    for r in lines:
        groups[_invoice_loose_line_key(r, alias_map)].append(r)
    out: list[InvoiceLine] = []
    for g in groups.values():
        if len(g) == 1:
            out.append(g[0])
            continue
        real = [x for x in g if not _is_invoice_zero_placeholder_line(x)]
        phantoms = [x for x in g if _is_invoice_zero_placeholder_line(x)]
        if real and phantoms:
            out.extend(real)
        else:
            out.extend(g)
    return out


def _drop_zero_qty_when_other_pay_type_has_hours(lines: list[InvoiceLine], alias_map: dict[str, str]) -> list[InvoiceLine]:
    """
    Per VMS / import side: omit 0-hr rows when that badge already has positive hours on another pay type
    (mapped pay type). Same pay-type 0 lines are left for `_dedupe_invoice_zero_placeholders`.
    """
    if not lines:
        return lines
    by_badge: defaultdict[str, list[InvoiceLine]] = defaultdict(list)
    for r in lines:
        by_badge[_invoice_badge_cluster_key(r)].append(r)
    out: list[InvoiceLine] = []
    for group in by_badge.values():
        pos_pts: set[str] = set()
        for r in group:
            if _as_decimal(r.quantity) > 0:
                _, pt = _invoice_crossref_key(r, alias_map)
                pos_pts.add(pt)
        for r in group:
            q = _as_decimal(r.quantity)
            _, pt = _invoice_crossref_key(r, alias_map)
            if q == 0 and pos_pts and pt not in pos_pts:
                continue
            out.append(r)
    return out


def _get_invoice_tolerance(s) -> Decimal:
    setting = s.execute(select(InvoiceSetting).where(InvoiceSetting.key == "mismatch_tolerance")).scalar_one_or_none()
    if not setting:
        return Decimal("0")
    try:
        value = Decimal(setting.value.strip())
    except Exception:
        return Decimal("0")
    return value if value >= 0 else Decimal("0")


def _get_paid_billed_company_filter(s) -> str:
    setting = s.execute(select(InvoiceSetting).where(InvoiceSetting.key == "paid_billed_company_filter")).scalar_one_or_none()
    if not setting:
        return ""
    return setting.value.strip()


# Reserved InvoiceSettingsPreset.name rows: one snapshot (tolerance, company filter, mappings) per company workflow.
_LAYOUT_SNAPSHOT_PRESET_NAMES: dict[str, str] = {
    "default": "__sys_invoice_layout_snapshot_default__",
    "material_motion": "__sys_invoice_layout_snapshot_material_motion__",
    "shoals": "__sys_invoice_layout_snapshot_shoals__",
}


def _save_invoice_layout_snapshot(s, layout: str) -> None:
    name = _LAYOUT_SNAPSHOT_PRESET_NAMES.get((layout or "").strip().lower())
    if not name:
        raise HTTPException(status_code=400, detail="Invalid target layout.")
    snap = _snapshot_invoice_crossref_settings(s)
    payload = json.dumps(snap)
    existing = s.execute(select(InvoiceSettingsPreset).where(InvoiceSettingsPreset.name == name)).scalar_one_or_none()
    if existing:
        existing.payload_json = payload
    else:
        s.add(InvoiceSettingsPreset(name=name, payload_json=payload))


def _apply_invoice_layout_snapshot_if_present(s, layout: str) -> None:
    name = _LAYOUT_SNAPSHOT_PRESET_NAMES.get((layout or "").strip().lower())
    if not name:
        return
    row = s.execute(select(InvoiceSettingsPreset).where(InvoiceSettingsPreset.name == name)).scalar_one_or_none()
    if not row or not (row.payload_json or "").strip():
        return
    try:
        data = json.loads(row.payload_json)
    except json.JSONDecodeError:
        return
    try:
        _apply_invoice_crossref_snapshot(s, data)
    except ValueError:
        return


def _snapshot_invoice_crossref_settings(s) -> dict:
    mappings = (
        s.execute(select(InvoicePayTypeMapping).order_by(InvoicePayTypeMapping.source_value.asc())).scalars().all()
    )
    return {
        "version": 1,
        "mismatch_tolerance": str(_get_invoice_tolerance(s)),
        "company_filter": _get_paid_billed_company_filter(s),
        "mappings": [{"source": m.source_value, "target": m.target_value} for m in mappings],
    }


def _apply_invoice_crossref_snapshot(s, payload: dict) -> None:
    if not isinstance(payload, dict):
        raise ValueError("Invalid preset data")
    raw_tol = str(payload.get("mismatch_tolerance", "0")).strip()
    try:
        tol = Decimal(raw_tol)
    except Exception as e:
        raise ValueError("Tolerance in preset must be a number.") from e
    if tol < 0:
        raise ValueError("Tolerance cannot be negative.")
    company = (payload.get("company_filter") or "").strip()[:200]

    st = s.execute(select(InvoiceSetting).where(InvoiceSetting.key == "mismatch_tolerance")).scalar_one_or_none()
    if not st:
        s.add(InvoiceSetting(key="mismatch_tolerance", value=str(tol)))
    else:
        st.value = str(tol)

    st2 = s.execute(select(InvoiceSetting).where(InvoiceSetting.key == "paid_billed_company_filter")).scalar_one_or_none()
    if not st2:
        s.add(InvoiceSetting(key="paid_billed_company_filter", value=company))
    else:
        st2.value = company

    s.execute(delete(InvoicePayTypeMapping))
    for row in payload.get("mappings", []):
        if not isinstance(row, dict):
            continue
        src = (row.get("source") or "").strip()
        tgt = (row.get("target") or "").strip()
        if not src or not tgt:
            continue
        s.add(InvoicePayTypeMapping(source_value=src, target_value=tgt))


def _split_first_last(full: str | None) -> tuple[str, str]:
    if not full:
        return "", ""
    parts = str(full).strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _try_int_badge(s: object) -> int | str | float:
    try:
        return int(float(str(s).strip()))
    except (ValueError, TypeError):
        return str(s).strip() if s is not None else ""


def _batch_week_date(batch: ImportBatch | None) -> date | None:
    """Calendar date from batch import time (no time-of-day); used as export week fallback."""
    if not batch or not batch.imported_at:
        return None
    d = batch.imported_at
    if getattr(d, "tzinfo", None) is not None:
        d = d.replace(tzinfo=None)
    return date(d.year, d.month, d.day)


EXPORT_SHORT_DATE_FMT = "m/d/yyyy"
# US accounting-style currency: $, thousands separator, negatives in parentheses (Excel format string).
EXPORT_ACCOUNTING_USD_FMT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
# Comma-style quantity / hours (thousands separator, two decimals).
EXPORT_COMMA_QTY_FMT = "#,##0.00"


def _format_export_numeric_column(ws, col_idx: int, fmt: str) -> None:
    """Apply Excel number format to numeric data cells (row 2+); skips None, dates, bool, text."""
    if ws.max_row < 2:
        return
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if v is None or isinstance(v, bool):
            continue
        if isinstance(v, date):
            continue
        if isinstance(v, (int, float)):
            cell.number_format = fmt


def _format_export_short_date_column(ws, col_idx: int) -> None:
    """US short date in Excel (e.g. 3/28/2026), no time component."""
    if ws.max_row < 2:
        return
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if v is None:
            continue
        if isinstance(v, datetime):
            cell.value = v.replace(tzinfo=None).date() if v.tzinfo else v.date()
        elif not isinstance(v, date):
            continue
        cell.number_format = EXPORT_SHORT_DATE_FMT


def _mim_payroll_lines(mim_lines: list[InvoiceLine]) -> list[InvoiceLine]:
    return [r for r in mim_lines if (r.pay_type or "").strip().lower() != "pto payout"]


def _mim_pto_payout_lines(mim_lines: list[InvoiceLine]) -> list[InvoiceLine]:
    return [r for r in mim_lines if (r.pay_type or "").strip().lower() == "pto payout"]


def _rates_from_mim(mim_lines: list[InvoiceLine]) -> list[InvoiceLine]:
    """One row per badge; prefer REG pay type for PayRate."""
    by_badge: dict[str, list[InvoiceLine]] = defaultdict(list)
    for r in mim_lines:
        by_badge[_invoice_badge_cluster_key(r)].append(r)
    out: list[InvoiceLine] = []
    for bid in sorted(by_badge.keys(), key=_badge_sort_key):
        rows = by_badge[bid]
        reg = [x for x in rows if (x.pay_type or "").upper() in ("REG", "REGULAR")]
        out.append(reg[0] if reg else rows[0])
    return out


def _first_paid_invoice_number(paid_lines: list[InvoiceLine]) -> str | None:
    for r in paid_lines:
        if r.invoice_number and str(r.invoice_number).strip():
            return str(r.invoice_number).strip()
    return None


def _badge_sort_key(bid: str) -> tuple:
    s = str(bid).strip()
    if s.isdigit():
        return (0, int(s))
    return (1, s)


def _fit_export_sheet_columns(ws, *, min_width: float = 10.0) -> None:
    """Bold first row (headers) and set column widths from cell contents (openpyxl has no true auto-fit)."""
    if ws.max_row < 1 or ws.max_column < 1:
        return
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    # Excel column width tops out around 255 character units.
    cap = 255.0
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            text = str(val)
            for line in text.splitlines():
                max_len = max(max_len, len(line))
        w = min(max(max_len + 2, min_width), cap)
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _invoice_cell_value(v: object | None) -> object | None:
    if v is None or not str(v).strip():
        return None
    s = str(v).strip().replace(",", "")
    try:
        if "." in s:
            f = float(s)
            if f == int(f):
                return int(f)
            return f
        return int(s)
    except (ValueError, TypeError):
        return str(v).strip()


def _payroll_export_invoice_no(
    mim_row: InvoiceLine,
    alias_map: dict[str, str],
    paid_invoice_by_key: dict[tuple, str],
    batch_invoice: str | None,
) -> object | None:
    """
    Column A on Payroll/PTO Payout: prefer invoice # from matching Paid/Billed line (same key),
    else MIM row, else first invoice from filtered paid/billed batch.
    """
    k = _invoice_crossref_key(mim_row, alias_map)
    if k in paid_invoice_by_key:
        return _invoice_cell_value(paid_invoice_by_key[k])
    k_loose = _invoice_key(mim_row.external_timesheet_id, mim_row.pay_type, alias_map)
    if k_loose in paid_invoice_by_key:
        return _invoice_cell_value(paid_invoice_by_key[k_loose])
    if mim_row.invoice_number and str(mim_row.invoice_number).strip():
        return _invoice_cell_value(mim_row.invoice_number)
    if batch_invoice:
        return _invoice_cell_value(batch_invoice)
    return None


def _crossref_comparison_is_ppe_row(c: dict, alias_map: dict[str, str]) -> bool:
    if c.get("is_ppe_row"):
        return True
    ppe_label = _normalize_pay_type("MIM PPE", alias_map).strip().lower()
    return (c.get("pay_type") or "").strip().lower() == ppe_label


def _crossref_comparison_is_pto_payout_row(c: dict) -> bool:
    return (c.get("pay_type") or "").strip().lower() == "pto payout"


def _week_from_crossref_comparison(c: dict, week_fallback: date | None) -> date | None:
    wk = (c.get("week_ended_iso") or "").strip()
    if wk:
        try:
            return date.fromisoformat(wk[:10])
        except ValueError:
            pass
    return week_fallback


def _company_name_for_payroll_export(
    c: dict,
    alias_map: dict[str, str],
    company_by_key: dict[tuple[str, str], str],
) -> str:
    explicit = (c.get("company_name") or "").strip()
    if explicit:
        return explicit
    ik = _invoice_crossref_key_from_comparison(c, alias_map)
    return (company_by_key.get(ik) or "").strip()


def _payroll_export_invoice_from_comparison(
    c: dict,
    alias_map: dict[str, str],
    paid_invoice_by_key: dict[tuple, str],
    batch_invoice: str | None,
) -> object | None:
    inv = (c.get("invoice_number") or "").strip()
    if inv:
        return _invoice_cell_value(inv)
    if c.get("shoals_hours_only"):
        badge, pt = _invoice_crossref_key_from_comparison(c, alias_map)
        wk = (str(c.get("week_ended_iso") or "")).strip()[:10]
        k3 = (badge, pt, wk)
        if k3 in paid_invoice_by_key:
            return _invoice_cell_value(paid_invoice_by_key[k3])
    wk_cmp: date | None = None
    wk_s = (str(c.get("week_ended_iso") or "")).strip()
    if wk_s:
        try:
            wk_cmp = date.fromisoformat(wk_s[:10])
        except ValueError:
            pass
    stub = SimpleNamespace(
        external_timesheet_id=c.get("external_timesheet_id"),
        pay_type=c.get("pay_type"),
        invoice_number=c.get("invoice_number"),
        employee_name=c.get("employee_name"),
        company_name=c.get("company_name"),
        week_ended_on=wk_cmp,
    )
    return _payroll_export_invoice_no(stub, alias_map, paid_invoice_by_key, batch_invoice)


def _export_float_cell(v: object | None) -> float | None:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _build_invoice_comparison(
    paid_billed_batch_id: int | None,
    mim_line_batch_ids: list[int] | None,
    *,
    crossref_layout: str = "material_motion",
):
    paid_map: dict[tuple[str, str, str], InvoiceLine | SimpleNamespace] = {}
    mim_map: dict[tuple[str, str, str], InvoiceLine | SimpleNamespace] = {}
    alias_map: dict[str, str] = {}
    tolerance = Decimal("0")
    selected_company = ""
    ppe_map: dict[str, Decimal] = {}
    ppe_keys: set[tuple[str, str, str]] = set()
    batch_invoice_no = ""
    paid_rows: list[InvoiceLine] = []
    mim_rows: list[InvoiceLine] = []
    shoals_hours_only = crossref_layout == "shoals"

    def _merge_key_for_crossref(line: InvoiceLine) -> tuple[str, str, str]:
        ts, pt = _invoice_crossref_key(line, alias_map)
        if shoals_hours_only:
            return (ts, pt, _invoice_week_key(line))
        return (ts, pt, "")

    with session_scope() as s:
        mappings = s.execute(select(InvoicePayTypeMapping)).scalars().all()
        alias_map = {_norm_pay_type_key(m.source_value): m.target_value.strip() for m in mappings}
        tolerance = _get_invoice_tolerance(s)
        selected_company = _get_paid_billed_company_filter(s)
        mapped_ppe_type = ""
        if not shoals_hours_only:
            ppe_batch = (
                s.execute(
                    select(ImportBatch).where(ImportBatch.note == INVOICE_NOTE_PPE).order_by(desc(ImportBatch.imported_at)).limit(1)
                )
                .scalars()
                .first()
            )
            if ppe_batch:
                ppe_rows = s.execute(select(PpeDeductionLine).where(PpeDeductionLine.import_batch_id == ppe_batch.id)).scalars().all()
                for p in ppe_rows:
                    key = _norm_timesheet_id(p.external_timesheet_id)
                    ppe_map[key] = ppe_map.get(key, Decimal("0")) + _as_decimal(p.deduction_amount)
            mapped_ppe_type = _normalize_pay_type("MIM PPE", alias_map).strip().lower()

        if paid_billed_batch_id:
            rows = (
                s.execute(
                    select(InvoiceLine)
                    .where(InvoiceLine.import_batch_id == paid_billed_batch_id)
                    .order_by(InvoiceLine.id.asc())
                )
                .scalars()
                .all()
            )
            for r in rows:
                if selected_company and (r.company_name or "").strip() != selected_company:
                    continue
                paid_rows.append(r)
            paid_rows = _dedupe_invoice_zero_placeholders(paid_rows, alias_map)
            paid_rows = _drop_zero_qty_when_other_pay_type_has_hours(paid_rows, alias_map)
            for r in paid_rows:
                if (
                    not batch_invoice_no
                    and not _invoice_line_total_billed_is_zero(r)
                    and r.invoice_number
                    and str(r.invoice_number).strip()
                ):
                    batch_invoice_no = str(r.invoice_number).strip()
                    break
            paid_groups: defaultdict[tuple[str, str, str], list[InvoiceLine]] = defaultdict(list)
            for r in paid_rows:
                paid_groups[_merge_key_for_crossref(r)].append(r)
            for _k, lst in paid_groups.items():
                lst2 = [x for x in lst if not _invoice_line_total_billed_is_zero(x)]
                if not lst2:
                    continue
                paid_map[_k] = _merge_invoice_lines_for_crossref(lst2)

        batch_ids = list(mim_line_batch_ids or [])
        if batch_ids:
            rows_all: list[InvoiceLine] = []
            for mim_batch_id in batch_ids:
                rows = (
                    s.execute(
                        select(InvoiceLine)
                        .where(InvoiceLine.import_batch_id == mim_batch_id)
                        .order_by(InvoiceLine.id.asc())
                    )
                    .scalars()
                    .all()
                )
                for r in rows:
                    rows_all.append(r)
            mim_rows = _dedupe_invoice_zero_placeholders(rows_all, alias_map)
            mim_rows = _drop_zero_qty_when_other_pay_type_has_hours(mim_rows, alias_map)
            mim_groups: defaultdict[tuple[str, str, str], list[InvoiceLine]] = defaultdict(list)
            for r in mim_rows:
                mim_groups[_merge_key_for_crossref(r)].append(r)
            for _k, lst in mim_groups.items():
                mim_map[_k] = _merge_invoice_lines_for_crossref(lst)

        if not shoals_hours_only:
            ppe_keys = {(kid, mapped_ppe_type, "") for kid in ppe_map.keys()}

        comparisons: list = []

        def _append_comparison_row(
            paid: InvoiceLine | SimpleNamespace | None,
            mim: InvoiceLine | SimpleNamespace | None,
            k: tuple[str, str, str],
            is_ppe_row: bool,
        ) -> None:
            paid_pay_rate = _as_decimal(paid.pay_rate if paid else None)
            mim_pay_rate = _as_decimal(mim.pay_rate if mim else None)
            paid_bill_rate = _as_decimal(paid.bill_rate if paid else None)
            mim_bill_rate = _as_decimal(mim.bill_rate if mim else None)
            paid_qty = _as_decimal(paid.quantity if paid else None)
            mim_qty = _as_decimal(mim.quantity if mim else None)
            mim_qty_import = mim_qty  # before PPE override; used for "MIM-only, qty 0" soft highlight
            paid_total = _as_decimal(paid.total_amount if paid else None)
            mim_total = _as_decimal(mim.total_amount if mim else None)

            pay_rate_diff = paid_pay_rate - mim_pay_rate
            bill_rate_diff = paid_bill_rate - mim_bill_rate

            if is_ppe_row:
                ppe_deduction = ppe_map.get(k[0], Decimal("0"))
                mim_total = -abs(ppe_deduction)
                mim_pay_rate = Decimal("0")
                mim_bill_rate = Decimal("0")
                mim_qty = Decimal("0")
                pay_rate_diff = paid_pay_rate - mim_pay_rate
                bill_rate_diff = paid_bill_rate - mim_bill_rate

            qty_diff = paid_qty - mim_qty
            total_diff = paid_total - mim_total
            pay_rate_mismatch = False
            bill_rate_mismatch = False
            markup_mismatch = False

            if shoals_hours_only and not is_ppe_row:
                paid_pay_rate = Decimal("0")
                mim_pay_rate = Decimal("0")
                paid_bill_rate = Decimal("0")
                mim_bill_rate = Decimal("0")
                pay_rate_diff = Decimal("0")
                bill_rate_diff = Decimal("0")
                paid_total = Decimal("0")
                mim_total = Decimal("0")
                total_diff = Decimal("0")
                total_mismatch = abs(qty_diff) > tolerance
            else:
                total_mismatch = abs(total_diff) > tolerance
                if mim and not is_ppe_row and pay_type_uses_1663_markup(mim.pay_type):
                    exp_total = markup_1663_expected_total(mim_pay_rate, mim_qty)
                    if exp_total is not None:
                        markup_mismatch = abs(mim_total - exp_total) > tolerance

            inv_num = ""
            if paid and paid.invoice_number and str(paid.invoice_number).strip():
                inv_num = str(paid.invoice_number).strip()
            elif mim and mim.invoice_number and str(mim.invoice_number).strip():
                inv_num = str(mim.invoice_number).strip()
            elif batch_invoice_no:
                inv_num = batch_invoice_no

            paid_week_ended = ""
            mim_week_ended = ""
            if not is_ppe_row:
                paid_week_ended = _invoice_week_display_line(paid) if paid else ""
                mim_week_ended = _invoice_week_display_line(mim) if mim else ""
            wk_iso = ""
            if not is_ppe_row:
                wk_iso = _invoice_week_key(paid) if paid else _invoice_week_key(mim) if mim else ""
            week_display = ""
            if wk_iso:
                try:
                    week_display = date.fromisoformat(wk_iso[:10]).strftime("%m/%d/%Y")
                except ValueError:
                    week_display = wk_iso

            mim_no_paid_zero_mim_qty = (
                paid is None
                and mim is not None
                and not is_ppe_row
                and mim_qty_import == Decimal("0")
            )

            row: dict = {
                "invoice_number": inv_num,
                "external_timesheet_id": (
                    k[0]
                    if is_ppe_row
                    else (paid.external_timesheet_id if paid else (mim.external_timesheet_id if mim else ""))
                ),
                "pay_type": (
                    _normalize_pay_type("MIM PPE", alias_map)
                    if is_ppe_row
                    else (
                        _normalize_pay_type(paid.pay_type, alias_map)
                        if paid
                        else (_normalize_pay_type(mim.pay_type, alias_map) if mim else "")
                    )
                ),
                "week_ended_iso": wk_iso,
                "week_ended": week_display,
                "paid_week_ended": paid_week_ended,
                "mim_week_ended": mim_week_ended,
                "employee_name": (paid.employee_name if paid and paid.employee_name else (mim.employee_name if mim else "")),
                "company_name": (
                    (paid.company_name if paid and (paid.company_name or "").strip() else None)
                    or (mim.company_name if mim and (mim.company_name or "").strip() else None)
                    or ""
                ),
                "is_ppe_row": is_ppe_row,
                "paid_pay_rate": paid_pay_rate,
                "mim_pay_rate": mim_pay_rate,
                "pay_rate_diff": pay_rate_diff,
                "paid_bill_rate": paid_bill_rate,
                "mim_bill_rate": mim_bill_rate,
                "bill_rate_diff": bill_rate_diff,
                "paid_qty": paid_qty,
                "mim_qty": mim_qty,
                "qty_diff": qty_diff,
                "mim_no_paid_zero_mim_qty": mim_no_paid_zero_mim_qty,
                "paid_total": paid_total,
                "mim_total": mim_total,
                "total_diff": total_diff,
                "has_paid": paid is not None,
                "has_mim": (mim is not None) or is_ppe_row,
                "pay_rate_mismatch": pay_rate_mismatch,
                "bill_rate_mismatch": bill_rate_mismatch,
                "total_mismatch": total_mismatch,
                "markup_mismatch": markup_mismatch,
                "shoals_hours_only": shoals_hours_only,
                "any_mismatch": (
                    ((paid is None) and not mim_no_paid_zero_mim_qty)
                    or ((mim is None) and (not is_ppe_row))
                    or total_mismatch
                    or markup_mismatch
                ),
            }
            comparisons.append(row)

        keys = sorted(set(paid_map.keys()) | set(mim_map.keys()) | ppe_keys)
        for k in keys:
            _append_comparison_row(paid_map.get(k), mim_map.get(k), k, k in ppe_keys)

        def _comparison_sort_key(c: dict) -> tuple:
            return (
                0
                if c["any_mismatch"]
                else (1 if c.get("mim_no_paid_zero_mim_qty") else 2),
                c["external_timesheet_id"],
                c["pay_type"],
                (c.get("week_ended_iso") or ""),
            )

        comparisons.sort(key=_comparison_sort_key)
        return comparisons, tolerance


def _comparison_to_jsonable(c: dict) -> dict:
    out: dict = {}
    for k, v in c.items():
        if isinstance(v, Decimal):
            out[k] = float(v)
        else:
            out[k] = v
    return out


_CROSSREF_PER_PAGE_ALLOWED = frozenset({25, 50, 100})


def _slice_crossref_page(full: list, page: int, per_page: int) -> tuple[list, int, int, int, int, int]:
    """Paginate snapshot rows. Returns (slice, row_count, total_pages, page, per_page, range_start, range_end)."""
    if per_page not in _CROSSREF_PER_PAGE_ALLOWED:
        per_page = 25
    row_count = len(full)
    if row_count == 0:
        return [], 0, 1, 1, per_page, 0, 0
    total_pages = (row_count + per_page - 1) // per_page
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    chunk = full[start : start + per_page]
    range_start = start + 1
    range_end = start + len(chunk)
    return chunk, row_count, total_pages, page, per_page, range_start, range_end


def _crossref_row_key_canonical(c: dict, alias_map: dict[str, str]) -> str:
    badge, pt = _invoice_crossref_key_from_comparison(c, alias_map)
    parts = [badge, pt]
    if c.get("shoals_hours_only"):
        parts.append((str(c.get("week_ended_iso") or "")).strip()[:10])
    return "\x1f".join(parts)


def _crossref_row_key_to_b64(canonical: str) -> str:
    return base64.urlsafe_b64encode(canonical.encode("utf-8")).decode("ascii").rstrip("=")


def _crossref_row_key_from_b64(token: str) -> str:
    t = token.strip()
    pad = "=" * ((4 - len(t) % 4) % 4)
    return base64.urlsafe_b64decode((t + pad).encode("ascii")).decode("utf-8")


def _attach_crossref_notes(s, comparisons: list[dict], alias_map: dict[str, str]) -> None:
    if not comparisons:
        return
    keys: list[str] = []
    for c in comparisons:
        canon = _crossref_row_key_canonical(c, alias_map)
        c["crossref_row_key_b64"] = _crossref_row_key_to_b64(canon)
        keys.append(canon)
    rows = s.execute(select(InvoiceCrossrefRowNote).where(InvoiceCrossrefRowNote.row_key.in_(keys))).scalars().all()
    by_key = {r.row_key: r for r in rows}
    for c, canon in zip(comparisons, keys):
        r = by_key.get(canon)
        c["crossref_note"] = (r.note if r else "") or ""
        c["crossref_solved"] = bool(r.solved) if r else False


def _delete_crossref_snapshot(s) -> None:
    s.execute(delete(InvoiceCrossrefSnapshot))


def _repair_crossref_snapshot_week_display(rows: list[dict]) -> None:
    """Backfill week columns from week_ended_iso when older snapshots omit display strings."""
    for row in rows:
        if row.get("is_ppe_row"):
            continue
        wk_iso = (str(row.get("week_ended_iso") or "")).strip()
        if not wk_iso or len(wk_iso) < 10:
            continue
        try:
            ds = date.fromisoformat(wk_iso[:10]).strftime("%m/%d/%Y")
        except ValueError:
            continue
        if not (str(row.get("week_ended") or "")).strip():
            row["week_ended"] = ds
        if row.get("has_paid") and not (str(row.get("paid_week_ended") or "")).strip():
            row["paid_week_ended"] = ds
        if row.get("has_mim") and not (str(row.get("mim_week_ended") or "")).strip():
            row["mim_week_ended"] = ds


def _refresh_invoice_crossref_snapshot() -> None:
    """Rebuild cross-reference snapshot from current DB imports (same as Calculate)."""
    with session_scope() as s:
        paid_billed_batch = (
            s.execute(
                select(ImportBatch)
                .where(ImportBatch.note == INVOICE_NOTE_PAID_BILLED)
                .order_by(desc(ImportBatch.imported_at))
                .limit(1)
            )
            .scalars()
            .first()
        )
        paid_id = paid_billed_batch.id if paid_billed_batch else None
        mim_ids, cr_layout = _crossref_line_batches_and_layout(s)

    comparisons, tolerance = _build_invoice_comparison(paid_id, mim_ids, crossref_layout=cr_layout)
    payload = json.dumps([_comparison_to_jsonable(c) for c in comparisons])
    with session_scope() as s:
        _delete_crossref_snapshot(s)
        s.add(InvoiceCrossrefSnapshot(payload_json=payload, tolerance_value=str(tolerance)))
        s.commit()


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    with session_scope() as s:
        imports = s.execute(select(ImportBatch).order_by(desc(ImportBatch.imported_at)).limit(10)).scalars().all()
        people_count = s.execute(select(func.count(Person.id))).scalar_one()
        attendance_count = s.execute(select(func.count(Attendance.id))).scalar_one()
    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "imports": imports,
            "people_count": people_count,
            "attendance_count": attendance_count,
        },
    )


@app.get("/home", response_class=HTMLResponse)
def home_page(request: Request):
    current_user = request.state.current_user
    can_view_projects = bool(getattr(current_user, "can_view_projects", False))
    can_manage_projects = bool(getattr(current_user, "can_manage_projects", False))
    can_view_all_projects = bool(getattr(current_user, "can_view_all_projects", False))
    can_view_people = bool(getattr(current_user, "can_view_people", False))
    can_view_imports = bool(getattr(current_user, "can_view_imports", False))
    can_view_invoices = bool(getattr(current_user, "can_view_invoices", False))
    if not any((can_view_projects, can_view_people, can_view_imports, can_view_invoices)):
        raise HTTPException(status_code=403, detail="You do not have access to Home.")

    today = date.today()
    week_end = today + timedelta(days=7)
    projects: list[Project] = []
    project_cards: list[dict[str, object]] = []
    task_items: list[dict[str, object]] = []
    people_workload: list[dict[str, object]] = []
    recent_notes: list[dict[str, object]] = []
    recent_imports: list[ImportBatch] = []
    invoice_snapshot_at: datetime | None = None
    invoice_exception_count = 0
    invoice_total_rows = 0
    people_count = 0
    active_people_count = 0

    with session_scope() as s:
        if can_view_projects:
            project_manager_person_id = getattr(current_user, "project_manager_person_id", None)
            user_person_id = getattr(current_user, "person_id", None)
            if can_view_all_projects or can_manage_projects:
                projects = s.execute(select(Project).order_by(desc(Project.updated_at), Project.name.asc())).scalars().all()
            elif project_manager_person_id or user_person_id:
                person = s.get(Person, project_manager_person_id or user_person_id)
                if person:
                    projects = (
                        s.execute(
                            select(Project)
                            .join(ProjectParticipant, Project.id == ProjectParticipant.project_id)
                            .where(ProjectParticipant.display_name == person.full_name)
                            .order_by(desc(Project.updated_at), Project.name.asc())
                        )
                        .scalars()
                        .all()
                    )

            project_ids = [project.id for project in projects]
            tasks_by_project: dict[int, list[dict[str, object]]] = {project_id: [] for project_id in project_ids}
            if project_ids:
                task_rows = (
                    s.execute(
                        select(ProjectTask, Project, ProjectParticipant.display_name)
                        .join(Project, ProjectTask.project_id == Project.id)
                        .outerjoin(ProjectParticipant, ProjectTask.owner_participant_id == ProjectParticipant.id)
                        .where(ProjectTask.project_id.in_(project_ids))
                        .order_by(ProjectTask.due_date.asc(), desc(ProjectTask.updated_at), ProjectTask.title.asc())
                    )
                    .all()
                )
                for task, project, owner_name in task_rows:
                    is_done = (task.status or "") == "done"
                    item = {
                        "id": task.id,
                        "project_id": project.id,
                        "project_name": project.name,
                        "title": task.title,
                        "href": f"/projects/{project.id}/tasks/{task.id}",
                        "status": task.status,
                        "status_label": _project_task_label(_PROJECT_TASK_STATUS_OPTIONS, task.status),
                        "priority": task.priority,
                        "priority_label": _project_task_label(_PROJECT_TASK_PRIORITY_OPTIONS, task.priority),
                        "risk_level": task.risk_level,
                        "risk_label": _project_task_label(_PROJECT_TASK_RISK_OPTIONS, task.risk_level),
                        "due_date": task.due_date,
                        "percent_complete": task.percent_complete or 0,
                        "owner_name": owner_name or "",
                        "is_done": is_done,
                        "is_overdue": bool(task.due_date and task.due_date < today and not is_done),
                        "is_due_this_week": bool(task.due_date and today <= task.due_date <= week_end and not is_done),
                        "is_stuck": (task.status or "") == "blocked",
                        "is_high_risk": (task.risk_level or "") in {"high", "critical"},
                        "updated_at": task.updated_at,
                    }
                    tasks_by_project.setdefault(project.id, []).append(item)
                    task_items.append(item)

                note_rows = (
                    s.execute(
                        select(ProjectTaskNote, ProjectTask, Project)
                        .join(ProjectTask, ProjectTaskNote.task_id == ProjectTask.id)
                        .join(Project, ProjectTask.project_id == Project.id)
                        .where(Project.id.in_(project_ids), ProjectTaskNote.is_section.is_(False))
                        .order_by(desc(ProjectTaskNote.updated_at), desc(ProjectTaskNote.created_at))
                        .limit(6)
                    )
                    .all()
                )
                for note, task, project in note_rows:
                    title = (note.title or "").strip() or (note.content or "").strip()[:80] or "Task note"
                    recent_notes.append(
                        {
                            "title": title,
                            "task_title": task.title,
                            "project_name": project.name,
                            "href": f"/projects/{project.id}/tasks/{task.id}/notes",
                            "updated_at": note.updated_at,
                            "is_pinned": bool(note.is_pinned),
                        }
                    )

            for project in projects:
                rows = tasks_by_project.get(project.id, [])
                total = len(rows)
                done = sum(1 for task in rows if task["is_done"])
                overdue = sum(1 for task in rows if task["is_overdue"])
                stuck = sum(1 for task in rows if task["is_stuck"])
                high_risk = sum(1 for task in rows if task["is_high_risk"])
                missing_owner = sum(1 for task in rows if not task["owner_name"] and not task["is_done"])
                avg_completion = round(sum(int(task["percent_complete"]) for task in rows) / total) if total else 0
                next_due = next((task for task in rows if task["due_date"] and not task["is_done"]), None)
                project_cards.append(
                    {
                        "id": project.id,
                        "name": project.name,
                        "client_name": project.client_name or "",
                        "status": project.status or "planning",
                        "notes": project.notes or "",
                        "href": f"/projects/{project.id}",
                        "timeline_href": f"/projects/{project.id}?tab=timeline",
                        "views_href": f"/projects/{project.id}?tab=board",
                        "total_tasks": total,
                        "open_tasks": total - done,
                        "done_tasks": done,
                        "overdue_tasks": overdue,
                        "stuck_tasks": stuck,
                        "high_risk_tasks": high_risk,
                        "missing_owner_tasks": missing_owner,
                        "avg_completion": avg_completion,
                        "next_due": next_due,
                    }
                )

            workload_by_owner: dict[str, dict[str, object]] = {}
            for task in task_items:
                if task["is_done"]:
                    continue
                owner = str(task["owner_name"] or "Unassigned")
                row = workload_by_owner.setdefault(
                    owner,
                    {"name": owner, "open_tasks": 0, "stuck_tasks": 0, "high_risk_tasks": 0, "overdue_tasks": 0, "avg_completion": 0, "_completion_total": 0},
                )
                row["open_tasks"] = int(row["open_tasks"]) + 1
                row["stuck_tasks"] = int(row["stuck_tasks"]) + (1 if task["is_stuck"] else 0)
                row["high_risk_tasks"] = int(row["high_risk_tasks"]) + (1 if task["is_high_risk"] else 0)
                row["overdue_tasks"] = int(row["overdue_tasks"]) + (1 if task["is_overdue"] else 0)
                row["_completion_total"] = int(row["_completion_total"]) + int(task["percent_complete"])
            for row in workload_by_owner.values():
                open_tasks = max(int(row["open_tasks"]), 1)
                row["avg_completion"] = round(int(row["_completion_total"]) / open_tasks)
                row.pop("_completion_total", None)
            people_workload = sorted(
                workload_by_owner.values(),
                key=lambda row: (int(row["overdue_tasks"]), int(row["stuck_tasks"]), int(row["open_tasks"])),
                reverse=True,
            )[:8]

        if can_view_people:
            people_count = s.execute(select(func.count(Person.id))).scalar_one()
            active_people_count = s.execute(
                select(func.count(Person.id)).where((Person.talent_status.is_(None)) | (Person.talent_status != "terminated"))
            ).scalar_one()

        if can_view_imports:
            recent_imports = s.execute(select(ImportBatch).order_by(desc(ImportBatch.imported_at)).limit(5)).scalars().all()

        if can_view_invoices:
            snap_row = (
                s.execute(select(InvoiceCrossrefSnapshot).order_by(desc(InvoiceCrossrefSnapshot.created_at)).limit(1))
                .scalars()
                .first()
            )
            if snap_row:
                invoice_snapshot_at = snap_row.created_at
                try:
                    invoice_rows = json.loads(snap_row.payload_json)
                except json.JSONDecodeError:
                    invoice_rows = []
                invoice_total_rows = len(invoice_rows)
                invoice_exception_count = sum(1 for row in invoice_rows if row.get("any_mismatch"))

    open_tasks = [task for task in task_items if not task["is_done"]]
    overdue_tasks = [task for task in open_tasks if task["is_overdue"]]
    stuck_tasks = [task for task in open_tasks if task["is_stuck"]]
    high_risk_tasks = [task for task in open_tasks if task["is_high_risk"]]
    missing_owner_tasks = [task for task in open_tasks if not task["owner_name"]]
    due_this_week_tasks = [task for task in open_tasks if task["is_due_this_week"]]
    recent_task_updates = sorted(task_items, key=lambda task: task["updated_at"] or datetime.min, reverse=True)[:6]
    portfolio_avg_completion = (
        round(sum(int(card["avg_completion"]) for card in project_cards) / len(project_cards)) if project_cards else 0
    )
    activity_feed: list[dict[str, object]] = []
    for task in recent_task_updates[:4]:
        activity_feed.append(
            {
                "kind": "Task",
                "title": task["title"],
                "subtitle": f"{task['project_name']} / {task['status_label']}",
                "href": task["href"],
                "when": task["updated_at"],
            }
        )
    for note in recent_notes[:3]:
        activity_feed.append(
            {
                "kind": "Note",
                "title": note["title"],
                "subtitle": f"{note['project_name']} / {note['task_title']}",
                "href": note["href"],
                "when": note["updated_at"],
            }
        )
    for batch in recent_imports[:2]:
        activity_feed.append(
            {
                "kind": "Import",
                "title": batch.filename,
                "subtitle": f"{batch.rows_total} rows",
                "href": "/imports",
                "when": batch.imported_at,
            }
        )
    activity_feed.sort(key=lambda item: item["when"] or datetime.min, reverse=True)

    return templates.TemplateResponse(
        request,
        "home.html",
        {
            "title": "Home",
            "today": today,
            "can_view_projects": can_view_projects,
            "can_manage_projects": can_manage_projects,
            "can_view_people": can_view_people,
            "can_view_imports": can_view_imports,
            "can_view_invoices": can_view_invoices,
            "project_cards": project_cards,
            "projects_count": len(project_cards),
            "active_projects_count": sum(1 for card in project_cards if card["status"] == "active"),
            "portfolio_avg_completion": portfolio_avg_completion,
            "open_tasks_count": len(open_tasks),
            "overdue_tasks": overdue_tasks[:6],
            "overdue_tasks_count": len(overdue_tasks),
            "stuck_tasks": stuck_tasks[:6],
            "stuck_tasks_count": len(stuck_tasks),
            "high_risk_tasks_count": len(high_risk_tasks),
            "missing_owner_tasks_count": len(missing_owner_tasks),
            "due_this_week_tasks": sorted(due_this_week_tasks, key=lambda task: task["due_date"] or today)[:6],
            "people_count": people_count,
            "active_people_count": active_people_count,
            "people_workload": people_workload,
            "recent_imports": recent_imports,
            "invoice_exception_count": invoice_exception_count,
            "invoice_total_rows": invoice_total_rows,
            "invoice_snapshot_at": invoice_snapshot_at,
            "activity_feed": activity_feed[:8],
        },
    )


def _project_task_label(options: tuple[tuple[str, str], ...], value: str | None) -> str:
    raw = (value or "").strip()
    for option_value, label in options:
        if option_value == raw:
            return label
    return raw.replace("_", " ").title() if raw else ""


def _coerce_project_task_percent(raw: str | int | None) -> int:
    try:
        value = int(str(raw if raw is not None else "0").strip() or "0")
    except ValueError:
        value = 0
    return max(0, min(100, value))


def _project_task_percent_for_status(status: str, current_percent: int | None) -> int:
    percent = _coerce_project_task_percent(current_percent)
    if status == "done":
        return 100
    if status == "in_progress" and percent == 0:
        return 10
    return min(percent, 99)


def _coerce_project_department_color(raw: str | None) -> str:
    color = (raw or "slate").strip().lower()
    return color if color in _PROJECT_DEPARTMENT_COLORS else "slate"


def _parse_project_id_list(values: object) -> list[int]:
    raw_values = values if isinstance(values, list) else [values]
    parsed: list[int] = []
    for raw in raw_values:
        for piece in str(raw or "").replace(",", " ").split():
            try:
                value = int(piece)
            except ValueError:
                continue
            if value not in parsed:
                parsed.append(value)
    return parsed


def _sync_project_task_supporters(s, task: ProjectTask, supporter_ids: list[int]) -> None:
    valid_ids = set(
        s.execute(
            select(ProjectParticipant.id).where(
                ProjectParticipant.project_id == task.project_id,
                ProjectParticipant.id.in_(supporter_ids),
            )
        ).scalars().all()
    )
    existing_links = list(task.supporter_links)
    for link in existing_links:
        if link.participant_id not in valid_ids:
            s.delete(link)
    existing_ids = {link.participant_id for link in existing_links if link.participant_id in valid_ids}
    for participant_id in supporter_ids:
        if participant_id in valid_ids and participant_id not in existing_ids:
            s.add(ProjectTaskSupporter(task=task, participant_id=participant_id))
            existing_ids.add(participant_id)


def _project_task_redirect(project_id: int, tab: str, message_key: str, message: str) -> RedirectResponse:
    return _redirect(f"/projects/{project_id}?tab={quote(tab)}&{message_key}=" + quote(message))


def _user_can_access_project(s, project: Project, current_user: object) -> bool:
    if bool(getattr(current_user, "can_manage_projects", False)) or bool(
        getattr(current_user, "can_view_all_projects", False)
    ):
        return True
    person_id = getattr(current_user, "project_manager_person_id", None) or getattr(current_user, "person_id", None)
    if not person_id:
        return False
    person = s.get(Person, person_id)
    if not person:
        return False
    participant = (
        s.execute(
            select(ProjectParticipant).where(
                ProjectParticipant.project_id == project.id,
                ProjectParticipant.display_name == person.full_name,
            )
        )
        .scalars()
        .first()
    )
    return participant is not None


def _task_detail_redirect(project_id: int, task_id: int, message_key: str, message: str) -> RedirectResponse:
    return _redirect(f"/projects/{project_id}/tasks/{task_id}?{message_key}=" + quote(message))


@app.get("/projects", response_class=HTMLResponse)
def projects_page(request: Request, created: str | None = None, error: str | None = None):
    current_user = request.state.current_user
    can_manage_projects = bool(getattr(current_user, "can_manage_projects", False))
    can_view_all_projects = bool(getattr(current_user, "can_view_all_projects", False))
    project_manager_person_id = getattr(current_user, "project_manager_person_id", None)
    user_person_id = getattr(current_user, "person_id", None)

    with session_scope() as s:
        if can_view_all_projects or can_manage_projects:
            # Project managers and those with projects_manage can see all projects
            projects = s.execute(select(Project).order_by(desc(Project.created_at), Project.name.asc())).scalars().all()
        elif project_manager_person_id:
            # Project managers see projects where they are a participant (by matching person name)
            person = s.get(Person, project_manager_person_id)
            if person:
                projects = (
                    s.execute(
                        select(Project)
                        .join(ProjectParticipant, Project.id == ProjectParticipant.project_id)
                        .where(ProjectParticipant.display_name == person.full_name)
                        .order_by(desc(Project.created_at), Project.name.asc())
                    )
                    .scalars()
                    .all()
                )
            else:
                projects = []
        elif user_person_id:
            # Regular users see projects they are assigned to via ProjectParticipant
            person = s.get(Person, user_person_id)
            if person:
                projects = (
                    s.execute(
                        select(Project)
                        .join(ProjectParticipant, Project.id == ProjectParticipant.project_id)
                        .where(ProjectParticipant.display_name == person.full_name)
                        .order_by(desc(Project.created_at), Project.name.asc())
                    )
                    .scalars()
                    .all()
                )
            else:
                projects = []
        else:
            projects = []

    active_count = sum(1 for project in projects if (project.status or "").strip().lower() == "active")
    project_cards: list[dict[str, object]] = []
    project_ids = [project.id for project in projects]
    task_rows_by_project: dict[int, list[ProjectTask]] = {project_id: [] for project_id in project_ids}
    if project_ids:
        with session_scope() as s:
            task_rows = (
                s.execute(
                    select(ProjectTask)
                    .where(ProjectTask.project_id.in_(project_ids))
                    .order_by(ProjectTask.due_date.asc(), ProjectTask.title.asc())
                )
                .scalars()
                .all()
            )
        for task in task_rows:
            task_rows_by_project.setdefault(task.project_id, []).append(task)
    today = date.today()
    for project in projects:
        rows = task_rows_by_project.get(project.id, [])
        total = len(rows)
        done = sum(1 for task in rows if (task.status or "") == "done")
        overdue = sum(1 for task in rows if task.due_date and task.due_date < today and (task.status or "") != "done")
        blocked = sum(1 for task in rows if (task.status or "") == "blocked")
        avg_completion = round(sum((task.percent_complete or 0) for task in rows) / total) if total else 0
        next_due = next((task for task in rows if task.due_date and (task.status or "") != "done"), None)
        project_cards.append(
            {
                "project": project,
                "total_tasks": total,
                "open_tasks": total - done,
                "done_tasks": done,
                "overdue_tasks": overdue,
                "blocked_tasks": blocked,
                "avg_completion": avg_completion,
                "next_due": next_due,
            }
        )
    portfolio_total_tasks = sum(int(card["total_tasks"]) for card in project_cards)
    portfolio_open_tasks = sum(int(card["open_tasks"]) for card in project_cards)
    portfolio_blocked_tasks = sum(int(card["blocked_tasks"]) for card in project_cards)
    portfolio_overdue_tasks = sum(int(card["overdue_tasks"]) for card in project_cards)
    portfolio_avg_completion = (
        round(sum(int(card["avg_completion"]) for card in project_cards) / len(project_cards)) if project_cards else 0
    )
    return templates.TemplateResponse(
        request,
        "projects.html",
        {
            "title": "Projects",
            "projects": projects,
            "project_cards": project_cards,
            "portfolio_total_tasks": portfolio_total_tasks,
            "portfolio_open_tasks": portfolio_open_tasks,
            "portfolio_blocked_tasks": portfolio_blocked_tasks,
            "portfolio_overdue_tasks": portfolio_overdue_tasks,
            "portfolio_avg_completion": portfolio_avg_completion,
            "project_status_options": _PROJECT_STATUS_OPTIONS,
            "created": created == "1",
            "error": (error or "").strip() or None,
            "projects_count": len(projects),
            "active_projects_count": active_count,
            "can_manage_projects": can_manage_projects,
        },
    )


@app.get("/projects/{project_id}", response_class=HTMLResponse)
def project_detail_page(
    request: Request,
    project_id: int,
    tab: str | None = None,
    notice: str | None = None,
    error: str | None = None,
):
    current_user = request.state.current_user
    can_manage_projects = bool(getattr(current_user, "can_manage_projects", False))
    can_view_all_projects = bool(getattr(current_user, "can_view_all_projects", False))
    project_manager_person_id = getattr(current_user, "project_manager_person_id", None)
    user_person_id = getattr(current_user, "person_id", None)

    with session_scope() as s:
        project = s.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        # Check if user has access to this project
        has_access = False
        if can_view_all_projects or can_manage_projects:
            has_access = True
        elif project_manager_person_id:
            # Check if user is a project participant (via display_name matching person's full_name)
            person = s.get(Person, project_manager_person_id)
            if person:
                participant = (
                    s.execute(
                        select(ProjectParticipant).where(
                            ProjectParticipant.project_id == project_id,
                            ProjectParticipant.display_name == person.full_name
                        )
                    )
                    .scalars()
                    .first()
                )
                has_access = participant is not None
        elif user_person_id:
            # Check if user is a project participant
            person = s.get(Person, user_person_id)
            if person:
                participant = (
                    s.execute(
                        select(ProjectParticipant).where(
                            ProjectParticipant.project_id == project_id,
                            ProjectParticipant.display_name == person.full_name
                        )
                    )
                    .scalars()
                    .first()
                )
                has_access = participant is not None

        if not has_access:
            raise HTTPException(status_code=403, detail="You do not have access to this project.")

        participant_rows = (
            s.execute(
                select(ProjectParticipant)
                .where(ProjectParticipant.project_id == project_id)
                .order_by(ProjectParticipant.display_name.asc(), ProjectParticipant.id.asc())
            )
            .scalars()
            .all()
        )

        department_rows = (
            s.execute(
                select(ProjectDepartment)
                .where(ProjectDepartment.project_id == project_id)
                .order_by(ProjectDepartment.sort_order.asc(), ProjectDepartment.name.asc(), ProjectDepartment.id.asc())
            )
            .scalars()
            .all()
        )

        task_rows = (
            s.execute(
                select(ProjectTask)
                .where(ProjectTask.project_id == project_id)
                .order_by(ProjectTask.sort_order.asc(), ProjectTask.due_date.asc(), ProjectTask.title.asc())
            )
            .scalars()
            .all()
        )
        participant_name_by_id = {row.id: row.display_name for row in participant_rows}
        department_by_id = {row.id: row for row in department_rows}
        supporter_rows = (
            s.execute(
                select(ProjectTaskSupporter)
                .join(ProjectTask, ProjectTask.id == ProjectTaskSupporter.task_id)
                .where(ProjectTask.project_id == project_id)
            )
            .scalars()
            .all()
        )
        supporter_ids_by_task: dict[int, list[int]] = defaultdict(list)
        for link in supporter_rows:
            supporter_ids_by_task[link.task_id].append(link.participant_id)
        task_cards: list[dict[str, object]] = []
        today = date.today()
        for task in task_rows:
            is_done = (task.status or "") == "done"
            is_overdue = bool(task.due_date and task.due_date < today and not is_done)
            is_due_soon = bool(task.due_date and today <= task.due_date <= today + timedelta(days=7) and not is_done)
            department = department_by_id.get(task.department_id) if task.department_id else None
            supporter_ids = supporter_ids_by_task.get(task.id, [])
            supporter_names = [participant_name_by_id.get(pid, "") for pid in supporter_ids]
            supporter_names = [name for name in supporter_names if name]
            task_cards.append(
                {
                    "id": task.id,
                    "title": task.title,
                    "description": task.description,
                    "department_id": task.department_id,
                    "department_name": department.name if department else "",
                    "department_color": department.color if department else "slate",
                    "status": task.status,
                    "status_label": _project_task_label(_PROJECT_TASK_STATUS_OPTIONS, task.status),
                    "priority": task.priority,
                    "priority_label": _project_task_label(_PROJECT_TASK_PRIORITY_OPTIONS, task.priority),
                    "start_date": task.start_date,
                    "due_date": task.due_date,
                    "due_offset": (task.due_date - today).days if task.due_date else None,
                    "percent_complete": task.percent_complete or 0,
                    "dependency_note": task.dependency_note,
                    "risk_level": task.risk_level,
                    "risk_label": _project_task_label(_PROJECT_TASK_RISK_OPTIONS, task.risk_level),
                    "risk_notes": task.risk_notes,
                    "owner_participant_id": task.owner_participant_id,
                    "owner_name": participant_name_by_id.get(task.owner_participant_id) if task.owner_participant_id else "",
                    "supporter_ids": supporter_ids,
                    "supporter_names": supporter_names,
                    "supporter_summary": ", ".join(supporter_names),
                    "is_overdue": is_overdue,
                    "is_due_soon": is_due_soon,
                    "updated_at": task.updated_at,
                }
            )

        task_total = len(task_cards)
        task_done = sum(1 for task in task_cards if task["status"] == "done")
        task_blocked = sum(1 for task in task_cards if task["status"] == "blocked")
        task_overdue = sum(1 for task in task_cards if task["is_overdue"])
        task_due_soon = sum(1 for task in task_cards if task["is_due_soon"])
        avg_completion = round(sum(int(task["percent_complete"]) for task in task_cards) / task_total) if task_total else 0
        task_stats = {
            "total": task_total,
            "done": task_done,
            "active": task_total - task_done,
            "blocked": task_blocked,
            "overdue": task_overdue,
            "due_soon": task_due_soon,
            "avg_completion": avg_completion,
        }
        task_columns = [
            {
                "value": value,
                "label": label,
                "tasks": [task for task in task_cards if task["status"] == value],
                "percent": round((sum(1 for task in task_cards if task["status"] == value) * 100) / task_total) if task_total else 0,
            }
            for value, label in _PROJECT_TASK_STATUS_OPTIONS
        ]
        chart_classes = {
            "backlog": "bg-slate-400",
            "todo": "bg-amber-500",
            "in_progress": "bg-blue-500",
            "blocked": "bg-red-500",
            "done": "bg-emerald-500",
            "critical": "bg-rose-500",
            "high": "bg-orange-500",
            "medium": "bg-indigo-500",
            "low": "bg-slate-400",
        }
        priority_chart = [
            {
                "value": value,
                "label": label,
                "count": sum(1 for task in task_cards if task["priority"] == value),
                "percent": round((sum(1 for task in task_cards if task["priority"] == value) * 100) / task_total) if task_total else 0,
                "bar_class": chart_classes.get(value, "bg-slate-400"),
            }
            for value, label in _PROJECT_TASK_PRIORITY_OPTIONS
        ]
        risk_chart = [
            {
                "value": value,
                "label": label,
                "count": sum(1 for task in task_cards if task["risk_level"] == value),
                "percent": round((sum(1 for task in task_cards if task["risk_level"] == value) * 100) / task_total) if task_total else 0,
                "bar_class": chart_classes.get(value, "bg-slate-400"),
            }
            for value, label in _PROJECT_TASK_RISK_OPTIONS
        ]
        for column in task_columns:
            column["bar_class"] = chart_classes.get(str(column["value"]), "bg-slate-400")
        owner_load: list[dict[str, object]] = []
        max_owner_open = 1
        for participant in participant_rows:
            owned = [task for task in task_cards if task["owner_participant_id"] == participant.id]
            open_owned = [task for task in owned if task["status"] != "done"]
            max_owner_open = max(max_owner_open, len(open_owned))
            owner_load.append(
                {
                    "id": participant.id,
                    "name": participant.display_name,
                    "open": len(open_owned),
                    "done": sum(1 for task in owned if task["status"] == "done"),
                    "overdue": sum(1 for task in open_owned if task["is_overdue"]),
                }
            )
        unassigned_open = [task for task in task_cards if not task["owner_participant_id"] and task["status"] != "done"]
        if unassigned_open:
            max_owner_open = max(max_owner_open, len(unassigned_open))
            owner_load.append({"id": None, "name": "Unassigned", "open": len(unassigned_open), "done": 0, "overdue": sum(1 for task in unassigned_open if task["is_overdue"])})
        for owner in owner_load:
            owner["percent"] = round((int(owner["open"]) * 100) / max_owner_open) if max_owner_open else 0

        department_load: list[dict[str, object]] = []
        max_department_open = 1
        for department in department_rows:
            rows = [task for task in task_cards if task["department_id"] == department.id]
            open_rows = [task for task in rows if task["status"] != "done"]
            max_department_open = max(max_department_open, len(open_rows))
            total_rows = len(rows)
            department_load.append(
                {
                    "id": department.id,
                    "name": department.name,
                    "color": department.color,
                    "total": total_rows,
                    "open": len(open_rows),
                    "done": sum(1 for task in rows if task["status"] == "done"),
                    "blocked": sum(1 for task in rows if task["status"] == "blocked"),
                    "overdue": sum(1 for task in open_rows if task["is_overdue"]),
                    "avg_completion": round(sum(int(task["percent_complete"]) for task in rows) / total_rows) if total_rows else 0,
                }
            )
        unassigned_department_rows = [task for task in task_cards if not task["department_id"]]
        if unassigned_department_rows:
            open_rows = [task for task in unassigned_department_rows if task["status"] != "done"]
            max_department_open = max(max_department_open, len(open_rows))
            total_rows = len(unassigned_department_rows)
            department_load.append(
                {
                    "id": None,
                    "name": "Unassigned",
                    "color": "slate",
                    "total": total_rows,
                    "open": len(open_rows),
                    "done": sum(1 for task in unassigned_department_rows if task["status"] == "done"),
                    "blocked": sum(1 for task in unassigned_department_rows if task["status"] == "blocked"),
                    "overdue": sum(1 for task in open_rows if task["is_overdue"]),
                    "avg_completion": round(sum(int(task["percent_complete"]) for task in unassigned_department_rows) / total_rows) if total_rows else 0,
                }
            )
        for department in department_load:
            department["percent"] = round((int(department["open"]) * 100) / max_department_open) if max_department_open else 0

        dated_tasks = [task for task in task_cards if task["start_date"] or task["due_date"]]
        project_timeline_start = min((task["start_date"] or task["due_date"] for task in dated_tasks), default=today)
        project_timeline_end = max((task["due_date"] or task["start_date"] for task in dated_tasks), default=today + timedelta(days=14))
        if project_timeline_end < project_timeline_start:
            project_timeline_start, project_timeline_end = project_timeline_end, project_timeline_start
        max_timeline_days = 183
        timeline_start = project_timeline_start
        timeline_end = min(project_timeline_end, timeline_start + timedelta(days=max_timeline_days - 1))
        gantt_is_capped = project_timeline_end > timeline_end
        timeline_days = max((timeline_end - timeline_start).days + 1, 1)
        gantt_grid_start = timeline_start - timedelta(days=timeline_start.weekday())
        gantt_grid_end = timeline_end + timedelta(days=(4 - timeline_end.weekday()) % 7)
        gantt_business_days: list[dict[str, object]] = []
        cursor = gantt_grid_start
        while cursor <= gantt_grid_end:
            if cursor.weekday() < 5:
                gantt_business_days.append(
                    {
                        "date": cursor,
                        "label": cursor.strftime("%a"),
                        "day": cursor.day,
                        "month": cursor.strftime("%b"),
                        "date_short": f"{cursor.month}/{cursor.day}",
                        "is_today": cursor == today,
                    }
                )
            cursor += timedelta(days=1)
        if not gantt_business_days:
            gantt_business_days = [{"date": today, "label": today.strftime("%a"), "day": today.day, "month": today.strftime("%b"), "date_short": f"{today.month}/{today.day}", "is_today": True}]
        business_index_by_date = {day["date"]: index + 1 for index, day in enumerate(gantt_business_days)}
        gantt_today_date = today if today in business_index_by_date else None
        if gantt_today_date is None and timeline_start <= today <= timeline_end:
            previous_business_days = [day["date"] for day in gantt_business_days if day["date"] <= today]
            if previous_business_days:
                gantt_today_date = max(previous_business_days)
        if gantt_today_date is not None:
            for day in gantt_business_days:
                day["is_today"] = day["date"] == gantt_today_date
        gantt_today_grid_start = business_index_by_date.get(gantt_today_date) if gantt_today_date is not None else None
        visible_business_start = gantt_business_days[0]["date"]
        visible_business_end = gantt_business_days[-1]["date"]
        visible_business_columns = len(gantt_business_days)

        def _business_span_for_dates(start_value: date, end_value: date) -> tuple[int, int, bool]:
            if end_value < start_value:
                start_value, end_value = end_value, start_value
            is_clipped = start_value < visible_business_start or end_value > visible_business_end
            if end_value < visible_business_start or start_value > visible_business_end:
                return 1, 0, True
            start_value = max(start_value, visible_business_start)
            end_value = min(end_value, visible_business_end)
            indexes = [
                business_index_by_date[day["date"]]
                for day in gantt_business_days
                if start_value <= day["date"] <= end_value
            ]
            if not indexes:
                nearest = min(
                    business_index_by_date.items(),
                    key=lambda item: abs((item[0] - start_value).days),
                )[1]
                return nearest, 1, is_clipped
            grid_start = max(1, min(min(indexes), visible_business_columns))
            grid_span = max(1, max(indexes) - min(indexes) + 1)
            grid_span = min(grid_span, visible_business_columns - grid_start + 1)
            return grid_start, grid_span, is_clipped

        gantt_weeks: list[dict[str, object]] = []
        current_week: dict[str, object] | None = None
        for index, day in enumerate(gantt_business_days, start=1):
            iso_year, iso_week, _ = day["date"].isocalendar()
            key = f"{iso_year}-W{iso_week:02d}"
            if current_week is None or current_week["key"] != key:
                week_start = day["date"]
                current_week = {
                    "key": key,
                    "label": f"Week of {week_start.month}/{week_start.day}/{str(week_start.year)[2:]}",
                    "week_number": iso_week,
                    "start": index,
                    "span": 1,
                }
                gantt_weeks.append(current_week)
            else:
                current_week["span"] = int(current_week["span"]) + 1

        for task in task_cards:
            start_value = task["start_date"] or task["due_date"]
            end_value = task["due_date"] or task["start_date"]
            if not start_value or not end_value:
                task["has_timeline"] = False
                task["gantt_left"] = 0
                task["gantt_width"] = 0
                task["gantt_days"] = 0
                continue
            if end_value < start_value:
                start_value, end_value = end_value, start_value
            task_days = max((end_value - start_value).days + 1, 1)
            grid_start, grid_span, is_clipped = _business_span_for_dates(start_value, end_value)
            visible_start = max(start_value, timeline_start)
            visible_end = min(end_value, timeline_end)
            visible_days = max((visible_end - visible_start).days + 1, 1)
            gantt_left = round(((visible_start - timeline_start).days * 100) / timeline_days, 2)
            gantt_width = max(round((visible_days * 100) / timeline_days, 2), 2)
            gantt_width = min(gantt_width, max(100 - gantt_left, 0))
            task["has_timeline"] = True
            task["timeline_start_label"] = start_value
            task["timeline_end_label"] = end_value
            task["gantt_left"] = gantt_left
            task["gantt_width"] = gantt_width
            task["gantt_days"] = task_days
            task["gantt_grid_start"] = grid_start
            task["gantt_grid_span"] = grid_span
            task["is_gantt_clipped"] = is_clipped
            if not grid_span:
                task["has_timeline"] = False
                task["gantt_left"] = 0
                task["gantt_width"] = 0

        gantt_departments: list[dict[str, object]] = []
        grouped_department_ids: set[int | None] = set()
        gantt_source_departments: list[dict[str, object]] = [
            {"id": department.id, "name": department.name, "color": department.color}
            for department in department_rows
        ]
        if any(not task["department_id"] for task in task_cards):
            gantt_source_departments.append({"id": None, "name": "Cross-functional", "color": "slate"})
        for department in gantt_source_departments:
            department_id = department["id"]
            department_tasks = [task for task in task_cards if task["department_id"] == department_id]
            grouped_department_ids.add(department_id)
            dated_department_tasks = [task for task in department_tasks if task["has_timeline"]]
            if dated_department_tasks:
                group_start = min(task["timeline_start_label"] for task in dated_department_tasks)
                group_end = max(task["timeline_end_label"] for task in dated_department_tasks)
                group_days = max((group_end - group_start).days + 1, 1)
                visible_group_start = max(group_start, timeline_start)
                visible_group_end = min(group_end, timeline_end)
                visible_group_days = max((visible_group_end - visible_group_start).days + 1, 1)
                group_left = round(((visible_group_start - timeline_start).days * 100) / timeline_days, 2)
                group_width = max(round((visible_group_days * 100) / timeline_days, 2), 2)
                group_width = min(group_width, max(100 - group_left, 0))
                group_grid_start, group_grid_span, group_is_clipped = _business_span_for_dates(group_start, group_end)
            else:
                group_start = None
                group_end = None
                group_days = 0
                group_left = 0
                group_width = 0
                group_grid_start = 1
                group_grid_span = 1
                group_is_clipped = False
            total_rows = len(department_tasks)
            done_rows = sum(1 for task in department_tasks if task["status"] == "done")
            open_rows = [task for task in department_tasks if task["status"] != "done"]
            gantt_departments.append(
                {
                    "id": department_id,
                    "key": str(department_id) if department_id is not None else "none",
                    "name": department["name"],
                    "color": department["color"],
                    "tasks": dated_department_tasks,
                    "total": total_rows,
                    "open": len(open_rows),
                    "done": done_rows,
                    "blocked": sum(1 for task in department_tasks if task["status"] == "blocked"),
                    "overdue": sum(1 for task in open_rows if task["is_overdue"]),
                    "avg_completion": round(sum(int(task["percent_complete"]) for task in department_tasks) / total_rows) if total_rows else 0,
                    "timeline_start_label": group_start,
                    "timeline_end_label": group_end,
                    "gantt_left": group_left,
                    "gantt_width": group_width,
                    "gantt_grid_start": group_grid_start,
                    "gantt_grid_span": group_grid_span,
                    "gantt_days": group_days,
                    "is_gantt_clipped": group_is_clipped,
                    "collapsed_default": bool(total_rows and done_rows == total_rows),
                }
            )
        timeline_ticks: list[dict[str, object]] = []
        tick_cursor = timeline_start
        while tick_cursor <= timeline_end:
            timeline_ticks.append(
                {
                    "label": f"{tick_cursor:%b} {tick_cursor.day}",
                    "date": tick_cursor,
                    "left": round(((tick_cursor - timeline_start).days * 100) / timeline_days, 2),
                }
            )
            tick_cursor += timedelta(days=7)
        today_marker_left = None
        if timeline_start <= today <= timeline_end:
            today_marker_left = round(((today - timeline_start).days * 100) / timeline_days, 2)

        upcoming_tasks = [
            task
            for task in sorted(
                task_cards,
                key=lambda row: (row["due_date"] is None, row["due_date"] or date.max, str(row["title"]).lower()),
            )
            if task["status"] != "done"
        ][:5]
        risk_watch = [
            task
            for task in sorted(
                task_cards,
                key=lambda row: (
                    0 if row["is_overdue"] else 1,
                    {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(str(row["risk_level"]), 4),
                    row["due_date"] or date.max,
                ),
            )
            if task["status"] != "done" and (task["is_overdue"] or task["status"] == "blocked" or task["risk_level"] in {"high", "critical"})
        ][:6]
        dependency_watch = [
            task
            for task in sorted(task_cards, key=lambda row: (row["due_date"] is None, row["due_date"] or date.max))
            if task["status"] != "done" and task["dependency_note"]
        ][:6]

        selected_member = None
        raw_tab = (tab or "").strip().lower()
        project_tab = "overview"
        if raw_tab in {"", "home", "overview"}:
            project_tab = "overview"
        elif raw_tab == "list":
            return _redirect(f"/projects/{project_id}?tab=board")
        elif raw_tab in {"dashboard", "board", "timeline", "people"}:
            if raw_tab == "dashboard":
                raw_tab = "overview"
            project_tab = raw_tab
        elif raw_tab == "manager" and can_manage_projects:
            project_tab = "manager"
    return templates.TemplateResponse(
        request,
        "project_detail.html",
        {
            "title": project.name,
            "project": project,
            "member_cards": participant_rows,
            "department_cards": department_rows,
            "department_load": department_load,
            "task_cards": task_cards,
            "task_columns": task_columns,
            "task_stats": task_stats,
            "upcoming_tasks": upcoming_tasks,
            "priority_chart": priority_chart,
            "risk_chart": risk_chart,
            "owner_load": owner_load,
            "risk_watch": risk_watch,
            "dependency_watch": dependency_watch,
            "timeline_start": timeline_start,
            "timeline_end": timeline_end,
            "timeline_ticks": timeline_ticks,
            "today_marker_left": today_marker_left,
            "gantt_departments": gantt_departments,
            "gantt_business_days": gantt_business_days,
            "gantt_weeks": gantt_weeks,
            "gantt_today_grid_start": gantt_today_grid_start,
            "gantt_is_capped": gantt_is_capped,
            "project_timeline_end": project_timeline_end,
            "project_status_options": _PROJECT_STATUS_OPTIONS,
            "project_task_status_options": _PROJECT_TASK_STATUS_OPTIONS,
            "project_task_priority_options": _PROJECT_TASK_PRIORITY_OPTIONS,
            "project_task_risk_options": _PROJECT_TASK_RISK_OPTIONS,
            "project_department_colors": _PROJECT_DEPARTMENT_COLORS,
            "project_tab": project_tab,
            "selected_member": selected_member,
            "notice": (notice or "").strip() or None,
            "error": (error or "").strip() or None,
            "can_manage_projects": can_manage_projects,
        },
    )


@app.post("/projects")
def projects_create(
    name: str = Form(""),
    client_name: str = Form(""),
    status: str = Form("planning"),
    notes: str = Form(""),
):
    name_clean = (name or "").strip()
    client_clean = (client_name or "").strip()
    notes_clean = (notes or "").strip()
    status_clean = (status or "planning").strip().lower()

    if not name_clean:
        return _redirect("/projects?error=" + quote("Project name is required."))
    if len(name_clean) > 200:
        return _redirect("/projects?error=" + quote("Project name is too long (max 200 characters)."))
    if len(client_clean) > 200:
        return _redirect("/projects?error=" + quote("Client name is too long (max 200 characters)."))
    if len(notes_clean) > 4000:
        return _redirect("/projects?error=" + quote("Notes are too long (max 4000 characters)."))
    if status_clean not in _ALLOWED_PROJECT_STATUSES:
        return _redirect("/projects?error=" + quote("Choose a valid project status."))

    with session_scope() as s:
        project = Project(
            name=name_clean,
            client_name=client_clean or None,
            status=status_clean,
            notes=notes_clean or None,
        )
        s.add(project)
        s.flush()
        for index, (department_name, color) in enumerate(_DEFAULT_PROJECT_DEPARTMENTS, start=1):
            s.add(ProjectDepartment(project_id=project.id, name=department_name, color=color, sort_order=index * 10))
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            return _redirect("/projects?error=" + quote("A project with that name already exists."))
    return _redirect("/projects?created=1")


@app.get("/projects/{project_id}/tasks/{task_id}", response_class=HTMLResponse)
def project_task_detail_page(
    request: Request,
    project_id: int,
    task_id: int,
    notice: str | None = None,
    error: str | None = None,
):
    current_user = request.state.current_user
    can_manage_projects = bool(getattr(current_user, "can_manage_projects", False))
    with session_scope() as s:
        project = s.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        if not _user_can_access_project(s, project, current_user):
            raise HTTPException(status_code=403, detail="You do not have access to this project.")
        task = s.get(ProjectTask, task_id)
        if not task or task.project_id != project_id:
            raise HTTPException(status_code=404, detail="Task not found")

        participant_rows = (
            s.execute(
                select(ProjectParticipant)
                .where(ProjectParticipant.project_id == project_id)
                .order_by(ProjectParticipant.display_name.asc(), ProjectParticipant.id.asc())
            )
            .scalars()
            .all()
        )
        department_rows = (
            s.execute(
                select(ProjectDepartment)
                .where(ProjectDepartment.project_id == project_id)
                .order_by(ProjectDepartment.sort_order.asc(), ProjectDepartment.name.asc(), ProjectDepartment.id.asc())
            )
            .scalars()
            .all()
        )
        supporter_rows = (
            s.execute(
                select(ProjectTaskSupporter)
                .where(ProjectTaskSupporter.task_id == task_id)
                .order_by(ProjectTaskSupporter.id.asc())
            )
            .scalars()
            .all()
        )
        note_rows = (
            s.execute(
                select(ProjectTaskNote)
                .where(ProjectTaskNote.task_id == task_id)
                .order_by(ProjectTaskNote.is_pinned.desc(), ProjectTaskNote.created_at.desc(), ProjectTaskNote.id.desc())
            )
            .scalars()
            .all()
        )

        participant_name_by_id = {row.id: row.display_name for row in participant_rows}
        department_by_id = {row.id: row for row in department_rows}
        department = department_by_id.get(task.department_id) if task.department_id else None
        supporter_ids = [row.participant_id for row in supporter_rows]
        supporter_names = [participant_name_by_id.get(pid, "") for pid in supporter_ids]
        supporter_names = [name for name in supporter_names if name]
        is_done = (task.status or "") == "done"
        today = date.today()
        task_card = {
            "id": task.id,
            "title": task.title,
            "description": task.description,
            "department_id": task.department_id,
            "department_name": department.name if department else "",
            "department_color": department.color if department else "slate",
            "status": task.status,
            "status_label": _project_task_label(_PROJECT_TASK_STATUS_OPTIONS, task.status),
            "priority": task.priority,
            "priority_label": _project_task_label(_PROJECT_TASK_PRIORITY_OPTIONS, task.priority),
            "start_date": task.start_date,
            "due_date": task.due_date,
            "percent_complete": task.percent_complete or 0,
            "dependency_note": task.dependency_note,
            "risk_level": task.risk_level,
            "risk_label": _project_task_label(_PROJECT_TASK_RISK_OPTIONS, task.risk_level),
            "risk_notes": task.risk_notes,
            "owner_participant_id": task.owner_participant_id,
            "owner_name": participant_name_by_id.get(task.owner_participant_id) if task.owner_participant_id else "",
            "supporter_ids": supporter_ids,
            "supporter_names": supporter_names,
            "supporter_summary": ", ".join(supporter_names),
            "is_overdue": bool(task.due_date and task.due_date < today and not is_done),
            "created_at": task.created_at,
            "updated_at": task.updated_at,
        }
        note_type_label_by_value = dict(_PROJECT_TASK_NOTE_TYPE_OPTIONS)
        note_cards = [
            {
                "id": note.id,
                "title": note.title,
                "content": note.content,
                "plain_text": _planner_plain_text_from_content(note.content),
                "note_type": note.note_type,
                "note_type_label": note_type_label_by_value.get(note.note_type, note.note_type.replace("_", " ").title()),
                "is_pinned": note.is_pinned,
                "author_label": note.author_label,
                "created_at": note.created_at,
                "updated_at": note.updated_at,
            }
            for note in note_rows
        ]

    return templates.TemplateResponse(
        request,
        "project_task_detail.html",
        {
            "title": task_card["title"],
            "project": project,
            "task": task_card,
            "notes": note_cards,
            "member_cards": participant_rows,
            "department_cards": department_rows,
            "project_task_status_options": _PROJECT_TASK_STATUS_OPTIONS,
            "project_task_priority_options": _PROJECT_TASK_PRIORITY_OPTIONS,
            "project_task_risk_options": _PROJECT_TASK_RISK_OPTIONS,
            "project_task_note_type_options": _PROJECT_TASK_NOTE_TYPE_OPTIONS,
            "notice": (notice or "").strip() or None,
            "error": (error or "").strip() or None,
            "can_manage_projects": can_manage_projects,
        },
    )


@app.post("/projects/{project_id}/details")
def project_update_details(
    project_id: int,
    name: str = Form(""),
    client_name: str = Form(""),
    status: str = Form("planning"),
    description: str = Form(""),
):
    name_clean = (name or "").strip()
    client_clean = (client_name or "").strip()
    description_clean = (description or "").strip()
    status_clean = (status or "planning").strip().lower()

    if not name_clean:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Project name is required."))
    if len(name_clean) > 200:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Project name is too long (max 200 characters)."))
    if len(client_clean) > 200:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Client / Internal Department is too long (max 200 characters)."))
    if len(description_clean) > 4000:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Description is too long (max 4000 characters)."))
    if status_clean not in _ALLOWED_PROJECT_STATUSES:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Choose a valid project status."))

    with session_scope() as s:
        project = s.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        project.name = name_clean
        project.client_name = client_clean or None
        project.status = status_clean
        project.notes = description_clean or None
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("A project with that name already exists."))

    return _redirect(f"/projects/{project_id}?tab=manager&notice=" + quote("Project details updated."))


@app.post("/projects/{project_id}/participants")
def project_add_participant(project_id: int, display_name: str = Form("")):
    name_clean = (display_name or "").strip()
    if not name_clean:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Enter a person name to add."))
    if len(name_clean) > 200:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Person name is too long (max 200 characters)."))

    with session_scope() as s:
        project = s.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        s.add(ProjectParticipant(project_id=project_id, display_name=name_clean))
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("That person is already on this project."))

    return _redirect(f"/projects/{project_id}?tab=manager&notice=" + quote("Person added to project."))


@app.post("/projects/{project_id}/participants/{participant_id}/remove")
def project_remove_participant(project_id: int, participant_id: int):
    with session_scope() as s:
        row = s.execute(
            select(ProjectParticipant).where(
                ProjectParticipant.project_id == project_id,
                ProjectParticipant.id == participant_id,
            )
        ).scalar_one_or_none()
        if row is not None:
            s.delete(row)
            s.commit()
    return _redirect(f"/projects/{project_id}?tab=manager&notice=" + quote("Person removed from project."))


@app.post("/projects/{project_id}/departments")
def project_add_department(
    request: Request,
    project_id: int,
    name: str = Form(""),
    color: str = Form("slate"),
):
    if not bool(getattr(request.state.current_user, "can_manage_projects", False)):
        raise HTTPException(status_code=403, detail="You do not have permission to manage project departments.")
    name_clean = (name or "").strip()
    if not name_clean:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Department name is required."))
    if len(name_clean) > 160:
        return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("Department name is too long."))
    color_clean = _coerce_project_department_color(color)
    with session_scope() as s:
        project = s.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        max_order = (
            s.execute(select(func.max(ProjectDepartment.sort_order)).where(ProjectDepartment.project_id == project_id)).scalar_one_or_none()
        ) or 0
        s.add(ProjectDepartment(project_id=project_id, name=name_clean, color=color_clean, sort_order=max_order + 10))
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            return _redirect(f"/projects/{project_id}?tab=manager&error=" + quote("That department already exists."))
    return _redirect(f"/projects/{project_id}?tab=manager&notice=" + quote("Department added."))


@app.post("/projects/{project_id}/departments/{department_id}/delete")
def project_delete_department(request: Request, project_id: int, department_id: int):
    if not bool(getattr(request.state.current_user, "can_manage_projects", False)):
        raise HTTPException(status_code=403, detail="You do not have permission to manage project departments.")
    with session_scope() as s:
        department = s.get(ProjectDepartment, department_id)
        if department and department.project_id == project_id:
            for task in s.execute(select(ProjectTask).where(ProjectTask.department_id == department_id)).scalars().all():
                task.department_id = None
            s.delete(department)
            s.commit()
    return _redirect(f"/projects/{project_id}?tab=manager&notice=" + quote("Department removed."))


@app.post("/projects/{project_id}/tasks")
def project_create_task(
    request: Request,
    project_id: int,
    title: str = Form(""),
    department_id: str = Form(""),
    owner_participant_id: str = Form(""),
    supporter_participant_ids: list[str] = Form([]),
    status: str = Form("todo"),
    priority: str = Form("medium"),
    start_date: str = Form(""),
    due_date: str = Form(""),
    percent_complete: str = Form("0"),
    dependency_note: str = Form(""),
    risk_level: str = Form("low"),
    risk_notes: str = Form(""),
    description: str = Form(""),
    return_tab: str = Form("board"),
):
    if not bool(getattr(request.state.current_user, "can_manage_projects", False)):
        raise HTTPException(status_code=403, detail="You do not have permission to create project tasks.")

    title_clean = (title or "").strip()
    description_clean = (description or "").strip()
    dependency_clean = (dependency_note or "").strip()
    risk_notes_clean = (risk_notes or "").strip()
    status_clean = (status or "todo").strip().lower()
    priority_clean = (priority or "medium").strip().lower()
    risk_clean = (risk_level or "low").strip().lower()
    tab_clean = (return_tab or "board").strip().lower()
    if tab_clean == "list":
        tab_clean = "board"
    if tab_clean not in {"overview", "board", "timeline", "people", "manager"}:
        tab_clean = "board"

    if not title_clean:
        return _project_task_redirect(project_id, tab_clean, "error", "Task title is required.")
    if len(title_clean) > 300:
        return _project_task_redirect(project_id, tab_clean, "error", "Task title is too long.")
    if status_clean not in _ALLOWED_PROJECT_TASK_STATUSES:
        return _project_task_redirect(project_id, tab_clean, "error", "Choose a valid task status.")
    if priority_clean not in _ALLOWED_PROJECT_TASK_PRIORITIES:
        return _project_task_redirect(project_id, tab_clean, "error", "Choose a valid task priority.")
    if risk_clean not in _ALLOWED_PROJECT_TASK_RISKS:
        return _project_task_redirect(project_id, tab_clean, "error", "Choose a valid risk level.")

    owner_id: int | None = None
    if owner_participant_id.strip():
        try:
            owner_id = int(owner_participant_id.strip())
        except ValueError:
            return _project_task_redirect(project_id, tab_clean, "error", "Choose a valid owner.")
    department_id_value: int | None = None
    if department_id.strip():
        try:
            department_id_value = int(department_id.strip())
        except ValueError:
            return _project_task_redirect(project_id, tab_clean, "error", "Choose a valid department.")
    supporter_ids = _parse_project_id_list(supporter_participant_ids)

    start_d = _parse_optional_date_field(start_date)
    due_d = _parse_optional_date_field(due_date)
    percent = _coerce_project_task_percent(percent_complete)
    if status_clean == "done":
        percent = 100

    with session_scope() as s:
        project = s.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        if owner_id is not None:
            owner = s.get(ProjectParticipant, owner_id)
            if not owner or owner.project_id != project_id:
                return _project_task_redirect(project_id, tab_clean, "error", "Choose a valid owner.")
        if department_id_value is not None:
            department = s.get(ProjectDepartment, department_id_value)
            if not department or department.project_id != project_id:
                return _project_task_redirect(project_id, tab_clean, "error", "Choose a valid department.")
        max_order = (
            s.execute(select(func.max(ProjectTask.sort_order)).where(ProjectTask.project_id == project_id)).scalar_one_or_none()
        ) or 0
        task = ProjectTask(
            project_id=project_id,
            department_id=department_id_value,
            owner_participant_id=owner_id,
            title=title_clean,
            description=description_clean or None,
            status=status_clean,
            priority=priority_clean,
            start_date=start_d,
            due_date=due_d,
            percent_complete=percent,
            dependency_note=dependency_clean or None,
            risk_level=risk_clean,
            risk_notes=risk_notes_clean or None,
            sort_order=max_order + 1,
        )
        s.add(task)
        s.flush()
        _sync_project_task_supporters(s, task, supporter_ids)
        s.commit()

    return _project_task_redirect(project_id, tab_clean, "notice", "Task created.")


@app.post("/projects/{project_id}/tasks/{task_id}")
def project_update_task(
    request: Request,
    project_id: int,
    task_id: int,
    title: str = Form(""),
    department_id: str = Form(""),
    owner_participant_id: str = Form(""),
    supporter_participant_ids: list[str] = Form([]),
    status: str = Form("todo"),
    priority: str = Form("medium"),
    start_date: str = Form(""),
    due_date: str = Form(""),
    percent_complete: str = Form("0"),
    dependency_note: str = Form(""),
    risk_level: str = Form("low"),
    risk_notes: str = Form(""),
    description: str = Form(""),
    return_tab: str = Form("board"),
):
    if not bool(getattr(request.state.current_user, "can_manage_projects", False)):
        raise HTTPException(status_code=403, detail="You do not have permission to update project tasks.")

    title_clean = (title or "").strip()
    description_clean = (description or "").strip()
    dependency_clean = (dependency_note or "").strip()
    risk_notes_clean = (risk_notes or "").strip()
    status_clean = (status or "todo").strip().lower()
    priority_clean = (priority or "medium").strip().lower()
    risk_clean = (risk_level or "low").strip().lower()
    tab_clean = (return_tab or "board").strip().lower()
    if tab_clean == "list":
        tab_clean = "board"
    if tab_clean not in {"overview", "board", "timeline", "people", "manager", "task"}:
        tab_clean = "board"

    def task_update_redirect(message_key: str, message: str) -> RedirectResponse:
        if tab_clean == "task":
            return _task_detail_redirect(project_id, task_id, message_key, message)
        return _project_task_redirect(project_id, tab_clean, message_key, message)

    if not title_clean:
        return task_update_redirect("error", "Task title is required.")
    if status_clean not in _ALLOWED_PROJECT_TASK_STATUSES:
        return task_update_redirect("error", "Choose a valid task status.")
    if priority_clean not in _ALLOWED_PROJECT_TASK_PRIORITIES:
        return task_update_redirect("error", "Choose a valid task priority.")
    if risk_clean not in _ALLOWED_PROJECT_TASK_RISKS:
        return task_update_redirect("error", "Choose a valid risk level.")

    owner_id: int | None = None
    if owner_participant_id.strip():
        try:
            owner_id = int(owner_participant_id.strip())
        except ValueError:
            return task_update_redirect("error", "Choose a valid owner.")
    department_id_value: int | None = None
    if department_id.strip():
        try:
            department_id_value = int(department_id.strip())
        except ValueError:
            return task_update_redirect("error", "Choose a valid department.")
    supporter_ids = _parse_project_id_list(supporter_participant_ids)

    start_d = _parse_optional_date_field(start_date)
    due_d = _parse_optional_date_field(due_date)
    percent = _coerce_project_task_percent(percent_complete)
    if status_clean == "done":
        percent = 100

    with session_scope() as s:
        task = s.get(ProjectTask, task_id)
        if not task or task.project_id != project_id:
            raise HTTPException(status_code=404, detail="Task not found")
        if owner_id is not None:
            owner = s.get(ProjectParticipant, owner_id)
            if not owner or owner.project_id != project_id:
                return task_update_redirect("error", "Choose a valid owner.")
        if department_id_value is not None:
            department = s.get(ProjectDepartment, department_id_value)
            if not department or department.project_id != project_id:
                return task_update_redirect("error", "Choose a valid department.")
        task.department_id = department_id_value
        task.owner_participant_id = owner_id
        task.title = title_clean
        task.description = description_clean or None
        task.status = status_clean
        task.priority = priority_clean
        task.start_date = start_d
        task.due_date = due_d
        task.percent_complete = percent
        task.dependency_note = dependency_clean or None
        task.risk_level = risk_clean
        task.risk_notes = risk_notes_clean or None
        task.updated_at = datetime.now()
        _sync_project_task_supporters(s, task, supporter_ids)
        s.commit()

    if tab_clean == "task":
        return _task_detail_redirect(project_id, task_id, "notice", "Task updated.")
    return _project_task_redirect(project_id, tab_clean, "notice", "Task updated.")


@app.post("/projects/{project_id}/tasks/{task_id}/move")
async def project_move_task(request: Request, project_id: int, task_id: int):
    if not bool(getattr(request.state.current_user, "can_manage_projects", False)):
        raise HTTPException(status_code=403, detail="You do not have permission to move project tasks.")

    payload: dict[str, object] = {}
    content_type = (request.headers.get("content-type") or "").lower()
    if "application/json" in content_type:
        try:
            raw_payload = await request.json()
            if isinstance(raw_payload, dict):
                payload = raw_payload
        except Exception:
            payload = {}
    else:
        form = await request.form()
        payload = dict(form)

    status_clean = str(payload.get("status") or "").strip().lower()
    if status_clean not in _ALLOWED_PROJECT_TASK_STATUSES:
        return JSONResponse({"ok": False, "error": "Choose a valid task status."}, status_code=400)

    ordered_raw = payload.get("ordered_ids") or payload.get("orderedIds") or []
    if isinstance(ordered_raw, str):
        ordered_values = [value.strip() for value in ordered_raw.split(",") if value.strip()]
    elif isinstance(ordered_raw, list):
        ordered_values = ordered_raw
    else:
        ordered_values = []
    ordered_ids: list[int] = []
    for value in ordered_values:
        try:
            task_id_value = int(value)
        except (TypeError, ValueError):
            continue
        if task_id_value not in ordered_ids:
            ordered_ids.append(task_id_value)
    if task_id not in ordered_ids:
        ordered_ids.append(task_id)

    with session_scope() as s:
        target_task = s.get(ProjectTask, task_id)
        if not target_task or target_task.project_id != project_id:
            return JSONResponse({"ok": False, "error": "Task not found."}, status_code=404)
        rows = (
            s.execute(
                select(ProjectTask).where(
                    ProjectTask.project_id == project_id,
                    ProjectTask.id.in_(ordered_ids),
                )
            )
            .scalars()
            .all()
        )
        rows_by_id = {row.id: row for row in rows}
        for index, moved_id in enumerate(ordered_ids):
            row = rows_by_id.get(moved_id)
            if row is None:
                continue
            row.status = status_clean
            row.sort_order = (index + 1) * 10
            if row.id == task_id:
                row.percent_complete = _project_task_percent_for_status(status_clean, row.percent_complete)
            row.updated_at = datetime.now()
        s.commit()

    return JSONResponse({"ok": True, "status": status_clean, "ordered_ids": ordered_ids})


@app.post("/projects/{project_id}/tasks/{task_id}/delete")
def project_delete_task(request: Request, project_id: int, task_id: int, return_tab: str = Form("board")):
    if not bool(getattr(request.state.current_user, "can_manage_projects", False)):
        raise HTTPException(status_code=403, detail="You do not have permission to delete project tasks.")
    tab_clean = (return_tab or "board").strip().lower()
    if tab_clean == "list" or tab_clean == "task":
        tab_clean = "board"
    if tab_clean not in {"overview", "board", "timeline", "people", "manager"}:
        tab_clean = "board"
    with session_scope() as s:
        task = s.get(ProjectTask, task_id)
        if task and task.project_id == project_id:
            s.delete(task)
            s.commit()
    return _project_task_redirect(project_id, tab_clean, "notice", "Task deleted.")


@app.post("/projects/{project_id}/tasks/{task_id}/notes")
def project_task_add_note(
    request: Request,
    project_id: int,
    task_id: int,
    note_type: str = Form("general"),
    title: str = Form(""),
    content: str = Form(""),
    is_pinned: str | None = Form(None),
):
    note_type_clean = (note_type or "general").strip().lower()
    title_clean = (title or "").strip()
    content_clean = (content or "").strip()
    pinned = is_pinned == "1"
    if note_type_clean not in _ALLOWED_PROJECT_TASK_NOTE_TYPES:
        return _task_detail_redirect(project_id, task_id, "error", "Choose a valid note type.")
    if not title_clean and not content_clean:
        return _task_detail_redirect(project_id, task_id, "error", "Add a note title or note body.")
    if len(title_clean) > 200:
        return _task_detail_redirect(project_id, task_id, "error", "Note title is too long.")
    if len(content_clean) > 8000:
        return _task_detail_redirect(project_id, task_id, "error", "Note body is too long.")

    current_user = request.state.current_user
    author_label = (
        getattr(current_user, "display_name", None)
        or getattr(current_user, "full_name", None)
        or getattr(current_user, "username", None)
        or "Local user"
    )
    with session_scope() as s:
        project = s.get(Project, project_id)
        task = s.get(ProjectTask, task_id)
        if not project or not task or task.project_id != project_id:
            raise HTTPException(status_code=404, detail="Task not found")
        if not _user_can_access_project(s, project, current_user):
            raise HTTPException(status_code=403, detail="You do not have access to this project.")
        if pinned:
            existing_pins = (
                s.execute(select(ProjectTaskNote).where(ProjectTaskNote.task_id == task_id, ProjectTaskNote.is_pinned == True))
                .scalars()
                .all()
            )
            for note in existing_pins:
                note.is_pinned = False
        s.add(
            ProjectTaskNote(
                task_id=task_id,
                note_type=note_type_clean,
                title=title_clean or None,
                content=content_clean,
                is_pinned=pinned,
                author_label=str(author_label)[:200],
            )
        )
        task.updated_at = datetime.now()
        s.commit()
    return _task_detail_redirect(project_id, task_id, "notice", "Note added.")


@app.post("/projects/{project_id}/tasks/{task_id}/notes/{note_id}/pin")
def project_task_pin_note(request: Request, project_id: int, task_id: int, note_id: int):
    current_user = request.state.current_user
    with session_scope() as s:
        project = s.get(Project, project_id)
        task = s.get(ProjectTask, task_id)
        note = s.get(ProjectTaskNote, note_id)
        if not project or not task or task.project_id != project_id or not note or note.task_id != task_id:
            raise HTTPException(status_code=404, detail="Note not found")
        if not _user_can_access_project(s, project, current_user):
            raise HTTPException(status_code=403, detail="You do not have access to this project.")
        existing_pins = (
            s.execute(select(ProjectTaskNote).where(ProjectTaskNote.task_id == task_id, ProjectTaskNote.is_pinned == True))
            .scalars()
            .all()
        )
        for row in existing_pins:
            row.is_pinned = False
        note.is_pinned = True
        note.updated_at = datetime.now()
        task.updated_at = datetime.now()
        s.commit()
    return _task_detail_redirect(project_id, task_id, "notice", "Note pinned.")


@app.post("/projects/{project_id}/tasks/{task_id}/notes/{note_id}/delete")
def project_task_delete_note(request: Request, project_id: int, task_id: int, note_id: int):
    if not bool(getattr(request.state.current_user, "can_manage_projects", False)):
        raise HTTPException(status_code=403, detail="You do not have permission to delete task notes.")
    with session_scope() as s:
        project = s.get(Project, project_id)
        task = s.get(ProjectTask, task_id)
        note = s.get(ProjectTaskNote, note_id)
        if not project or not task or task.project_id != project_id or not note or note.task_id != task_id:
            raise HTTPException(status_code=404, detail="Note not found")
        s.delete(note)
        task.updated_at = datetime.now()
        s.commit()
    return _task_detail_redirect(project_id, task_id, "notice", "Note deleted.")


def _task_notebook_reindex_parent(s, task_id: int, parent_id: int | None, moving_entry_id: int | None = None) -> None:
    siblings = (
        s.execute(
            select(ProjectTaskNote)
            .where(
                ProjectTaskNote.task_id == task_id,
                ProjectTaskNote.parent_id == parent_id,
            )
            .order_by(ProjectTaskNote.sort_order.asc(), ProjectTaskNote.title.asc(), ProjectTaskNote.id.asc())
        )
        .scalars()
        .all()
    )
    if moving_entry_id is not None:
        siblings = [row for row in siblings if row.id != moving_entry_id]
    for index, sibling in enumerate(siblings, start=1):
        sibling.sort_order = index


def _task_notebook_move_entry(
    s,
    *,
    task_id: int,
    entry: ProjectTaskNote,
    new_parent_id: int | None,
    before_entry_id: int | None = None,
) -> str | None:
    entries = s.execute(select(ProjectTaskNote).where(ProjectTaskNote.task_id == task_id)).scalars().all()
    _, children_by_parent = _planner_load_tree(entries)
    descendant_ids = _planner_descendant_ids(entry.id, children_by_parent)

    if new_parent_id is not None:
        target_parent = next((row for row in entries if row.id == new_parent_id), None)
        if target_parent is None:
            return "The selected parent no longer exists."
        if target_parent.id == entry.id or target_parent.id in descendant_ids:
            return "You cannot move a note inside itself or one of its children."

    before_entry = None
    if before_entry_id is not None:
        before_entry = next((row for row in entries if row.id == before_entry_id), None)
        if before_entry is None:
            return "The selected drop target no longer exists."
        expected_parent_id = before_entry.parent_id
        if expected_parent_id == before_entry.id or expected_parent_id not in {row.id for row in entries}:
            expected_parent_id = None
        if before_entry.id == entry.id or before_entry.id in descendant_ids:
            return "That drop target is not valid for this note."
        if expected_parent_id != new_parent_id:
            return "That drop target is out of date. Please try again."

    old_parent_id = entry.parent_id
    destination_siblings = [
        row
        for row in (
            s.execute(
                select(ProjectTaskNote)
                .where(ProjectTaskNote.task_id == task_id, ProjectTaskNote.parent_id == new_parent_id)
                .order_by(ProjectTaskNote.sort_order.asc(), ProjectTaskNote.title.asc(), ProjectTaskNote.id.asc())
            )
            .scalars()
            .all()
        )
        if row.id != entry.id
    ]

    insert_at = len(destination_siblings)
    if before_entry is not None:
        for index, sibling in enumerate(destination_siblings):
            if sibling.id == before_entry.id:
                insert_at = index
                break

    entry.parent_id = new_parent_id
    destination_siblings.insert(insert_at, entry)
    for index, sibling in enumerate(destination_siblings, start=1):
        sibling.sort_order = index
    if old_parent_id != new_parent_id:
        _task_notebook_reindex_parent(s, task_id, old_parent_id, moving_entry_id=entry.id)
    entry.updated_at = datetime.now()
    return None


@app.get("/projects/{project_id}/tasks/{task_id}/notebook", response_class=HTMLResponse)
def project_task_notebook_legacy_redirect(request: Request, project_id: int, task_id: int):
    query = f"?{request.url.query}" if request.url.query else ""
    return _redirect(f"/projects/{project_id}/tasks/{task_id}/notes{query}")


@app.get("/projects/{project_id}/tasks/{task_id}/notes", response_class=HTMLResponse)
def project_task_notebook_page(
    request: Request,
    project_id: int,
    task_id: int,
    entry_id: int | None = None,
    saved: str | None = None,
    error: str | None = None,
):
    current_user = request.state.current_user
    with session_scope() as s:
        project = s.get(Project, project_id)
        task = s.get(ProjectTask, task_id)
        if not project or not task or task.project_id != project_id:
            raise HTTPException(status_code=404, detail="Task not found")
        if not _user_can_access_project(s, project, current_user):
            raise HTTPException(status_code=403, detail="You do not have access to this project.")

        entries = (
            s.execute(
                select(ProjectTaskNote)
                .where(ProjectTaskNote.task_id == task_id)
                .order_by(ProjectTaskNote.sort_order.asc(), ProjectTaskNote.title.asc(), ProjectTaskNote.id.asc())
            )
            .scalars()
            .all()
        )
        entry_tree, children_by_parent = _planner_load_tree(entries)
        current_entry = None
        invalid_parent_ids: set[int] = set()
        current_entry_note_role = "Note"
        current_entry_document = {"version": 1, "boxes": []}
        if entry_id:
            current_entry = s.get(ProjectTaskNote, entry_id)
            if not current_entry or current_entry.task_id != task_id:
                current_entry = None
            else:
                invalid_parent_ids = _planner_descendant_ids(current_entry.id, children_by_parent)
                invalid_parent_ids.add(current_entry.id)
                if children_by_parent.get(current_entry.id):
                    current_entry_note_role = "Parent note"
                elif current_entry.parent_id is not None:
                    current_entry_note_role = "Child note"
                current_entry_document = _planner_document_from_content(current_entry.content or "")

    return templates.TemplateResponse(
        request,
        "project_task_notebook.html",
        {
            "title": f"Notes - {task.title}",
            "project": project,
            "task": task,
            "entry_tree": entry_tree,
            "entry_options": _planner_flatten_tree(entry_tree),
            "current_entry": current_entry,
            "current_entry_note_role": current_entry_note_role,
            "current_entry_document": current_entry_document,
            "invalid_parent_ids": invalid_parent_ids,
            "saved": saved == "1",
            "error": (error or "").strip() or None,
        },
    )


@app.post("/projects/{project_id}/tasks/{task_id}/notebook")
def project_task_notebook_save(
    request: Request,
    project_id: int,
    task_id: int,
    title: str = Form(""),
    content: str = Form(""),
    entry_id: str = Form(""),
    parent_id: str = Form(""),
    is_section: str = Form(""),
    action: str = Form(""),
    move_parent_id: str = Form(""),
    move_before_id: str = Form(""),
):
    current_user = request.state.current_user
    is_autosave = request.headers.get("x-planner-autosave") == "1"

    def _parse_int(value: str) -> int | None:
        value_clean = (value or "").strip()
        if not value_clean:
            return None
        try:
            return int(value_clean)
        except ValueError:
            return None

    with session_scope() as s:
        project = s.get(Project, project_id)
        task = s.get(ProjectTask, task_id)
        if not project or not task or task.project_id != project_id:
            raise HTTPException(status_code=404, detail="Task not found")
        if not _user_can_access_project(s, project, current_user):
            raise HTTPException(status_code=403, detail="You do not have access to this project.")

        parsed_entry_id = _parse_int(entry_id)
        parsed_move_parent_id = _parse_int(move_parent_id)
        parsed_move_before_id = _parse_int(move_before_id)

        if action == "move" and parsed_entry_id is not None:
            entry = s.get(ProjectTaskNote, parsed_entry_id)
            if not entry or entry.task_id != task_id:
                return _redirect(
                    f"/projects/{project_id}/tasks/{task_id}/notes?error="
                    + quote("That note could not be moved because it no longer exists.")
                )
            move_error = _task_notebook_move_entry(
                s,
                task_id=task_id,
                entry=entry,
                new_parent_id=parsed_move_parent_id,
                before_entry_id=parsed_move_before_id,
            )
            if move_error:
                s.rollback()
                return _redirect(f"/projects/{project_id}/tasks/{task_id}/notes?entry_id={entry.id}&error=" + quote(move_error))
            task.updated_at = datetime.now()
            s.commit()
            return _redirect(f"/projects/{project_id}/tasks/{task_id}/notes?entry_id={entry.id}&saved=1")

        if action == "delete" and parsed_entry_id is not None:
            entry = s.get(ProjectTaskNote, parsed_entry_id)
            if entry and entry.task_id == task_id:
                s.delete(entry)
                task.updated_at = datetime.now()
                s.commit()
            return _redirect(f"/projects/{project_id}/tasks/{task_id}/notes?saved=1")

        title_clean = (title or "").strip()
        content_clean = content or ""
        is_section_bool = is_section.strip().lower() in {"1", "true", "yes", "on"}

        if is_autosave:
            if parsed_entry_id is None:
                return JSONResponse({"ok": False, "error": "The note could not be found."}, status_code=400)
            entry = s.get(ProjectTaskNote, parsed_entry_id)
            if not entry or entry.task_id != task_id:
                return JSONResponse({"ok": False, "error": "The note no longer exists."}, status_code=404)
            document_payload = _planner_document_payload_to_storage(content_clean)
            if document_payload is None:
                return JSONResponse({"ok": False, "error": "The note layout could not be saved."}, status_code=400)
            entry.title = title_clean or "Untitled note"
            entry.content = document_payload
            entry.updated_at = datetime.now()
            task.updated_at = datetime.now()
            s.commit()
            return JSONResponse({"ok": True, "entry_id": entry.id, "title": entry.title, "content": document_payload})

        if title_clean:
            if parsed_entry_id is not None:
                entry = s.get(ProjectTaskNote, parsed_entry_id)
                if entry and entry.task_id == task_id:
                    entry.title = title_clean
                    entry.content = content_clean
                    entry.updated_at = datetime.now()
                    task.updated_at = datetime.now()
                    s.commit()
                    return _redirect(f"/projects/{project_id}/tasks/{task_id}/notes?entry_id={entry.id}&saved=1")
            else:
                max_order = (
                    s.execute(
                        select(func.max(ProjectTaskNote.sort_order)).where(
                            ProjectTaskNote.task_id == task_id,
                            ProjectTaskNote.parent_id == None,
                        )
                    ).scalar_one_or_none()
                ) or 0
                new_entry = ProjectTaskNote(
                    task_id=task_id,
                    parent_id=None,
                    is_section=is_section_bool,
                    note_type="general",
                    title=title_clean,
                    content=content_clean,
                    sort_order=max_order + 1,
                    author_label=(
                        getattr(current_user, "display_name", None)
                        or getattr(current_user, "full_name", None)
                        or getattr(current_user, "username", None)
                        or "Local user"
                    ),
                )
                s.add(new_entry)
                task.updated_at = datetime.now()
                s.commit()
                return _redirect(f"/projects/{project_id}/tasks/{task_id}/notes?entry_id={new_entry.id}&saved=1")

    return _redirect(f"/projects/{project_id}/tasks/{task_id}/notes?saved=1")


def _planner_entry_sort_key(entry: PlannerEntry) -> tuple[int, str, int]:
    return (entry.sort_order or 0, (entry.title or "").lower(), entry.id)


def _planner_load_tree(entries: list[PlannerEntry]) -> tuple[list[dict[str, object]], dict[int | None, list[PlannerEntry]]]:
    entry_ids = {entry.id for entry in entries}
    children_by_parent: dict[int | None, list[PlannerEntry]] = defaultdict(list)
    for entry in entries:
        parent_id = entry.parent_id
        if parent_id == entry.id or parent_id not in entry_ids:
            parent_id = None
        children_by_parent[parent_id].append(entry)

    for children in children_by_parent.values():
        children.sort(key=_planner_entry_sort_key)

    def build(parent_id: int | None, depth: int) -> list[dict[str, object]]:
        nodes: list[dict[str, object]] = []
        for child in children_by_parent.get(parent_id, []):
            child_parent_id = child.parent_id
            if child_parent_id == child.id or child_parent_id not in entry_ids:
                child_parent_id = None
            nodes.append(
                {
                    "entry": child,
                    "parent_id": child_parent_id,
                    "depth": depth,
                    "children": build(child.id, depth + 1),
                }
            )
        return nodes

    return build(None, 0), children_by_parent


def _planner_flatten_tree(nodes: list[dict[str, object]]) -> list[dict[str, object]]:
    flat_nodes: list[dict[str, object]] = []
    for node in nodes:
        entry = node["entry"]
        depth = int(node["depth"])
        label_prefix = "  " * depth
        flat_nodes.append(
            {
                "id": entry.id,
                "depth": depth,
                "title": entry.title,
                "is_section": entry.is_section,
                "label": f"{label_prefix}{entry.title}",
            }
        )
        flat_nodes.extend(_planner_flatten_tree(node["children"]))
    return flat_nodes


def _planner_descendant_ids(entry_id: int, children_by_parent: dict[int | None, list[PlannerEntry]]) -> set[int]:
    descendants: set[int] = set()
    stack = [entry_id]
    while stack:
        current_id = stack.pop()
        for child in children_by_parent.get(current_id, []):
            if child.id in descendants:
                continue
            descendants.add(child.id)
            stack.append(child.id)
    return descendants


def _planner_reindex_parent(s, participant_id: int, parent_id: int | None, moving_entry_id: int | None = None) -> None:
    siblings = (
        s.execute(
            select(PlannerEntry)
            .where(
                PlannerEntry.participant_id == participant_id,
                PlannerEntry.parent_id == parent_id,
            )
            .order_by(PlannerEntry.sort_order.asc(), PlannerEntry.title.asc(), PlannerEntry.id.asc())
        )
        .scalars()
        .all()
    )
    if moving_entry_id is not None:
        siblings = [row for row in siblings if row.id != moving_entry_id]
    for index, sibling in enumerate(siblings, start=1):
        sibling.sort_order = index


def _planner_move_entry(
    s,
    *,
    participant_id: int,
    entry: PlannerEntry,
    new_parent_id: int | None,
    before_entry_id: int | None = None,
) -> str | None:
    entries = (
        s.execute(select(PlannerEntry).where(PlannerEntry.participant_id == participant_id))
        .scalars()
        .all()
    )
    _, children_by_parent = _planner_load_tree(entries)
    descendant_ids = _planner_descendant_ids(entry.id, children_by_parent)

    target_parent = None
    if new_parent_id is not None:
        target_parent = next((row for row in entries if row.id == new_parent_id), None)
        if target_parent is None:
            return "The selected parent no longer exists."
        if target_parent.id == entry.id or target_parent.id in descendant_ids:
            return "You cannot move a note inside itself or one of its children."

    before_entry = None
    if before_entry_id is not None:
        before_entry = next((row for row in entries if row.id == before_entry_id), None)
        if before_entry is None:
            return "The selected drop target no longer exists."
        expected_parent_id = before_entry.parent_id
        if expected_parent_id == before_entry.id or expected_parent_id not in {row.id for row in entries}:
            expected_parent_id = None
        if before_entry.id == entry.id or before_entry.id in descendant_ids:
            return "That drop target is not valid for this note."
        if expected_parent_id != new_parent_id:
            return "That drop target is out of date. Please try again."

    old_parent_id = entry.parent_id
    destination_siblings = [
        row
        for row in (
            s.execute(
                select(PlannerEntry)
                .where(
                    PlannerEntry.participant_id == participant_id,
                    PlannerEntry.parent_id == new_parent_id,
                )
                .order_by(PlannerEntry.sort_order.asc(), PlannerEntry.title.asc(), PlannerEntry.id.asc())
            )
            .scalars()
            .all()
        )
        if row.id != entry.id
    ]

    insert_at = len(destination_siblings)
    if before_entry is not None:
        for index, sibling in enumerate(destination_siblings):
            if sibling.id == before_entry.id:
                insert_at = index
                break

    entry.parent_id = new_parent_id
    destination_siblings.insert(insert_at, entry)
    for index, sibling in enumerate(destination_siblings, start=1):
        sibling.sort_order = index
    if old_parent_id != new_parent_id:
        _planner_reindex_parent(s, participant_id, old_parent_id, moving_entry_id=entry.id)
    entry.updated_at = datetime.now()
    return None


def _planner_coerce_int(value, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(round(float(value)))
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, parsed))


def _planner_box_text(value) -> str:
    return str(value or "").replace("\r\n", "\n").replace("\r", "\n")[:20000]


def _planner_box_color(value) -> str:
    color = str(value or "").strip()
    if len(color) == 7 and color.startswith("#"):
        hex_part = color[1:]
        if all(ch in "0123456789abcdefABCDEF" for ch in hex_part):
            return "#" + hex_part.lower()
    return "#f8fafc"


def _planner_box_font_family(value) -> str:
    font = str(value or "").strip()
    if len(font) > 80:
        font = font[:80]
    return font or "Georgia"


def _planner_box_id(value) -> str:
    raw = str(value or "").strip()
    if not raw:
        return uuid.uuid4().hex
    return raw[:80]


def _planner_normalize_box(raw_box: object) -> dict[str, object] | None:
    if not isinstance(raw_box, dict):
        return None
    return {
        "id": _planner_box_id(raw_box.get("id")),
        "x": _planner_coerce_int(raw_box.get("x"), 32, 0, 4000),
        "y": _planner_coerce_int(raw_box.get("y"), 48, 0, 4000),
        "w": _planner_coerce_int(raw_box.get("w"), 320, 180, 1600),
        "h": _planner_coerce_int(raw_box.get("h"), 180, 120, 1400),
        "text": _planner_box_text(raw_box.get("text")),
        "fontFamily": _planner_box_font_family(raw_box.get("fontFamily")),
        "fontSize": _planner_coerce_int(raw_box.get("fontSize"), 18, 12, 72),
        "fontWeight": "700" if str(raw_box.get("fontWeight", "")).strip() in {"700", "bold"} else "400",
        "fontStyle": "italic" if str(raw_box.get("fontStyle", "")).strip() == "italic" else "normal",
        "textDecoration": "underline" if str(raw_box.get("textDecoration", "")).strip() == "underline" else "none",
        "color": _planner_box_color(raw_box.get("color")),
    }


def _planner_document_from_content(raw_content: str | None) -> dict[str, object]:
    content = (raw_content or "").strip()
    if not content:
        return {"version": 1, "canvasHeight": 528, "boxes": []}
    try:
        payload = json.loads(content)
    except json.JSONDecodeError:
        payload = None

    raw_boxes = None
    if isinstance(payload, dict):
        raw_boxes = payload.get("boxes")
    elif isinstance(payload, list):
        raw_boxes = payload

    if isinstance(raw_boxes, list):
        boxes = [box for box in (_planner_normalize_box(item) for item in raw_boxes) if box is not None]
        canvas_height = 528
        if isinstance(payload, dict):
            canvas_height = _planner_coerce_int(payload.get("canvasHeight"), 528, 360, 5000)
        return {"version": 1, "canvasHeight": canvas_height, "boxes": boxes}

    fallback_text = _planner_box_text(raw_content)
    if not fallback_text.strip():
        return {"version": 1, "canvasHeight": 528, "boxes": []}
    return {
        "version": 1,
        "canvasHeight": 528,
        "boxes": [
            {
                "id": uuid.uuid4().hex,
                "x": 32,
                "y": 48,
                "w": 360,
                "h": 220,
                "text": fallback_text,
                "fontFamily": "Georgia",
                "fontSize": 18,
                "fontWeight": "400",
                "fontStyle": "normal",
                "textDecoration": "none",
                "color": "#f8fafc",
            }
        ],
    }


def _planner_document_payload_to_storage(raw_payload: str | None) -> str | None:
    try:
        payload = json.loads(raw_payload or "")
    except json.JSONDecodeError:
        return None

    raw_boxes = payload.get("boxes") if isinstance(payload, dict) else payload if isinstance(payload, list) else None
    if not isinstance(raw_boxes, list):
        return None
    boxes = [box for box in (_planner_normalize_box(item) for item in raw_boxes) if box is not None]
    canvas_height = 528
    if isinstance(payload, dict):
        canvas_height = _planner_coerce_int(payload.get("canvasHeight"), 528, 360, 5000)
    return json.dumps({"version": 1, "canvasHeight": canvas_height, "boxes": boxes}, separators=(",", ":"))


def _planner_plain_text_from_content(raw_content: str | None) -> str:
    content = (raw_content or "").strip()
    if not content:
        return ""
    try:
        payload = json.loads(content)
    except json.JSONDecodeError:
        return content
    raw_boxes = payload.get("boxes") if isinstance(payload, dict) else payload if isinstance(payload, list) else None
    if not isinstance(raw_boxes, list):
        return content
    pieces: list[str] = []
    for item in raw_boxes:
        if isinstance(item, dict):
            text_value = str(item.get("text") or "").strip()
            if text_value:
                pieces.append(text_value)
    return "\n\n".join(pieces)


@app.get("/projects/{project_id}/participants/{participant_id}/planner", response_class=HTMLResponse)
def project_participant_planner(
    request: Request,
    project_id: int,
    participant_id: int,
    entry_id: int | None = None,
    saved: str | None = None,
    error: str | None = None,
):
    current_user = request.state.current_user
    project_manager_person_id = getattr(current_user, "project_manager_person_id", None)
    user_person_id = getattr(current_user, "person_id", None)
    can_view_all_projects = bool(getattr(current_user, "can_view_all_projects", False))
    can_manage_projects = bool(getattr(current_user, "can_manage_projects", False))

    with session_scope() as s:
        project = s.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        participant = s.get(ProjectParticipant, participant_id)
        if not participant or participant.project_id != project_id:
            raise HTTPException(status_code=404, detail="Participant not found")

        # Check access - user must be the participant or a project manager
        has_access = False
        if can_view_all_projects or can_manage_projects:
            has_access = True
        elif project_manager_person_id:
            person = s.get(Person, project_manager_person_id)
            if person and person.full_name == participant.display_name:
                has_access = True
        elif user_person_id:
            person = s.get(Person, user_person_id)
            if person and person.full_name == participant.display_name:
                has_access = True

        if not has_access:
            raise HTTPException(status_code=403, detail="You do not have access to this planner.")

        entries = (
            s.execute(
                select(PlannerEntry)
                .where(
                    PlannerEntry.participant_id == participant_id,
                )
                .order_by(PlannerEntry.sort_order.asc(), PlannerEntry.title.asc())
            )
            .scalars()
            .all()
        )
        entry_tree, children_by_parent = _planner_load_tree(entries)
        entry_options = _planner_flatten_tree(entry_tree)

        # Get current entry if selected
        current_entry = None
        invalid_parent_ids: set[int] = set()
        current_entry_note_role = "Note"
        current_entry_document = {"version": 1, "boxes": []}
        if entry_id:
            current_entry = s.get(PlannerEntry, entry_id)
            if not current_entry or current_entry.participant_id != participant_id:
                current_entry = None
            else:
                invalid_parent_ids = _planner_descendant_ids(current_entry.id, children_by_parent)
                invalid_parent_ids.add(current_entry.id)
                current_has_children = bool(children_by_parent.get(current_entry.id))
                if current_has_children:
                    current_entry_note_role = "Parent note"
                elif current_entry.parent_id is not None:
                    current_entry_note_role = "Child note"
                current_entry_document = _planner_document_from_content(current_entry.content or "")

    return templates.TemplateResponse(
        request,
        "planner.html",
        {
            "title": f"Tasks and Planning - {participant.display_name}",
            "project": project,
            "participant": participant,
            "entry_tree": entry_tree,
            "entry_options": entry_options,
            "current_entry": current_entry,
            "current_entry_note_role": current_entry_note_role,
            "current_entry_document": current_entry_document,
            "invalid_parent_ids": invalid_parent_ids,
            "saved": saved == "1",
            "error": (error or "").strip() or None,
        },
    )


@app.post("/projects/{project_id}/participants/{participant_id}/planner")
def project_participant_planner_save(
    request: Request,
    project_id: int,
    participant_id: int,
    title: str = Form(""),
    content: str = Form(""),
    entry_id: str = Form(""),
    parent_id: str = Form(""),
    is_section: str = Form(""),
    action: str = Form(""),
    move_parent_id: str = Form(""),
    move_before_id: str = Form(""),
):
    current_user = request.state.current_user
    project_manager_person_id = getattr(current_user, "project_manager_person_id", None)
    user_person_id = getattr(current_user, "person_id", None)
    can_view_all_projects = bool(getattr(current_user, "can_view_all_projects", False))
    can_manage_projects = bool(getattr(current_user, "can_manage_projects", False))
    is_autosave = request.headers.get("x-planner-autosave") == "1"

    with session_scope() as s:
        project = s.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        participant = s.get(ProjectParticipant, participant_id)
        if not participant or participant.project_id != project_id:
            raise HTTPException(status_code=404, detail="Participant not found")

        # Check access
        has_access = False
        if can_view_all_projects or can_manage_projects:
            has_access = True
        elif project_manager_person_id:
            person = s.get(Person, project_manager_person_id)
            if person and person.full_name == participant.display_name:
                has_access = True
        elif user_person_id:
            person = s.get(Person, user_person_id)
            if person and person.full_name == participant.display_name:
                has_access = True

        if not has_access:
            raise HTTPException(status_code=403, detail="You do not have access to this planner.")

        def _parse_int(value: str) -> int | None:
            value_clean = (value or "").strip()
            if not value_clean:
                return None
            try:
                return int(value_clean)
            except ValueError:
                return None

        parsed_entry_id = _parse_int(entry_id)
        parsed_parent_id = _parse_int(parent_id)
        parsed_move_parent_id = _parse_int(move_parent_id)
        parsed_move_before_id = _parse_int(move_before_id)

        if action == "move" and parsed_entry_id is not None:
            entry = s.get(PlannerEntry, parsed_entry_id)
            if not entry or entry.participant_id != participant_id:
                return _redirect(
                    f"/projects/{project_id}/participants/{participant_id}/planner?error="
                    + quote("That note could not be moved because it no longer exists.")
                )
            move_error = _planner_move_entry(
                s,
                participant_id=participant_id,
                entry=entry,
                new_parent_id=parsed_move_parent_id,
                before_entry_id=parsed_move_before_id,
            )
            if move_error:
                s.rollback()
                return _redirect(
                    f"/projects/{project_id}/participants/{participant_id}/planner?entry_id={entry.id}&error="
                    + quote(move_error)
                )
            s.commit()
            return _redirect(f"/projects/{project_id}/participants/{participant_id}/planner?entry_id={entry.id}&saved=1")

        # Handle delete action
        if action == "delete" and parsed_entry_id is not None:
            entry = s.get(PlannerEntry, parsed_entry_id)
            if entry and entry.participant_id == participant_id:
                s.delete(entry)
                s.commit()
                return _redirect(f"/projects/{project_id}/participants/{participant_id}/planner?saved=1")

        title_clean = (title or "").strip()
        content_clean = content or ""
        is_section_bool = is_section.strip().lower() in {"1", "true", "yes", "on"}

        if is_autosave:
            if parsed_entry_id is None:
                return JSONResponse({"ok": False, "error": "The note could not be found."}, status_code=400)
            entry = s.get(PlannerEntry, parsed_entry_id)
            if not entry or entry.participant_id != participant_id:
                return JSONResponse({"ok": False, "error": "The note no longer exists."}, status_code=404)
            document_payload = _planner_document_payload_to_storage(content_clean)
            if document_payload is None:
                return JSONResponse({"ok": False, "error": "The note layout could not be saved."}, status_code=400)
            entry.title = title_clean or "Untitled note"
            entry.content = document_payload
            entry.updated_at = datetime.now()
            s.commit()
            return JSONResponse(
                {
                    "ok": True,
                    "entry_id": entry.id,
                    "title": entry.title,
                    "content": document_payload,
                }
            )

        if title_clean:
            if parsed_entry_id is not None:
                # Update existing entry
                entry = s.get(PlannerEntry, parsed_entry_id)
                if entry and entry.participant_id == participant_id:
                    entry.title = title_clean
                    entry.content = content_clean
                    entry.updated_at = datetime.now()
                    s.commit()
            else:
                # Create new entry
                # Get max sort order for the parent
                max_order = (
                    s.execute(
                        select(func.max(PlannerEntry.sort_order)).where(
                            PlannerEntry.participant_id == participant_id,
                            PlannerEntry.parent_id == None,
                        )
                    ).scalar_one_or_none()
                ) or 0

                new_entry = PlannerEntry(
                    participant_id=participant_id,
                    parent_id=None,
                    is_section=False,
                    title=title_clean,
                    content=content_clean,
                    sort_order=max_order + 1,
                )
                s.add(new_entry)
                s.commit()
                # Get the ID for redirect
                entry_id_val = new_entry.id

        # Redirect back to the entry or to the main planner page
        if entry_id.strip():
            return _redirect(f"/projects/{project_id}/participants/{participant_id}/planner?entry_id={entry_id.strip()}&saved=1")
        elif 'entry_id_val' in locals():
            return _redirect(f"/projects/{project_id}/participants/{participant_id}/planner?entry_id={entry_id_val}&saved=1")
        else:
            return _redirect(f"/projects/{project_id}/participants/{participant_id}/planner?saved=1")


@app.get("/people", response_class=HTMLResponse)
def people(request: Request, q: str | None = None):
    with session_scope() as s:
        stmt: Select[tuple[Person, int]] = (
            select(Person, func.count(Attendance.id).label("attendance_count"))
            .join(Attendance, Attendance.person_id == Person.id, isouter=True)
            .group_by(Person.id)
            .order_by(Person.full_name.asc())
        )
        if q:
            like = f"%{q.strip()}%"
            stmt = stmt.where((Person.full_name.ilike(like)) | (Person.email.ilike(like)))
        rows = s.execute(stmt).all()
    return templates.TemplateResponse(
        request,
        "people.html",
        {"rows": rows, "q": q or "", "talent_status_options": _TALENT_STATUS_OPTIONS},
    )


_PROFILE_TABS = frozenset(
    {"attendance", "employment", "hours", "payroll", "onboarding", "personal", "emergency"}
)


def _person_stint_has_snapshot_data(p: Person) -> bool:
    """True if there is any employment or termination data worth archiving on re-hire."""
    return bool(
        (p.employment_prior_assignment or "").strip()
        or p.employment_initial_hire_date
        or p.employment_prior_end_date
        or (p.employment_prior_job_title or "").strip()
        or p.termination_end_date
        or (p.termination_end_reason or "").strip()
        or p.termination_recorded_at
        or p.employment_recorded_at
    )


@app.get("/people/{person_id}", response_class=HTMLResponse)
def person_profile(
    request: Request,
    person_id: int,
    edit_termination: str | None = Query(None),
    edit_employment: str | None = Query(None),
    edit_personal: str | None = Query(None),
    edit_archive: int | None = Query(None),
    tab: str | None = Query(None),
):
    with session_scope() as s:
        person = s.get(Person, person_id)
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")
        attendance = (
            s.execute(select(Attendance).where(Attendance.person_id == person_id).order_by(desc(Attendance.attended_on)))
            .scalars()
            .all()
        )
        attendance_occurrences_total, attendance_occurrence_breakdown = _attendance_occurrence_totals(attendance)
        onboarding_documents = (
            s.execute(
                select(PersonOnboardingDocument)
                .where(PersonOnboardingDocument.person_id == person_id)
                .order_by(desc(PersonOnboardingDocument.uploaded_at))
            )
            .scalars()
            .all()
        )
        employment_archive_rows = (
            s.execute(
                select(EmploymentAssignmentArchive)
                .where(EmploymentAssignmentArchive.person_id == person_id)
                .order_by(EmploymentAssignmentArchive.id.asc())
            )
            .scalars()
            .all()
        )
        valid_archive_ids = {a.id for a in employment_archive_rows}
        employment_archives_display = [
            {
                "id": a.id,
                "prior_assignment": a.prior_assignment,
                "initial_hire_date": a.initial_hire_date,
                "assignment_end_date": a.assignment_end_date,
                "job_title": a.job_title,
                "termination_end_date": a.termination_end_date,
                "termination_end_reason": a.termination_end_reason,
                "termination_reason_display": _termination_reason_label(a.termination_end_reason),
                "archived_at": a.archived_at,
            }
            for a in employment_archive_rows
        ]
    today = date.today()
    employment_assignment_days = _employment_assignment_calendar_days(
        person.employment_initial_hire_date,
        person.employment_prior_end_date,
        today=today,
    )
    last_week_range_label, last_week_punches = _last_week_punch_rows(attendance)
    termination_reason_display = _termination_reason_label(person.termination_end_reason)
    can_manage_people = bool(getattr(request.state.current_user, "can_manage_people", False))
    can_view_sensitive_people = bool(getattr(request.state.current_user, "can_view_sensitive_people", False))
    can_manage_sensitive_people = bool(getattr(request.state.current_user, "can_manage_sensitive_people", False))
    termination_edit_mode = (edit_termination or "").strip().lower() in ("1", "true", "yes", "on")
    termination_details_saved = person.termination_recorded_at is not None
    show_termination_form = can_manage_people and (person.talent_status or "active") == "terminated" and (
        not termination_details_saved or termination_edit_mode
    )
    employment_edit_mode = (edit_employment or "").strip().lower() in ("1", "true", "yes", "on")
    employment_details_saved = person.employment_recorded_at is not None
    show_employment_form = can_manage_people and (not employment_details_saved or employment_edit_mode)
    personal_edit_mode = (edit_personal or "").strip().lower() in ("1", "true", "yes", "on")
    personal_details_saved = person.personal_recorded_at is not None
    show_personal_form = can_manage_sensitive_people and (not personal_details_saved or personal_edit_mode)
    personal_ssn_masked = _ssn_mask_display(person.social_security_number)
    raw_tab = (tab or "").strip().lower()
    allowed_tabs = {"attendance", "employment", "hours", "payroll"}
    if can_view_sensitive_people:
        allowed_tabs.update({"onboarding", "personal", "emergency"})
    profile_tab = raw_tab if raw_tab in allowed_tabs else "attendance"
    if employment_edit_mode:
        profile_tab = "employment"
    if termination_edit_mode:
        profile_tab = "employment"
    if personal_edit_mode and can_manage_sensitive_people:
        profile_tab = "personal"
    edit_archive_id = edit_archive if edit_archive is not None and edit_archive in valid_archive_ids else None
    if edit_archive_id is not None:
        profile_tab = "employment"
    emergency_contacts = _parse_emergency_contacts_json(getattr(person, "emergency_contacts_json", None))
    emergency_contacts_json = json.dumps(emergency_contacts).replace("<", "\\u003c")
    direct_deposit_accounts = _parse_direct_deposit_json(getattr(person, "direct_deposit_json", None))
    direct_deposit_json_script = json.dumps(direct_deposit_accounts).replace("<", "\\u003c")
    return templates.TemplateResponse(
        request,
        "person.html",
        {
            "person": person,
            "attendance": attendance,
            "employment_assignment_days": employment_assignment_days,
            "employment_assignment_today": today,
            "attendance_occurrences_total": attendance_occurrences_total,
            "attendance_occurrence_breakdown": attendance_occurrence_breakdown,
            "talent_status_options": _TALENT_STATUS_OPTIONS,
            "termination_end_reason_options": _TERMINATION_END_REASONS,
            "termination_reason_display": termination_reason_display,
            "show_termination_form": show_termination_form,
            "termination_details_saved": termination_details_saved,
            "show_employment_form": show_employment_form,
            "employment_details_saved": employment_details_saved,
            "employment_archives_display": employment_archives_display,
            "edit_archive_id": edit_archive_id,
            "last_week_range_label": last_week_range_label,
            "last_week_punches": last_week_punches,
            "profile_tab": profile_tab,
            "can_manage_people": can_manage_people,
            "can_view_sensitive_people": can_view_sensitive_people,
            "can_manage_sensitive_people": can_manage_sensitive_people,
            "show_personal_form": show_personal_form,
            "personal_details_saved": personal_details_saved,
            "personal_ssn_masked": personal_ssn_masked,
            "onboarding_documents": onboarding_documents,
            "emergency_contacts_json": emergency_contacts_json,
            "direct_deposit_accounts": direct_deposit_accounts,
            "direct_deposit_json_script": direct_deposit_json_script,
        },
    )


@app.post("/people/{person_id}/talent-status")
def person_talent_status(
    person_id: int,
    talent_status: str = Form(...),
    redirect_to: str = Form("profile"),
    return_q: str = Form(""),
):
    if talent_status not in _ALLOWED_TALENT_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid talent status")
    if redirect_to not in ("profile", "people"):
        redirect_to = "profile"
    with session_scope() as s:
        person = s.get(Person, person_id)
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")
        prev_status = person.talent_status or "active"
        person.talent_status = talent_status
        if talent_status != "terminated":
            person.termination_end_date = None
            person.termination_end_reason = None
            person.termination_recorded_at = None
        elif prev_status != "terminated":
            person.termination_end_date = None
            person.termination_end_reason = None
            person.termination_recorded_at = None
        s.commit()
    if redirect_to == "people":
        rq = (return_q or "").strip()
        if rq:
            return RedirectResponse(url=f"/people?q={quote(rq)}", status_code=303)
        return RedirectResponse(url="/people", status_code=303)
    profile_url = f"/people/{person_id}"
    if talent_status == "terminated" and prev_status != "terminated":
        profile_url = f"/people/{person_id}?tab=employment"
    return RedirectResponse(url=profile_url, status_code=303)


@app.post("/people/{person_id}/termination-details")
def person_termination_details(
    person_id: int,
    termination_end_date: str = Form(""),
    termination_end_reason: str = Form(""),
):
    raw_date = (termination_end_date or "").strip()
    end_d: date | None = None
    if raw_date:
        try:
            end_d = date.fromisoformat(raw_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end date") from None

    reason_key = (termination_end_reason or "").strip()
    if reason_key and reason_key not in _ALLOWED_TERMINATION_END_REASONS:
        raise HTTPException(status_code=400, detail="Invalid end reason")

    with session_scope() as s:
        person = s.get(Person, person_id)
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")
        if (person.talent_status or "active") != "terminated":
            raise HTTPException(status_code=400, detail="Termination details apply only when talent status is Terminated.")
        person.termination_end_date = end_d
        person.employment_prior_end_date = end_d
        person.termination_end_reason = reason_key or None
        person.termination_recorded_at = datetime.now()
        s.commit()
    return RedirectResponse(url=f"/people/{person_id}?tab=employment", status_code=303)


@app.post("/people/{person_id}/rehire")
def person_rehire(person_id: int):
    with session_scope() as s:
        person = s.get(Person, person_id)
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")
        if (person.talent_status or "active") != "terminated":
            raise HTTPException(status_code=400, detail="Re-hire is only available when talent status is Terminated.")
        if _person_stint_has_snapshot_data(person):
            end_on_job = person.employment_prior_end_date or person.termination_end_date
            s.add(
                EmploymentAssignmentArchive(
                    person_id=person.id,
                    prior_assignment=(person.employment_prior_assignment or "").strip() or None,
                    initial_hire_date=person.employment_initial_hire_date,
                    assignment_end_date=end_on_job,
                    job_title=(person.employment_prior_job_title or "").strip() or None,
                    termination_end_date=person.termination_end_date,
                    termination_end_reason=(person.termination_end_reason or "").strip() or None,
                    archived_at=datetime.now(),
                )
            )
        person.talent_status = "active"
        person.termination_end_date = None
        person.termination_end_reason = None
        person.termination_recorded_at = None
        person.employment_prior_assignment = None
        person.employment_initial_hire_date = None
        person.employment_prior_end_date = None
        person.employment_prior_job_title = None
        person.employment_recorded_at = None
        s.commit()
    return RedirectResponse(url=f"/people/{person_id}?tab=employment", status_code=303)


@app.post("/people/{person_id}/employment-archive/{archive_id}/delete")
def delete_employment_archive(person_id: int, archive_id: int):
    with session_scope() as s:
        row = s.get(EmploymentAssignmentArchive, archive_id)
        if not row or row.person_id != person_id:
            raise HTTPException(status_code=404, detail="Archive record not found.")
        s.delete(row)
        s.commit()
    return RedirectResponse(url=f"/people/{person_id}?tab=employment", status_code=303)


@app.post("/people/{person_id}/employment-archive/{archive_id}/edit")
def edit_employment_archive(
    person_id: int,
    archive_id: int,
    prior_assignment: str = Form(""),
    initial_hire_date: str = Form(""),
    assignment_end_date: str = Form(""),
    job_title: str = Form(""),
    termination_end_date: str = Form(""),
    termination_end_reason: str = Form(""),
):
    reason_key = (termination_end_reason or "").strip()
    if reason_key and reason_key not in _ALLOWED_TERMINATION_END_REASONS:
        raise HTTPException(status_code=400, detail="Invalid termination end reason.")
    with session_scope() as s:
        row = s.get(EmploymentAssignmentArchive, archive_id)
        if not row or row.person_id != person_id:
            raise HTTPException(status_code=404, detail="Archive record not found.")
        row.prior_assignment = (prior_assignment or "").strip() or None
        row.job_title = (job_title or "").strip() or None
        row.initial_hire_date = _parse_optional_date_field(initial_hire_date)
        row.assignment_end_date = _parse_optional_date_field(assignment_end_date)
        row.termination_end_date = _parse_optional_date_field(termination_end_date)
        row.termination_end_reason = reason_key or None
        s.commit()
    return RedirectResponse(url=f"/people/{person_id}?tab=employment", status_code=303)


def _parse_optional_date_field(raw: str) -> date | None:
    s = (raw or "").strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date") from None


def _normalize_ssn_optional(raw: str) -> str | None:
    s = (raw or "").strip()
    if not s:
        return None
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) != 9:
        raise HTTPException(
            status_code=400,
            detail="Social Security Number must be exactly 9 digits, or leave blank.",
        ) from None
    return f"{digits[0:3]}-{digits[3:5]}-{digits[5:9]}"


def _ssn_mask_display(stored: str | None) -> str:
    if not stored:
        return ""
    digits = "".join(c for c in stored if c.isdigit())
    if len(digits) >= 4:
        return f"***-**-{digits[-4:]}"
    return "***-**-****"


def _parse_emergency_contacts_json(raw: str | None) -> list[dict[str, str]]:
    if not raw or not str(raw).strip():
        return []
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []
    out: list[dict[str, str]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        out.append(
            {
                "name": str(item.get("name") or "").strip(),
                "phone": str(item.get("phone") or "").strip(),
                "email": str(item.get("email") or "").strip(),
            }
        )
    return out


def _normalize_emergency_contacts_from_form(raw_json: str) -> list[dict[str, str]]:
    try:
        data = json.loads(raw_json or "[]")
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail="Invalid emergency contacts data.") from e
    if not isinstance(data, list):
        raise HTTPException(status_code=400, detail="Invalid emergency contacts data.")
    if len(data) > 50:
        raise HTTPException(status_code=400, detail="Too many emergency contacts (max 50).")
    cleaned: list[dict[str, str]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        phone = str(item.get("phone") or "").strip()
        email = str(item.get("email") or "").strip()
        if len(name) > 200 or len(phone) > 80 or len(email) > 320:
            raise HTTPException(status_code=400, detail="Emergency contact field too long.")
        if not name and not phone and not email:
            continue
        cleaned.append({"name": name, "phone": phone, "email": email})
    return cleaned


_ALLOWED_DIRECT_DEPOSIT_ALLOCATION = frozenset({"percent", "fixed_amount", "remainder"})


def _parse_direct_deposit_json(raw: str | None) -> list[dict[str, object]]:
    if not raw or not str(raw).strip():
        return []
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []
    return [x for x in data if isinstance(x, dict)]


def _normalize_direct_deposit_from_form(raw_json: str) -> list[dict[str, object]]:
    try:
        data = json.loads(raw_json or "[]")
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail="Invalid direct deposit data.") from e
    if not isinstance(data, list):
        raise HTTPException(status_code=400, detail="Invalid direct deposit data.")
    if len(data) > 20:
        raise HTTPException(status_code=400, detail="Too many direct deposit accounts (max 20).")
    cleaned: list[dict[str, object]] = []
    remainder_count = 0
    for item in data:
        if not isinstance(item, dict):
            continue
        bank_name = str(item.get("bank_name") or "").strip()
        routing = "".join(c for c in str(item.get("routing_number") or "") if c.isdigit())
        account_number = str(item.get("account_number") or "").strip()
        if len(bank_name) > 200 or len(routing) > 20 or len(account_number) > 34:
            raise HTTPException(status_code=400, detail="Direct deposit field too long.")
        if not bank_name and not routing and not account_number:
            continue
        pr = item.get("priority")
        try:
            pri = int(pr)
        except (TypeError, ValueError) as e:
            raise HTTPException(status_code=400, detail="Priority must be a whole number from 1 to 99.") from e
        if pri < 1 or pri > 99:
            raise HTTPException(status_code=400, detail="Priority must be from 1 to 99.")
        at = str(item.get("allocation_type") or "percent").strip().lower()
        if at not in _ALLOWED_DIRECT_DEPOSIT_ALLOCATION:
            raise HTTPException(status_code=400, detail="Invalid allocation type.")
        amount_val: float | None
        if at == "remainder":
            remainder_count += 1
            amount_val = None
        else:
            av = item.get("amount_value")
            try:
                amount_val = float(av) if av is not None and str(av).strip() != "" else None
            except (TypeError, ValueError) as e:
                raise HTTPException(status_code=400, detail="Invalid amount value.") from e
            if amount_val is None:
                raise HTTPException(status_code=400, detail="Enter an amount for percent or fixed-dollar allocation.")
            if at == "percent":
                if amount_val < 0 or amount_val > 100:
                    raise HTTPException(status_code=400, detail="Percent must be between 0 and 100.")
            elif at == "fixed_amount" and amount_val < 0:
                raise HTTPException(status_code=400, detail="Fixed amount cannot be negative.")
        if routing and len(routing) != 9:
            raise HTTPException(status_code=400, detail="Routing number must be exactly 9 digits.")
        cleaned.append(
            {
                "priority": pri,
                "bank_name": bank_name,
                "routing_number": routing,
                "account_number": account_number,
                "allocation_type": at,
                "amount_value": amount_val,
            }
        )
    if remainder_count > 1:
        raise HTTPException(status_code=400, detail="Only one account can use “Remaining balance”.")
    cleaned.sort(key=lambda x: (int(x["priority"]), str(x["bank_name"])))
    return cleaned


def _client_upload_filename(raw: str | None) -> str:
    name = Path(raw or "").name or "document"
    return name[:280] if len(name) > 280 else name


def _safe_upload_suffix(filename: str | None) -> str:
    suf = Path(filename or "").suffix.lower()
    if not suf or len(suf) > 12 or not suf.startswith("."):
        return ""
    if not all(c.isalnum() or c == "." for c in suf):
        return ""
    return suf


async def _read_upload_limited(file: UploadFile, max_bytes: int) -> bytes:
    out = bytearray()
    while True:
        chunk = await file.read(64 * 1024)
        if not chunk:
            break
        if len(out) + len(chunk) > max_bytes:
            raise HTTPException(status_code=413, detail="File too large")
        out.extend(chunk)
    return bytes(out)


@app.post("/people/{person_id}/personal-info")
def person_personal_info(
    person_id: int,
    email: str = Form(""),
    phone: str = Form(""),
    date_of_birth: str = Form(""),
    social_security_number: str = Form(""),
    direct_deposit_json: str = Form("[]"),
):
    dob = _parse_optional_date_field(date_of_birth)
    ssn = _normalize_ssn_optional(social_security_number)
    email_clean = (email or "").strip() or None
    phone_clean = (phone or "").strip() or None
    dd_cleaned = _normalize_direct_deposit_from_form(direct_deposit_json)
    dd_store = json.dumps(dd_cleaned) if dd_cleaned else None
    with session_scope() as s:
        person = s.get(Person, person_id)
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")
        person.email = email_clean
        person.phone = phone_clean
        person.date_of_birth = dob
        person.social_security_number = ssn
        person.direct_deposit_json = dd_store
        person.personal_recorded_at = datetime.now()
        try:
            s.commit()
        except IntegrityError as e:
            s.rollback()
            raise HTTPException(
                status_code=400,
                detail="That email is already in use by another person.",
            ) from e
    return RedirectResponse(url=f"/people/{person_id}?tab=personal", status_code=303)


@app.post("/people/{person_id}/emergency-contacts")
def person_emergency_contacts(person_id: int, contacts_json: str = Form("")):
    cleaned = _normalize_emergency_contacts_from_form(contacts_json)
    with session_scope() as s:
        person = s.get(Person, person_id)
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")
        person.emergency_contacts_json = json.dumps(cleaned) if cleaned else None
        s.commit()
    return RedirectResponse(url=f"/people/{person_id}?tab=emergency", status_code=303)


@app.post("/people/{person_id}/onboarding-documents")
async def upload_person_onboarding_document(person_id: int, file: UploadFile = File(...)):
    ONBOARDING_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    raw_name = _client_upload_filename(file.filename)
    suffix = _safe_upload_suffix(file.filename)
    data = await _read_upload_limited(file, MAX_ONBOARDING_UPLOAD_BYTES)
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")
    stored = f"{uuid.uuid4().hex}{suffix}"
    path = ONBOARDING_UPLOAD_DIR / stored
    try:
        path.write_bytes(data)
    except OSError:
        raise HTTPException(status_code=500, detail="Could not store file") from None
    with session_scope() as s:
        person = s.get(Person, person_id)
        if not person:
            path.unlink(missing_ok=True)
            raise HTTPException(status_code=404, detail="Person not found")
        doc = PersonOnboardingDocument(
            person_id=person_id,
            original_filename=raw_name,
            stored_filename=stored,
            content_type=(file.content_type or "").strip() or None,
        )
        s.add(doc)
        try:
            s.commit()
        except IntegrityError:
            s.rollback()
            path.unlink(missing_ok=True)
            raise HTTPException(status_code=500, detail="Could not save upload") from None
    return RedirectResponse(url=f"/people/{person_id}?tab=onboarding", status_code=303)


@app.get("/people/{person_id}/onboarding-documents/{document_id}/download")
def download_person_onboarding_document(person_id: int, document_id: int):
    with session_scope() as s:
        doc = (
            s.execute(
                select(PersonOnboardingDocument).where(
                    PersonOnboardingDocument.id == document_id,
                    PersonOnboardingDocument.person_id == person_id,
                )
            )
            .scalars()
            .first()
        )
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        original = doc.original_filename
        stored = doc.stored_filename
        ctype = doc.content_type or "application/octet-stream"
    path = ONBOARDING_UPLOAD_DIR / stored
    if not path.is_file():
        raise HTTPException(status_code=404, detail="File missing on disk")
    return FileResponse(path, media_type=ctype, filename=original)


@app.post("/people/{person_id}/employment-history")
def person_employment_history(
    person_id: int,
    employment_prior_assignment: str = Form(""),
    employment_initial_hire_date: str = Form(""),
    employment_prior_end_date: str = Form(""),
    employment_prior_job_title: str = Form(""),
):
    hire_d = _parse_optional_date_field(employment_initial_hire_date)
    with session_scope() as s:
        person = s.get(Person, person_id)
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")
        if person.termination_recorded_at is not None:
            end_d = person.employment_prior_end_date
        else:
            end_d = _parse_optional_date_field(employment_prior_end_date)
        person.employment_prior_assignment = (employment_prior_assignment or "").strip() or None
        person.employment_initial_hire_date = hire_d
        person.employment_prior_end_date = end_d
        person.employment_prior_job_title = (employment_prior_job_title or "").strip() or None
        person.employment_recorded_at = datetime.now()
        s.commit()
    return RedirectResponse(url=f"/people/{person_id}?tab=employment", status_code=303)


@app.get("/imports", response_class=HTMLResponse)
def imports(request: Request):
    with session_scope() as s:
        batches = s.execute(select(ImportBatch).order_by(desc(ImportBatch.imported_at)).limit(200)).scalars().all()
    return templates.TemplateResponse(request, "imports.html", {"batches": batches})


_INVOICE_CROSSREF_SETTINGS_OPEN_QS = "settings=1"
_INVOICE_CROSSREF_SETTINGS_OPEN_URL = f"/invoice-crossreferencing?{_INVOICE_CROSSREF_SETTINGS_OPEN_QS}"


def _invoice_crossref_redirect_keep_settings_open() -> RedirectResponse:
    """After POSTs from the Company Selection <details>, reopen that panel on the next GET."""
    return _redirect(_INVOICE_CROSSREF_SETTINGS_OPEN_URL)


@app.get("/invoice-crossreferencing", response_class=HTMLResponse)
def invoice_crossreferencing(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1),
    crossref_error: str | None = Query(None),
    settings: str | None = Query(None),
):
    with session_scope() as s:
        paid_billed_batch = (
            s.execute(
                select(ImportBatch)
                .where(ImportBatch.note == INVOICE_NOTE_PAID_BILLED)
                .order_by(desc(ImportBatch.imported_at))
                .limit(1)
            )
            .scalars()
            .first()
        )
        mim_batch = _latest_import_batch_by_note(s, INVOICE_NOTE_MIM)
        shoals_core_batch = _latest_import_batch_by_note(s, INVOICE_NOTE_SHOALS_CORE)
        shoals_weekend_batch = _latest_import_batch_by_note(s, INVOICE_NOTE_SHOALS_WEEKEND)
        active_invoice_layout = _get_active_invoice_layout(s)
        pay_type_mappings = (
            s.execute(select(InvoicePayTypeMapping).order_by(InvoicePayTypeMapping.source_value.asc())).scalars().all()
        )
        raw_pay_types = s.execute(select(InvoiceLine.pay_type).distinct().order_by(InvoiceLine.pay_type.asc())).scalars().all()
        paid_billed_companies = (
            s.execute(
                select(InvoiceLine.company_name)
                .where(InvoiceLine.source_kind == "paid_billed", InvoiceLine.company_name.is_not(None))
                .distinct()
                .order_by(InvoiceLine.company_name.asc())
            )
            .scalars()
            .all()
        )
        selected_company = _get_paid_billed_company_filter(s)
        ppe_batch = _latest_import_batch_by_note(s, INVOICE_NOTE_PPE)
        mapped_targets = [m.target_value for m in pay_type_mappings]
        snap_row = (
            s.execute(select(InvoiceCrossrefSnapshot).order_by(desc(InvoiceCrossrefSnapshot.created_at)).limit(1))
            .scalars()
            .first()
        )
        settings_tolerance = _get_invoice_tolerance(s)

    alias_map = {_norm_pay_type_key(m.source_value): m.target_value.strip() for m in pay_type_mappings}

    comparisons: list = []
    row_count = 0
    snapshot_at: datetime | None = None
    has_snapshot = False
    crossref_total_pages = 1
    crossref_page = 1
    crossref_per_page = 25
    crossref_range_start = 0
    crossref_range_end = 0
    if snap_row:
        full = json.loads(snap_row.payload_json)
        _repair_crossref_snapshot_week_display(full)
        _shoals_default = active_invoice_layout == "shoals"
        for row in full:
            row.setdefault("shoals_hours_only", _shoals_default)
        comparisons, row_count, crossref_total_pages, crossref_page, crossref_per_page, crossref_range_start, crossref_range_end = (
            _slice_crossref_page(full, page, per_page)
        )
        snapshot_at = snap_row.created_at
        has_snapshot = True
        tolerance = Decimal(snap_row.tolerance_value)
        with session_scope() as s:
            _attach_crossref_notes(s, comparisons, alias_map)
    else:
        tolerance = settings_tolerance

    available_pay_types = sorted({*(p.strip() for p in raw_pay_types if p and p.strip()), *(t.strip() for t in mapped_targets if t and t.strip())})
    if ppe_batch:
        available_pay_types = sorted({*available_pay_types, "MIM PPE"})
    crossref_settings_open = (settings or "").strip().lower() in ("1", "true", "yes", "on")
    return templates.TemplateResponse(
        request,
        "invoice_crossreferencing.html",
        {
            "paid_billed_batch": paid_billed_batch,
            "mim_batch": mim_batch,
            "shoals_core_batch": shoals_core_batch,
            "shoals_weekend_batch": shoals_weekend_batch,
            "active_invoice_layout": active_invoice_layout,
            "pay_type_mappings": pay_type_mappings,
            "mismatch_tolerance": str(settings_tolerance),
            "paid_billed_companies": [c for c in paid_billed_companies if c and c.strip()],
            "selected_company": selected_company,
            "ppe_batch": ppe_batch,
            "available_pay_types": available_pay_types,
            "comparisons": comparisons,
            "row_count": row_count,
            "crossref_tolerance_display": str(tolerance),
            "has_snapshot": has_snapshot,
            "snapshot_at": snapshot_at,
            "crossref_page": crossref_page,
            "crossref_per_page": crossref_per_page,
            "crossref_total_pages": crossref_total_pages,
            "crossref_range_start": crossref_range_start,
            "crossref_range_end": crossref_range_end,
            "crossref_per_page_options": sorted(_CROSSREF_PER_PAGE_ALLOWED),
            "crossref_error": crossref_error,
            "crossref_settings_open": crossref_settings_open,
        },
    )


def _invoice_crossref_note_redirect_url(q: str, scroll_anchor: int) -> str:
    """Return URL after row-note POST; fragment scrolls back to the same table row (avoids jarring jump to top)."""
    u = f"/invoice-crossreferencing{q}"
    if scroll_anchor >= 0:
        u += f"#invoice-crossref-row-{scroll_anchor}"
    return u


@app.post("/invoice-crossreferencing/row-note")
def save_invoice_crossref_row_note(
    row_key: str = Form(""),
    note: str = Form(""),
    solved: str | None = Form(None),
    page: int = Form(1),
    per_page: int = Form(25),
    scroll_anchor: int = Form(0),
):
    if per_page not in _CROSSREF_PER_PAGE_ALLOWED:
        per_page = 25
    page = max(1, page)
    q = f"?page={page}&per_page={per_page}"
    anchor = max(0, scroll_anchor)
    try:
        canonical = _crossref_row_key_from_b64(row_key.strip())
    except Exception:
        return _redirect(_invoice_crossref_note_redirect_url(f"{q}&crossref_error=invalid_key", anchor))
    note_stripped = (note or "").strip()
    solved_bool = solved in ("1", "on", "true", "yes")

    if solved_bool and not note_stripped:
        return _redirect(_invoice_crossref_note_redirect_url(f"{q}&crossref_error=note_required", anchor))

    with session_scope() as s:
        row = s.get(InvoiceCrossrefRowNote, canonical)
        if not note_stripped:
            if row is not None:
                s.delete(row)
            s.commit()
            return _redirect(_invoice_crossref_note_redirect_url(q, anchor))

        solved_bool = bool(solved_bool and note_stripped)
        now = datetime.now()
        if row is None:
            s.add(InvoiceCrossrefRowNote(row_key=canonical, note=note_stripped, solved=solved_bool, updated_at=now))
        else:
            row.note = note_stripped
            row.solved = solved_bool
            row.updated_at = now
        s.commit()
    return _redirect(_invoice_crossref_note_redirect_url(q, anchor))


@app.post("/invoice-crossreferencing/layout-settings/save")
def save_invoice_layout_settings_for_workflow(target_layout: str = Form("")):
    tl = (target_layout or "").strip().lower()
    if tl not in ("default", "material_motion", "shoals"):
        raise HTTPException(status_code=400, detail="Invalid target layout.")
    with session_scope() as s:
        _save_invoice_layout_snapshot(s, tl)
        s.commit()
    return _invoice_crossref_redirect_keep_settings_open()


@app.post("/invoice-crossreferencing/paytype-mappings/add")
def add_invoice_paytype_mapping(source_value: str = Form(...), target_value: str = Form(...)):
    source_value = source_value.strip()
    target_value = target_value.strip()
    if not source_value or not target_value:
        raise HTTPException(status_code=400, detail="Both source and target pay type are required.")

    with session_scope() as s:
        existing = s.execute(
            select(InvoicePayTypeMapping).where(InvoicePayTypeMapping.source_value == source_value)
        ).scalar_one_or_none()
        if existing:
            existing.target_value = target_value
        else:
            s.add(InvoicePayTypeMapping(source_value=source_value, target_value=target_value))
        _delete_crossref_snapshot(s)
        s.commit()
    return _invoice_crossref_redirect_keep_settings_open()


@app.post("/invoice-crossreferencing/paytype-mappings/{mapping_id}/delete")
def delete_invoice_paytype_mapping(mapping_id: int):
    with session_scope() as s:
        mapping = s.get(InvoicePayTypeMapping, mapping_id)
        if mapping:
            s.delete(mapping)
        _delete_crossref_snapshot(s)
        s.commit()
    return _invoice_crossref_redirect_keep_settings_open()


@app.post("/invoice-crossreferencing/paytype-mappings/clear")
def clear_invoice_paytype_mappings():
    with session_scope() as s:
        s.execute(delete(InvoicePayTypeMapping))
        _delete_crossref_snapshot(s)
        s.commit()
    return _invoice_crossref_redirect_keep_settings_open()


@app.post("/invoice-crossreferencing/settings/tolerance")
def set_invoice_mismatch_tolerance(mismatch_tolerance: str = Form("0")):
    raw = (mismatch_tolerance or "0").strip()
    try:
        value = Decimal(raw)
    except Exception as e:
        raise HTTPException(status_code=400, detail="Tolerance must be a number.") from e
    if value < 0:
        raise HTTPException(status_code=400, detail="Tolerance cannot be negative.")

    with session_scope() as s:
        setting = s.execute(select(InvoiceSetting).where(InvoiceSetting.key == "mismatch_tolerance")).scalar_one_or_none()
        if not setting:
            setting = InvoiceSetting(key="mismatch_tolerance", value=str(value))
            s.add(setting)
        else:
            setting.value = str(value)
        _delete_crossref_snapshot(s)
        s.commit()
    return _invoice_crossref_redirect_keep_settings_open()


@app.post("/invoice-crossreferencing/settings/company-filter")
def set_paid_billed_company_filter(company_name: str = Form("")):
    value = (company_name or "").strip()
    with session_scope() as s:
        setting = s.execute(select(InvoiceSetting).where(InvoiceSetting.key == "paid_billed_company_filter")).scalar_one_or_none()
        if not setting:
            setting = InvoiceSetting(key="paid_billed_company_filter", value=value)
            s.add(setting)
        else:
            setting.value = value
        _delete_crossref_snapshot(s)
        s.commit()
    return _invoice_crossref_redirect_keep_settings_open()


@app.post("/invoice-crossreferencing/import/ppe")
async def import_invoice_ppe(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    content = await file.read()
    try:
        rows = parse_ppe_csv(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    with session_scope() as s:
        batch = ImportBatch(
            filename=file.filename,
            rows_total=len(rows),
            rows_created=len(rows),
            rows_updated=0,
            rows_skipped=0,
            note="invoice_ppe",
        )
        s.add(batch)
        s.flush()

        for r in rows:
            s.add(
                PpeDeductionLine(
                    import_batch_id=batch.id,
                    external_timesheet_id=r.external_timesheet_id,
                    employee_name=r.employee_name,
                    deduction_amount=float(r.deduction_amount),
                )
            )
        s.commit()
    _refresh_invoice_crossref_snapshot()
    return _redirect("/invoice-crossreferencing")


@app.post("/invoice-crossreferencing/calculate")
def invoice_crossref_calculate():
    _refresh_invoice_crossref_snapshot()
    return _redirect("/invoice-crossreferencing")


@app.post("/invoice-crossreferencing/clear")
def invoice_crossref_clear():
    with session_scope() as s:
        s.execute(delete(InvoiceLine))
        s.execute(delete(PpeDeductionLine))
        s.execute(delete(ImportBatch).where(ImportBatch.note.in_(list(_INVOICE_IMPORT_NOTES_ALL))))
        _delete_crossref_snapshot(s)
        _set_active_invoice_layout(s, "default")
        s.commit()
    return _redirect("/invoice-crossreferencing")


@app.post("/invoice-crossreferencing/import/paid-billed")
async def import_invoice_paid_billed(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    content = await file.read()
    try:
        rows = parse_invoice_workbook(content, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    with session_scope() as s:
        batch = ImportBatch(
            filename=file.filename,
            rows_total=len(rows),
            rows_created=len(rows),
            rows_updated=0,
            rows_skipped=0,
            note="invoice_paid_billed",
        )
        s.add(batch)
        s.flush()

        for r in rows:
            s.add(
                InvoiceLine(
                    import_batch_id=batch.id,
                    source_kind="paid_billed",
                    external_timesheet_id=r.external_timesheet_id,
                    pay_type=r.pay_type,
                    employee_name=r.employee_name,
                    company_name=r.company_name,
                    invoice_number=r.invoice_number,
                    pay_rate=float(r.pay_rate) if r.pay_rate is not None else None,
                    bill_rate=float(r.bill_rate) if r.bill_rate is not None else None,
                    quantity=float(r.quantity) if r.quantity is not None else None,
                    total_amount=float(r.total_amount) if r.total_amount is not None else None,
                    week_ended_on=r.week_ended_on,
                )
            )
        s.commit()
    _refresh_invoice_crossref_snapshot()
    return _redirect("/invoice-crossreferencing")


@app.post("/invoice-crossreferencing/import/mim")
async def import_invoice_mim(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    content = await file.read()
    try:
        rows = parse_invoice_workbook(content, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    with session_scope() as s:
        batch = ImportBatch(
            filename=file.filename,
            rows_total=len(rows),
            rows_created=len(rows),
            rows_updated=0,
            rows_skipped=0,
            note="invoice_mim",
        )
        s.add(batch)
        s.flush()

        for r in rows:
            s.add(
                InvoiceLine(
                    import_batch_id=batch.id,
                    source_kind="mim",
                    external_timesheet_id=r.external_timesheet_id,
                    pay_type=r.pay_type,
                    employee_name=r.employee_name,
                    company_name=r.company_name,
                    invoice_number=r.invoice_number,
                    pay_rate=float(r.pay_rate) if r.pay_rate is not None else None,
                    bill_rate=float(r.bill_rate) if r.bill_rate is not None else None,
                    quantity=float(r.quantity) if r.quantity is not None else None,
                    total_amount=float(r.total_amount) if r.total_amount is not None else None,
                    week_ended_on=r.week_ended_on,
                )
            )
        s.commit()
    _refresh_invoice_crossref_snapshot()
    return _redirect("/invoice-crossreferencing")


@app.post("/invoice-crossreferencing/import/shoals-core")
async def import_invoice_shoals_core(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    content = await file.read()
    try:
        rows = parse_shoals_timecard_workbook(content, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    with session_scope() as s:
        batch = ImportBatch(
            filename=file.filename,
            rows_total=len(rows),
            rows_created=len(rows),
            rows_updated=0,
            rows_skipped=0,
            note=INVOICE_NOTE_SHOALS_CORE,
        )
        s.add(batch)
        s.flush()

        for r in rows:
            s.add(
                InvoiceLine(
                    import_batch_id=batch.id,
                    source_kind="mim",
                    external_timesheet_id=r.external_timesheet_id,
                    pay_type=r.pay_type,
                    employee_name=r.employee_name,
                    company_name=r.company_name,
                    invoice_number=r.invoice_number,
                    pay_rate=float(r.pay_rate) if r.pay_rate is not None else None,
                    bill_rate=float(r.bill_rate) if r.bill_rate is not None else None,
                    quantity=float(r.quantity) if r.quantity is not None else None,
                    total_amount=float(r.total_amount) if r.total_amount is not None else None,
                    week_ended_on=r.week_ended_on,
                )
            )
        s.commit()
    _refresh_invoice_crossref_snapshot()
    return _redirect("/invoice-crossreferencing")


@app.post("/invoice-crossreferencing/import/shoals-weekend")
async def import_invoice_shoals_weekend(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    content = await file.read()
    try:
        rows = parse_shoals_timecard_workbook(content, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    with session_scope() as s:
        batch = ImportBatch(
            filename=file.filename,
            rows_total=len(rows),
            rows_created=len(rows),
            rows_updated=0,
            rows_skipped=0,
            note=INVOICE_NOTE_SHOALS_WEEKEND,
        )
        s.add(batch)
        s.flush()

        for r in rows:
            s.add(
                InvoiceLine(
                    import_batch_id=batch.id,
                    source_kind="mim",
                    external_timesheet_id=r.external_timesheet_id,
                    pay_type=r.pay_type,
                    employee_name=r.employee_name,
                    company_name=r.company_name,
                    invoice_number=r.invoice_number,
                    pay_rate=float(r.pay_rate) if r.pay_rate is not None else None,
                    bill_rate=float(r.bill_rate) if r.bill_rate is not None else None,
                    quantity=float(r.quantity) if r.quantity is not None else None,
                    total_amount=float(r.total_amount) if r.total_amount is not None else None,
                    week_ended_on=r.week_ended_on,
                )
            )
        s.commit()
    _refresh_invoice_crossref_snapshot()
    return _redirect("/invoice-crossreferencing")


@app.post("/invoice-crossreferencing/layout-switch")
def invoice_crossref_layout_switch(target_layout: str = Form("")):
    tl = (target_layout or "").strip().lower()
    if tl not in ("default", "material_motion", "shoals"):
        raise HTTPException(status_code=400, detail="Invalid target layout.")
    notes = _invoice_layout_switch_delete_notes(tl)
    with session_scope() as s:
        if notes:
            s.execute(delete(ImportBatch).where(ImportBatch.note.in_(notes)))
        _delete_crossref_snapshot(s)
        _set_active_invoice_layout(s, tl)
        _apply_invoice_layout_snapshot_if_present(s, tl)
        s.commit()
    return _redirect(_INVOICE_CROSSREF_SETTINGS_OPEN_URL)


def _stream_shoals_audit_workbook(
    *,
    comparisons: list[dict],
    paid_lines: list[InvoiceLine],
    mim_lines: list[InvoiceLine],
    alias_map: dict[str, str],
    week_dt: date | None,
    crossref_note_by_key: dict[str, str],
) -> StreamingResponse:
    """
    Excel layout aligned with the manual Shoals invoice audit: Analytics (pivot-style), Invoice (BO),
    and Paycom Timesheets (merged core + weekend lines).
    """
    for c in comparisons:
        c["shoals_hours_only"] = True

    sum_inv = Decimal("0")
    sum_pc = Decimal("0")
    sum_diff = Decimal("0")
    data_rows: list[list[object]] = []

    for c in comparisons:
        has_paid = bool(c.get("has_paid"))
        has_mim = bool(c.get("has_mim"))
        fn_pc, ln_pc = ("", "")
        if has_mim:
            fn_pc, ln_pc = _split_first_last(c.get("employee_name"))
        elif has_paid:
            fn_pc, ln_pc = _split_first_last(c.get("employee_name"))

        inv_hrs = _as_decimal(c.get("paid_qty"))
        pc_hrs = _as_decimal(c.get("mim_qty"))
        if has_paid:
            sum_inv += inv_hrs
        if has_mim:
            sum_pc += pc_hrs
        qd = _as_decimal(c.get("qty_diff"))
        sum_diff += qd

        wk_inv: date | str | None = None
        wk_iso = (str(c.get("week_ended_iso") or "")).strip()[:10]
        if wk_iso:
            try:
                wk_inv = date.fromisoformat(wk_iso)
            except ValueError:
                wk_inv = wk_iso
        if wk_inv is None and (c.get("paid_week_ended") or "").strip():
            pw = str(c.get("paid_week_ended")).strip()
            try:
                wk_inv = datetime.strptime(pw, "%m/%d/%Y").date()
            except ValueError:
                wk_inv = pw

        wk_pc: date | str | None = None
        if (c.get("mim_week_ended") or "").strip():
            mw = str(c.get("mim_week_ended")).strip()
            try:
                wk_pc = datetime.strptime(mw, "%m/%d/%Y").date()
            except ValueError:
                wk_pc = mw
        if wk_pc is None and isinstance(wk_inv, date):
            wk_pc = wk_inv
        elif wk_pc is None and wk_iso:
            try:
                wk_pc = date.fromisoformat(wk_iso)
            except ValueError:
                wk_pc = None

        ext = c.get("external_timesheet_id") or ""
        emp = c.get("employee_name") or ""
        row_key = _crossref_row_key_canonical(c, alias_map)
        comment = (crossref_note_by_key.get(row_key) or "").strip()

        data_rows.append(
            [
                _try_int_badge(ext) if ext else None,
                emp or None,
                wk_inv,
                float(inv_hrs) if has_paid else None,
                None,
                _try_int_badge(ext) if has_mim and ext else None,
                fn_pc or None,
                ln_pc or None,
                wk_pc,
                float(pc_hrs) if has_mim else None,
                None,
                float(qd),
                comment or None,
            ]
        )

    wb = Workbook()
    wb.remove(wb.active)

    ws_a = wb.create_sheet("Analytics", 0)
    _shoals_analytics_header_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_a.merge_cells("A1:D3")
    ws_a["A1"].value = "INVOICE DATA"
    ws_a["A1"].alignment = _shoals_analytics_header_center
    ws_a.merge_cells("F1:J3")
    ws_a["F1"].value = "PAYCOM DATA"
    ws_a["F1"].alignment = _shoals_analytics_header_center
    ws_a.merge_cells("L1:M3")
    ws_a["L1"].value = "Analysis & Comments"
    ws_a["L1"].alignment = _shoals_analytics_header_center
    ws_a.append(
        [
            "",
            "Subtotal",
            "",
            float(sum_inv),
            "",
            "",
            "Subtotal",
            "",
            "",
            "",
            float(sum_pc),
            "",
            float(sum_diff),
            "",
        ]
    )
    ws_a.append(
        [
            "ExternalTimesheetID",
            "EmployeeName",
            "WeekWorked",
            "Sum of BillUnit",
            "",
            "EECode",
            "Firstname",
            "Lastname",
            "Week Worked",
            "Sum of EarnHours",
            "",
            "Difference",
            "Comments",
        ]
    )
    for row in data_rows:
        ws_a.append(row)

    _format_export_short_date_column(ws_a, 3)
    _format_export_short_date_column(ws_a, 9)
    _format_export_numeric_column(ws_a, 4, EXPORT_COMMA_QTY_FMT)
    _format_export_numeric_column(ws_a, 10, EXPORT_COMMA_QTY_FMT)
    _format_export_numeric_column(ws_a, 12, EXPORT_COMMA_QTY_FMT)
    _fit_export_sheet_columns(ws_a)

    ws_inv = wb.create_sheet("Invoice", 1)
    ws_inv.append(
        [
            "BillToName",
            "ExternalTimesheetID",
            "EmployeeName",
            "InvoiceNumber",
            "PayCode",
            "BillUnit",
            "PayRate",
            "BillRate",
            "ItemBill",
            "WeekWorked",
        ]
    )
    for r in sorted(
        paid_lines,
        key=lambda x: (
            _badge_sort_key(str(x.external_timesheet_id)),
            (x.pay_type or ""),
            (x.week_ended_on.isoformat() if x.week_ended_on else ""),
        ),
    ):
        inv_cell = _invoice_cell_value(r.invoice_number)
        week_cell: date | None = r.week_ended_on or week_dt
        ws_inv.append(
            [
                r.company_name,
                _try_int_badge(r.external_timesheet_id),
                r.employee_name,
                inv_cell,
                r.pay_type,
                float(r.quantity) if r.quantity is not None else None,
                float(r.pay_rate) if r.pay_rate is not None else None,
                float(r.bill_rate) if r.bill_rate is not None else None,
                float(r.total_amount) if r.total_amount is not None else None,
                week_cell,
            ]
        )
    _format_export_short_date_column(ws_inv, 10)
    _format_export_numeric_column(ws_inv, 6, EXPORT_COMMA_QTY_FMT)
    for _money_col in (7, 8, 9):
        _format_export_numeric_column(ws_inv, _money_col, EXPORT_ACCOUNTING_USD_FMT)
    _fit_export_sheet_columns(ws_inv)

    ws_p = wb.create_sheet("Paycom Timesheets", 2)
    ws_p.append(["EECode", "Lastname", "Firstname", "EarnCode", "EarnHours", "Week Worked"])
    for r in sorted(
        mim_lines,
        key=lambda x: (
            _badge_sort_key(str(x.external_timesheet_id)),
            (x.pay_type or ""),
            x.week_ended_on.isoformat() if x.week_ended_on else "",
        ),
    ):
        fn, ln = _split_first_last(r.employee_name)
        ws_p.append(
            [
                _try_int_badge(r.external_timesheet_id),
                ln or None,
                fn or None,
                r.pay_type,
                float(r.quantity) if r.quantity is not None else None,
                r.week_ended_on or week_dt,
            ]
        )
    _format_export_short_date_column(ws_p, 6)
    _format_export_numeric_column(ws_p, 5, EXPORT_COMMA_QTY_FMT)
    _fit_export_sheet_columns(ws_p)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    we_label = week_dt.strftime("%m.%d.%y") if week_dt else ""
    fname = f"Shoals Invoice Audit WE {we_label}.xlsx" if we_label else "Shoals Invoice Audit.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _run_invoice_export(*, kind: str) -> StreamingResponse:
    """kind: material_motion (MIM + PPE workbook) or shoals (Shoals core/weekend workbook)."""
    k = (kind or "").strip().lower()
    if k not in ("material_motion", "shoals"):
        raise HTTPException(status_code=400, detail="Invalid export.")

    paid_billed_batch: ImportBatch | None = None
    mim_batch: ImportBatch | None = None
    shoals_core_batch: ImportBatch | None = None
    shoals_weekend_batch: ImportBatch | None = None
    ppe_batch: ImportBatch | None = None
    company_filter = ""
    paid_lines: list[InvoiceLine] = []
    mim_lines: list[InvoiceLine] = []
    ppe_lines: list[PpeDeductionLine] = []
    weekend_invoice_lines: list[InvoiceLine] = []
    alias_map: dict[str, str] = {}
    paid_invoice_by_key: dict[tuple, str] = {}
    snap_row: InvoiceCrossrefSnapshot | None = None
    crossref_solved_note_by_key: dict[str, str] = {}
    crossref_note_by_key: dict[str, str] = {}
    mim_batch_ids_for_cmp: list[int] = []

    with session_scope() as s:
        if k == "material_motion":
            if _get_active_invoice_layout(s) != "material_motion":
                raise HTTPException(
                    status_code=400,
                    detail="Open Company & import layout and select Material and Motion to use this export. Use the Shoals export when you are on the Shoals layout.",
                )
        else:
            if _get_active_invoice_layout(s) != "shoals":
                raise HTTPException(
                    status_code=400,
                    detail="Open Company & import layout and select Shoals to use this export. Use the Material & Motion export when you are on that layout.",
                )
        company_filter = _get_paid_billed_company_filter(s)
        mappings = s.execute(select(InvoicePayTypeMapping)).scalars().all()
        alias_map = {_norm_pay_type_key(m.source_value): m.target_value.strip() for m in mappings}
        paid_billed_batch = (
            s.execute(
                select(ImportBatch)
                .where(ImportBatch.note == INVOICE_NOTE_PAID_BILLED)
                .order_by(desc(ImportBatch.imported_at))
                .limit(1)
            )
            .scalars()
            .first()
        )
        mim_batch = _latest_import_batch_by_note(s, INVOICE_NOTE_MIM)
        shoals_core_batch = _latest_import_batch_by_note(s, INVOICE_NOTE_SHOALS_CORE)
        shoals_weekend_batch = _latest_import_batch_by_note(s, INVOICE_NOTE_SHOALS_WEEKEND)
        ppe_batch = _latest_import_batch_by_note(s, INVOICE_NOTE_PPE)
        mim_batch_ids_for_cmp = _material_motion_only_batch_ids(s) if k == "material_motion" else _shoals_only_batch_ids(s)

        if paid_billed_batch:
            rows = (
                s.execute(
                    select(InvoiceLine)
                    .where(InvoiceLine.import_batch_id == paid_billed_batch.id, InvoiceLine.source_kind == "paid_billed")
                    .order_by(InvoiceLine.id.asc())
                )
                .scalars()
                .all()
            )
            for r in rows:
                if company_filter and (r.company_name or "").strip() != company_filter:
                    continue
                if _invoice_line_total_billed_is_zero(r):
                    continue
                paid_lines.append(r)
            paid_lines = _dedupe_invoice_zero_placeholders(paid_lines, alias_map)
            paid_lines = _drop_zero_qty_when_other_pay_type_has_hours(paid_lines, alias_map)

        if mim_batch_ids_for_cmp:
            mim_lines = []
            for bid in mim_batch_ids_for_cmp:
                rows = list(
                    s.execute(
                        select(InvoiceLine)
                        .where(InvoiceLine.import_batch_id == bid, InvoiceLine.source_kind == "mim")
                        .order_by(InvoiceLine.id.asc())
                    )
                    .scalars()
                    .all()
                )
                for r in rows:
                    mim_lines.append(r)
            mim_lines = _dedupe_invoice_zero_placeholders(mim_lines, alias_map)
            mim_lines = _drop_zero_qty_when_other_pay_type_has_hours(mim_lines, alias_map)

        if k == "material_motion" and ppe_batch:
            ppe_lines = list(s.execute(select(PpeDeductionLine).where(PpeDeductionLine.import_batch_id == ppe_batch.id)).scalars().all())

        if k == "shoals" and shoals_weekend_batch:
            weekend_invoice_lines = list(
                s.execute(
                    select(InvoiceLine)
                    .where(InvoiceLine.import_batch_id == shoals_weekend_batch.id, InvoiceLine.source_kind == "mim")
                    .order_by(InvoiceLine.id.asc())
                )
                .scalars()
                .all()
            )
            weekend_invoice_lines = _dedupe_invoice_zero_placeholders(weekend_invoice_lines, alias_map)
            weekend_invoice_lines = _drop_zero_qty_when_other_pay_type_has_hours(weekend_invoice_lines, alias_map)

        for r in paid_lines:
            inv = str(r.invoice_number).strip() if r.invoice_number else ""
            if not inv:
                continue
            if k == "shoals":
                pk = (*_invoice_crossref_key(r, alias_map), _invoice_week_key(r))
            else:
                pk = _invoice_crossref_key(r, alias_map)
            if pk not in paid_invoice_by_key:
                paid_invoice_by_key[pk] = inv
            pair = _invoice_crossref_key(r, alias_map)
            if pair not in paid_invoice_by_key:
                paid_invoice_by_key[pair] = inv

        snap_row = (
            s.execute(select(InvoiceCrossrefSnapshot).order_by(desc(InvoiceCrossrefSnapshot.created_at)).limit(1))
            .scalars()
            .first()
        )

        solved_notes = (
            s.execute(
                select(InvoiceCrossrefRowNote).where(
                    InvoiceCrossrefRowNote.solved.is_(True),
                    InvoiceCrossrefRowNote.note != "",
                )
            )
            .scalars()
            .all()
        )
        crossref_solved_note_by_key = {
            n.row_key: (n.note or "").strip() for n in solved_notes if (n.note or "").strip()
        }
        all_notes = s.execute(select(InvoiceCrossrefRowNote)).scalars().all()
        crossref_note_by_key = {(n.row_key or ""): (n.note or "").strip() for n in all_notes if n.row_key}

    cr_layout = "shoals" if k == "shoals" else "material_motion"
    if snap_row:
        comparisons = json.loads(snap_row.payload_json)
        if k == "shoals":
            for c in comparisons:
                c["shoals_hours_only"] = True
        else:
            for c in comparisons:
                c.setdefault("shoals_hours_only", False)
    else:
        comparisons, _ = _build_invoice_comparison(
            paid_billed_batch.id if paid_billed_batch else None,
            mim_batch_ids_for_cmp,
            crossref_layout=cr_layout,
        )

    if k == "shoals":
        week_dt_shoals = (
            _batch_week_date(shoals_core_batch)
            or _batch_week_date(shoals_weekend_batch)
            or _batch_week_date(paid_billed_batch)
        )
        return _stream_shoals_audit_workbook(
            comparisons=comparisons,
            paid_lines=paid_lines,
            mim_lines=mim_lines,
            alias_map=alias_map,
            week_dt=week_dt_shoals,
            crossref_note_by_key=crossref_note_by_key,
        )

    wb = Workbook()
    wb.remove(wb.active)

    week_dt = _batch_week_date(mim_batch) or _batch_week_date(paid_billed_batch)
    bill_to = company_filter or "Material In Motion"
    ppe_invoice = _first_paid_invoice_number(paid_lines)
    side_label = "MIM"

    # --- Payroll (MIM detail, excludes synthetic PTO Payout tab rows) ---
    ws_pl = wb.create_sheet("Payroll", 0)
    ws_pl.append(
        [
            "Invoice No",
            "Agency Name",
            "Week End Date",
            "Badge ID",
            "First Name",
            "Last Name",
            "Type of Hour",
            "Qty",
            "PayRate",
            "BillRate",
            "Extended",
            "ShiftNo",
            "Date of Hire",
            "Notes",
        ]
    )
    payroll_noted_row_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    # Payroll rows must match merged cross-reference keys (same as UI / Crossreference sheet), not raw MIM import lines.
    company_by_key: dict[tuple[str, str], str] = {}
    for r in mim_lines:
        ik = _invoice_crossref_key(r, alias_map)
        if ik not in company_by_key and (r.company_name or "").strip():
            company_by_key[ik] = (r.company_name or "").strip()
    for r in paid_lines:
        ik = _invoice_crossref_key(r, alias_map)
        if ik not in company_by_key and (r.company_name or "").strip():
            company_by_key[ik] = (r.company_name or "").strip()

    payroll_pairs: list[tuple[int, dict]] = []
    for i, c in enumerate(comparisons):
        if not c.get("has_mim"):
            continue
        if _crossref_comparison_is_ppe_row(c, alias_map):
            continue
        if _crossref_comparison_is_pto_payout_row(c):
            continue
        payroll_pairs.append((i, c))

    def _payroll_comparison_has_solved_note(c: dict) -> bool:
        k = _crossref_row_key_canonical(c, alias_map)
        v = crossref_solved_note_by_key.get(k)
        return bool(v and str(v).strip())

    payroll_sorted = sorted(
        payroll_pairs,
        key=lambda t: (0 if _payroll_comparison_has_solved_note(t[1]) else 1, t[0]),
    )
    for _, c in payroll_sorted:
        fn, ln = _split_first_last(c.get("employee_name"))
        inv_num = _payroll_export_invoice_from_comparison(c, alias_map, paid_invoice_by_key, ppe_invoice)
        row_key = _crossref_row_key_canonical(c, alias_map)
        notes_cell = crossref_solved_note_by_key.get(row_key)
        week_end_cell = _week_from_crossref_comparison(c, week_dt)
        agency = _company_name_for_payroll_export(c, alias_map, company_by_key)
        ws_pl.append(
            [
                inv_num,
                agency or None,
                week_end_cell,
                _try_int_badge(c.get("external_timesheet_id")),
                fn,
                ln,
                c.get("pay_type"),
                _export_float_cell(c.get("mim_qty")),
                _export_float_cell(c.get("mim_pay_rate")),
                _export_float_cell(c.get("mim_bill_rate")),
                _export_float_cell(c.get("mim_total")),
                None,  # L ShiftNo
                None,  # M Date of Hire
                notes_cell,  # N Notes
            ]
        )
        if notes_cell:
            row_idx = ws_pl.max_row
            for col_idx in range(1, 15):
                ws_pl.cell(row=row_idx, column=col_idx).fill = payroll_noted_row_fill

    _format_export_short_date_column(ws_pl, 3)
    _format_export_numeric_column(ws_pl, 8, EXPORT_COMMA_QTY_FMT)
    for _money_col in (9, 10, 11):
        _format_export_numeric_column(ws_pl, _money_col, EXPORT_ACCOUNTING_USD_FMT)
    _fit_export_sheet_columns(ws_pl)

    # --- Invoice (paid/billed / BO lines) ---
    ws_inv = wb.create_sheet("Invoice", 1)
    ws_inv.append(
        [
            "BillToName",
            "ExternalTimesheetID",
            "EmployeeName",
            "InvoiceNumber",
            "PayCode",
            "BillUnit",
            "PayRate",
            "BillRate",
            "ItemBill",
            "WeekWorked",
        ]
    )
    for r in sorted(
        paid_lines,
        key=lambda x: (
            _badge_sort_key(str(x.external_timesheet_id)),
            (x.pay_type or ""),
            (x.week_ended_on.isoformat() if x.week_ended_on else ""),
        ),
    ):
        inv_cell = _invoice_cell_value(r.invoice_number)
        week_cell: date | None = r.week_ended_on or week_dt
        ws_inv.append(
            [
                r.company_name,
                _try_int_badge(r.external_timesheet_id),
                r.employee_name,
                inv_cell,
                r.pay_type,
                float(r.quantity) if r.quantity is not None else None,
                float(r.pay_rate) if r.pay_rate is not None else None,
                float(r.bill_rate) if r.bill_rate is not None else None,
                float(r.total_amount) if r.total_amount is not None else None,
                week_cell,
            ]
        )

    _format_export_short_date_column(ws_inv, 10)
    _format_export_numeric_column(ws_inv, 6, EXPORT_COMMA_QTY_FMT)
    for _money_col in (7, 8, 9):
        _format_export_numeric_column(ws_inv, _money_col, EXPORT_ACCOUNTING_USD_FMT)
    _fit_export_sheet_columns(ws_inv)

    # --- PPE Reimbursement ---
    ws_ppe = wb.create_sheet("PPE Reimbursement", 2)
    ws_ppe.append(
        [
            "BillToName",
            "ExternalTimesheetID",
            "EmployeeName",
            "InvoiceNumber",
            "PayCode",
            "BillUnit",
            "PayRate",
            "BillRate",
            "ItemBill",
            "WeekWorked",
        ]
    )
    ppe_inv_cell = _invoice_cell_value(ppe_invoice)
    for p in sorted(ppe_lines, key=lambda x: _badge_sort_key(str(x.external_timesheet_id))):
        amt = abs(float(p.deduction_amount))
        ws_ppe.append(
            [
                bill_to,
                _try_int_badge(p.external_timesheet_id),
                p.employee_name,
                ppe_inv_cell,
                "Equipment Fee",
                -1,
                amt,
                amt,
                -amt,
                week_dt,
            ]
        )

    _format_export_short_date_column(ws_ppe, 10)
    _format_export_numeric_column(ws_ppe, 6, EXPORT_COMMA_QTY_FMT)
    for _money_col in (7, 8, 9):
        _format_export_numeric_column(ws_ppe, _money_col, EXPORT_ACCOUNTING_USD_FMT)
    _fit_export_sheet_columns(ws_ppe)

    # --- Rates ---
    ws_rates = wb.create_sheet("Rates", 3)
    ws_rates.append(["EMP_EmpID", "EMP_LongName", "EMP_Agency", "PayRate"])
    for r in _rates_from_mim(mim_lines):
        nm = (r.employee_name or "").strip()
        ws_rates.append(
            [
                _try_int_badge(r.external_timesheet_id),
                nm.upper().replace("  ", " ") if nm else None,
                r.company_name,
                float(r.pay_rate) if r.pay_rate is not None else None,
            ]
        )

    _format_export_numeric_column(ws_rates, 4, EXPORT_ACCOUNTING_USD_FMT)
    _fit_export_sheet_columns(ws_rates)

    # --- PTO Payout ---
    ws_pto = wb.create_sheet("PTO Payout", 4)
    ws_pto.append(
        [
            "Invoice No",
            "Agency Name",
            "Week End Date",
            "Badge ID",
            "First Name",
            "Last Name",
            "Type of Hour",
            "Qty",
            "PayRate",
            "BillRate",
            "Extended",
        ]
    )
    for r in _mim_pto_payout_lines(mim_lines):
        fn, ln = _split_first_last(r.employee_name)
        inv_num = _payroll_export_invoice_no(r, alias_map, paid_invoice_by_key, ppe_invoice)
        pto_week = r.week_ended_on or week_dt
        ws_pto.append(
            [
                inv_num,
                r.company_name,
                pto_week,
                _try_int_badge(r.external_timesheet_id),
                fn,
                ln,
                "PTO",
                float(r.quantity) if r.quantity is not None else None,
                float(r.pay_rate) if r.pay_rate is not None else None,
                float(r.bill_rate) if r.bill_rate is not None else None,
                float(r.total_amount) if r.total_amount is not None else None,
            ]
        )

    _format_export_short_date_column(ws_pto, 3)
    _format_export_numeric_column(ws_pto, 8, EXPORT_COMMA_QTY_FMT)
    for _money_col in (9, 10, 11):
        _format_export_numeric_column(ws_pto, _money_col, EXPORT_ACCOUNTING_USD_FMT)
    _fit_export_sheet_columns(ws_pto)

    # --- Crossreference (diff / reconciliation; mismatch highlighting) ---
    ws_x = wb.create_sheet("Crossreference", 5)
    ws_x.append(
        [
            "Invoice Number",
            "External Timesheet ID (VMS ID)",
            "Pay Type",
            "Week End (Paid/Billed)",
            f"Week End ({side_label})",
            "Employee Name",
            "Paid Pay Rate",
            f"{side_label} Pay Rate",
            "Pay Rate Diff",
            "Paid Bill Rate",
            f"{side_label} Bill Rate",
            "Bill Rate Diff",
            "Billed Hours (Paid/Billed)",
            f"QTY ({side_label})",
            "Qty Diff",
            "Paid Total Amount",
            f"{side_label} Total Amount (incl PPE)" if k == "material_motion" else f"{side_label} Total Amount",
            "Total Amount Diff",
            "Missing In Paid/Billed?",
            f"Missing In {side_label}?",
        ]
    )
    for c in comparisons:
        pwk = (c.get("paid_week_ended") or "").strip()
        mwk = (c.get("mim_week_ended") or "").strip()
        if not pwk and not mwk:
            leg = (c.get("week_ended") or "").strip()
            if leg:
                pwk = leg
        ws_x.append(
            [
                c["invoice_number"],
                c["external_timesheet_id"],
                c["pay_type"],
                pwk or None,
                mwk or None,
                c["employee_name"],
                float(c["paid_pay_rate"]),
                float(c["mim_pay_rate"]),
                float(c["pay_rate_diff"]),
                float(c["paid_bill_rate"]),
                float(c["mim_bill_rate"]),
                float(c["bill_rate_diff"]),
                float(c["paid_qty"]),
                float(c["mim_qty"]),
                float(c["qty_diff"]),
                float(c["paid_total"]),
                float(c["mim_total"]),
                float(c["total_diff"]),
                "YES" if not c["has_paid"] else "",
                "YES" if not c["has_mim"] else "",
            ]
        )
        if c["any_mismatch"]:
            red_fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
            row_idx = ws_x.max_row
            for col_idx in range(1, 21):
                ws_x.cell(row=row_idx, column=col_idx).fill = red_fill
        elif c.get("mim_no_paid_zero_mim_qty"):
            yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
            row_idx = ws_x.max_row
            for col_idx in range(1, 21):
                ws_x.cell(row=row_idx, column=col_idx).fill = yellow_fill

    for _money_col in (7, 8, 9, 10, 11, 12, 16, 17, 18):
        _format_export_numeric_column(ws_x, _money_col, EXPORT_ACCOUNTING_USD_FMT)
    for _qty_col in (13, 14, 15):
        _format_export_numeric_column(ws_x, _qty_col, EXPORT_COMMA_QTY_FMT)
    _fit_export_sheet_columns(ws_x)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    we_label = ""
    if week_dt:
        we_label = week_dt.strftime("%m.%d.%y")
    if k == "material_motion":
        fname = f"MIM Invoice Report WE {we_label}.xlsx" if we_label else "MIM Invoice Report.xlsx"
    else:
        fname = f"Shoals Invoice Report WE {we_label}.xlsx" if we_label else "Shoals Invoice Report.xlsx"

    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/invoice-crossreferencing/export")
def export_invoice_crossreference():
    return _run_invoice_export(kind="material_motion")


@app.get("/invoice-crossreferencing/export/shoals")
def export_invoice_crossreference_shoals():
    return _run_invoice_export(kind="shoals")


@app.post("/imports/attendance")
async def import_attendance(file: UploadFile = File(...), default_status: str = Form("present")):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    content = await file.read()
    try:
        rows = parse_attendance_csv(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    default_status = (default_status or "present").strip().lower()

    created = 0
    updated = 0
    skipped = 0

    with session_scope() as s:
        batch = ImportBatch(filename=file.filename, rows_total=len(rows))
        s.add(batch)
        s.flush()

        for r in rows:
            status = (r.status or default_status).strip().lower() or "present"

            person: Person | None = None
            if r.email:
                person = s.execute(select(Person).where(Person.email == r.email)).scalar_one_or_none()
            if not person:
                person = s.execute(select(Person).where(Person.full_name == r.full_name, Person.email.is_(None))).scalar_one_or_none()

            if not person:
                person = Person(full_name=r.full_name, email=r.email)
                s.add(person)
                s.flush()

            existing = s.execute(
                select(Attendance).where(Attendance.person_id == person.id, Attendance.attended_on == r.attended_on)
            ).scalar_one_or_none()

            if not existing:
                s.add(
                    Attendance(
                        person_id=person.id,
                        attended_on=r.attended_on,
                        status=status,
                        source_filename=file.filename,
                    )
                )
                created += 1
            else:
                if existing.status != status or existing.source_filename != file.filename:
                    existing.status = status
                    existing.source_filename = file.filename
                    updated += 1
                else:
                    skipped += 1

        batch.rows_created = created
        batch.rows_updated = updated
        batch.rows_skipped = skipped

        try:
            s.commit()
        except IntegrityError as e:
            s.rollback()
            raise HTTPException(status_code=400, detail="Import failed due to a data conflict.") from e

    return _redirect("/imports")


@app.post("/imports/roster")
async def import_roster(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    content = await file.read()
    try:
        rows = parse_roster_csv(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    created = 0
    updated = 0
    skipped = 0

    with session_scope() as s:
        batch = ImportBatch(filename=file.filename, rows_total=len(rows), note="roster")
        s.add(batch)
        s.flush()

        for r in rows:
            email = (r.email or "").strip() or None
            full_name = r.full_name.strip()

            person: Person | None = None
            if email:
                person = s.execute(select(Person).where(Person.email == email)).scalar_one_or_none()
            if not person:
                person = (
                    s.execute(select(Person).where(Person.full_name == full_name, Person.email.is_(None)))
                    .scalar_one_or_none()
                )

            if not person:
                person = Person(
                    full_name=full_name,
                    email=email,
                    phone=r.phone,
                    roster_status=r.roster_status,
                    external_id=r.external_id,
                )
                s.add(person)
                created += 1
                continue

            changed = False
            if not person.email and email:
                person.email = email
                changed = True
            if r.phone and person.phone != r.phone:
                person.phone = r.phone
                changed = True
            if r.roster_status and person.roster_status != r.roster_status:
                person.roster_status = r.roster_status
                changed = True
            if r.external_id and person.external_id != r.external_id:
                person.external_id = r.external_id
                changed = True
            if person.full_name != full_name:
                person.full_name = full_name
                changed = True

            if changed:
                updated += 1
            else:
                skipped += 1

        batch.rows_created = created
        batch.rows_updated = updated
        batch.rows_skipped = skipped

        try:
            s.commit()
        except IntegrityError as e:
            s.rollback()
            raise HTTPException(status_code=400, detail="Roster import failed due to a data conflict.") from e

    return _redirect("/people")


@app.post("/people/add")
def people_add(
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    social_security_number: str = Form(""),
    date_of_birth: str = Form(""),
):
    first = (first_name or "").strip()
    last = (last_name or "").strip()
    if not first or not last:
        raise HTTPException(status_code=400, detail="First and last name are required.")
    full_name = f"{first} {last}".strip()
    if len(full_name) > 200:
        raise HTTPException(status_code=400, detail="Combined name is too long (max 200 characters).")

    email_clean = (email or "").strip()
    phone_clean = (phone or "").strip()
    if not email_clean:
        raise HTTPException(status_code=400, detail="Email is required.")
    if not phone_clean:
        raise HTTPException(status_code=400, detail="Phone number is required.")
    if len(phone_clean) > 40:
        raise HTTPException(status_code=400, detail="Phone number is too long.")

    dob = _parse_optional_date_field(date_of_birth)
    ssn = _normalize_ssn_optional(social_security_number)

    with session_scope() as s:
        person = Person(full_name=full_name, email=email_clean, phone=phone_clean)
        person.date_of_birth = dob
        person.social_security_number = ssn
        s.add(person)
        try:
            s.flush()
            new_id = person.id
            s.commit()
        except IntegrityError as e:
            s.rollback()
            raise HTTPException(status_code=400, detail="That email already exists.") from e

    return RedirectResponse(url=f"/people/{new_id}", status_code=303)
